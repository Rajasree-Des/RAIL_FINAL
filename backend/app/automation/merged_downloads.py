"""Build single merged PDF and Excel workbook from run artifacts."""

from __future__ import annotations

import logging
import re
from copy import copy
from datetime import UTC, datetime, timedelta
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pypdf import PageObject, PdfReader, PdfWriter
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from app.automation.merged_report_catalog import (
    MergedReportEntry,
    consolidated_download_catalog,
)
from app.automation.report_keys import canonicalize_report_key
from app.automation.run_registry import validate_artifact_file
from app.infrastructure.database.models import AutomationArtifactModel

logger = logging.getLogger(__name__)


class MergedDownloadError(Exception):
    """Raised when merged output cannot be built."""


# Content signals — real report pages usually include at least one of these.
_REPORT_CONTENT_MARKERS = (
    "received",
    "division",
    "s.no",
    "sno",
    "total",
    "% share",
    "zone",
    "train",
    "station",
    "complaint",
    "watering",
    "vande bharat",
    "comprehensive",
    "feedback",
    "opening balance",
    "closing balance",
)

_REPORT_NUM_LINE = re.compile(r"^Report\s+(\d+[\u2013\-–]?\d*|\d+)\s*$", re.IGNORECASE)
_DASH_LINE = re.compile(r"^-{3,}$")
_REPORT_TITLE_ONLY_NUM = re.compile(r"^Report\s+\d+", re.IGNORECASE)


def merged_report_filename_date(*, now: datetime | None = None) -> str:
    moment = now or datetime.now(UTC)
    if moment.tzinfo is None:
        moment = moment.replace(tzinfo=UTC)
    report_day = moment - timedelta(days=1)
    return report_day.strftime("%Y-%m-%d")


def cover_report_date_display(*, generated_at: datetime) -> str:
    """DD-MM-YYYY date shown on the merged PDF cover (report-as-of date)."""
    moment = generated_at
    if moment.tzinfo is None:
        moment = moment.replace(tzinfo=UTC)
    report_day = moment - timedelta(days=1)
    return report_day.strftime("%d-%m-%Y")


def merged_pdf_filename(*, now: datetime | None = None) -> str:
    return f"RailMadad_Report_{merged_report_filename_date(now=now)}.pdf"


def merged_excel_filename(*, now: datetime | None = None) -> str:
    """CRB_RM_Reports_DD-MM-YYYY.xlsx — same report-as-of date as the merged PDF cover."""
    moment = now or datetime.now(UTC)
    return f"CRB_RM_Reports_{cover_report_date_display(generated_at=moment)}.xlsx"


def _index_artifacts(
    artifacts: list[AutomationArtifactModel],
) -> dict[str, dict[str, AutomationArtifactModel]]:
    """Map slug → {pdf|excel → newest ready artifact}."""
    index: dict[str, dict[str, AutomationArtifactModel]] = {}
    for artifact in artifacts:
        if (artifact.status or "ready") != "ready":
            continue
        if artifact.artifact_type not in {"pdf", "excel"}:
            continue
        slug = canonicalize_report_key(artifact.report_slug or "")
        if not slug:
            continue
        bucket = index.setdefault(slug, {})
        if artifact.artifact_type not in bucket:
            bucket[artifact.artifact_type] = artifact
    return index


def _validated_path(artifact: AutomationArtifactModel) -> Path | None:
    try:
        return validate_artifact_file(
            Path(artifact.file_path),
            require_pdf_header=artifact.artifact_type == "pdf",
            file_type=artifact.artifact_type,
        )
    except Exception:
        return None


def _ordered_artifact_paths(
    artifacts: list[AutomationArtifactModel],
    *,
    file_type: str,
) -> list[tuple[MergedReportEntry, Path]]:
    index = _index_artifacts(artifacts)
    ordered: list[tuple[MergedReportEntry, Path]] = []
    for entry in consolidated_download_catalog():
        art = index.get(entry.slug, {}).get(file_type)
        if art is None:
            continue
        path = _validated_path(art)
        if path is None:
            continue
        ordered.append((entry, path))
    return ordered


def _build_cover_pdf(*, generated_at: datetime) -> bytes:
    buffer = BytesIO()
    width, height = A4
    pdf = canvas.Canvas(buffer, pagesize=A4)
    title = f"CRB RM Reports as on {cover_report_date_display(generated_at=generated_at)}"
    pdf.setFont("Helvetica-Bold", 22)
    pdf.drawCentredString(width / 2, height / 2, title)
    pdf.showPage()
    pdf.save()
    return buffer.getvalue()


def _build_toc_pdf(entries: list[MergedReportEntry]) -> bytes:
    buffer = BytesIO()
    width, height = A4
    pdf = canvas.Canvas(buffer, pagesize=A4)

    line_height = 22
    title_gap = 38
    title_block = 20
    content_height = title_block + title_gap + max(len(entries), 1) * line_height
    top_y = (height + content_height) / 2

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawCentredString(width / 2, top_y, "TABLE OF CONTENTS")
    pdf.setFont("Helvetica", 12)
    y = top_y - title_gap
    for entry in entries:
        pdf.drawCentredString(width / 2, y, entry.toc_line)
        y -= line_height
        if y < 72:
            pdf.showPage()
            pdf.setFont("Helvetica", 12)
            y = height - 72

    pdf.showPage()
    pdf.save()
    return buffer.getvalue()


def _page_text(page: PageObject) -> str:
    try:
        return (page.extract_text() or "").strip()
    except Exception:
        return ""


def is_separator_title_page(page: PageObject) -> bool:
    """True when a PDF page is a standalone report title/divider (no real content).

    Detects pages that contain only:
      Report <number>
      <report name>
      -----------
    and lack table/report body markers. Future-proof — not tied to page indices.
    """
    text = _page_text(page)
    if not text:
        return True

    compact = re.sub(r"\s+", " ", text).strip()
    lower = compact.lower()
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    if len(lines) > 8:
        return False

    if any(marker in lower for marker in _REPORT_CONTENT_MARKERS):
        # Allow short title + one marker only when still looks like a divider blob.
        if len(lines) > 4:
            return False

    has_report_heading = any(
        _REPORT_NUM_LINE.match(ln) or _REPORT_TITLE_ONLY_NUM.match(ln) for ln in lines
    )
    if not has_report_heading and "report vande bharat" not in lower:
        return False

    has_dashes = any(_DASH_LINE.match(ln.replace(" ", "")) for ln in lines) or "----" in compact

    # Title-only pages: few lines, report heading, optional dash rule, no dense data.
    numeric_tokens = len(re.findall(r"\b\d+(?:\.\d+)?\b", compact))
    if has_report_heading and len(lines) <= 5 and numeric_tokens <= 1:
        return True
    if has_report_heading and has_dashes and len(lines) <= 6 and numeric_tokens <= 2:
        return True

    return False


def _append_pdf_bytes(writer: PdfWriter, payload: bytes, *, strip_separators: bool = False) -> None:
    reader = PdfReader(BytesIO(payload))
    for page in reader.pages:
        if strip_separators and is_separator_title_page(page):
            continue
        writer.add_page(page)


def _append_pdf_file(writer: PdfWriter, path: Path, *, strip_separators: bool = True) -> None:
    reader = PdfReader(str(path))
    kept = 0
    for page in reader.pages:
        if strip_separators and is_separator_title_page(page):
            logger.debug("Skipping separator title page in %s", path.name)
            continue
        writer.add_page(page)
        kept += 1
    if kept == 0:
        logger.warning("All pages stripped from %s; appending raw PDF", path.name)
        for page in reader.pages:
            writer.add_page(page)


def build_merged_pdf(
    artifacts: list[AutomationArtifactModel],
    *,
    run_id: str,
    generated_at: datetime | None = None,
) -> bytes:
    """Merge cover, TOC, and report PDFs (vector-preserving, no divider pages)."""
    _ = run_id  # retained for API compatibility; not shown on cover
    when = generated_at or datetime.now(UTC)
    ordered = _ordered_artifact_paths(artifacts, file_type="pdf")
    if not ordered:
        raise MergedDownloadError("No PDF artifacts available for merge")

    entries = [entry for entry, _ in ordered]
    writer = PdfWriter()
    _append_pdf_bytes(writer, _build_cover_pdf(generated_at=when))
    _append_pdf_bytes(writer, _build_toc_pdf(entries))

    for _entry, path in ordered:
        _append_pdf_file(writer, path, strip_separators=True)

    output = BytesIO()
    writer.write(output)
    return output.getvalue()


def _copy_cell_style(source_cell, target_cell) -> None:
    if not source_cell.has_style:
        return
    target_cell.font = copy(source_cell.font)
    target_cell.border = copy(source_cell.border)
    target_cell.fill = copy(source_cell.fill)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)
    target_cell.alignment = copy(source_cell.alignment)


def _copy_worksheet(source: Worksheet, target: Worksheet) -> None:
    for col_letter, dim in source.column_dimensions.items():
        target.column_dimensions[col_letter].width = dim.width
        target.column_dimensions[col_letter].hidden = dim.hidden
        target.column_dimensions[col_letter].bestFit = dim.bestFit
        # customWidth is a read-only computed property in openpyxl >= 3.1
        # (derived from width being set) — it has no setter.

    for row_idx, dim in source.row_dimensions.items():
        target.row_dimensions[row_idx].height = dim.height
        target.row_dimensions[row_idx].hidden = dim.hidden

    for row in source.iter_rows():
        for cell in row:
            target_cell = target.cell(row=cell.row, column=cell.column, value=cell.value)
            _copy_cell_style(cell, target_cell)

    for merged_range in source.merged_cells.ranges:
        target.merge_cells(str(merged_range))

    target.freeze_panes = source.freeze_panes
    if source.auto_filter and source.auto_filter.ref:
        target.auto_filter.ref = source.auto_filter.ref

    _autofit_columns(target)


def _autofit_columns(worksheet: Worksheet) -> None:
    """Widen columns to fit content. Merged-cell members have no letter/value; skip them."""
    for column_cells in worksheet.columns:
        letter = None
        max_length = 0
        for cell in column_cells:
            if isinstance(cell, MergedCell):
                continue
            if letter is None:
                letter = cell.column_letter
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        if letter is None:
            letter = get_column_letter(column_cells[0].column)
        if max_length:
            worksheet.column_dimensions[letter].width = min(max(max_length + 2, 8), 60)


def _safe_sheet_title(entry: MergedReportEntry) -> str:
    title = entry.sheet_name[:31]
    invalid = ":\\/?*[]"
    for char in invalid:
        title = title.replace(char, "-")
    return title


def _worksheet_is_empty(worksheet: Worksheet) -> bool:
    """True when the worksheet has no cell with a non-blank value."""
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                return False
    return True


def _validate_merged_workbook(
    workbook: Workbook,
    expected_titles: list[str],
) -> None:
    """Verify workbook/worksheet integrity before returning bytes to the caller.

    Checks (per spec):
      - workbook exists and has at least one worksheet
      - every expected worksheet exists
      - worksheet order matches the catalog (ascending report numbering)
      - worksheet names match the expected report titles
      - no worksheet is empty
    """
    if workbook is None:
        raise MergedDownloadError("Merged workbook was not created")
    if not expected_titles:
        raise MergedDownloadError("No Excel artifacts available for merge")
    if workbook.sheetnames != expected_titles:
        raise MergedDownloadError(
            "Merged workbook worksheet order/names do not match report catalog: "
            f"expected {expected_titles}, got {workbook.sheetnames}"
        )
    for title in expected_titles:
        worksheet = workbook[title]
        if _worksheet_is_empty(worksheet):
            raise MergedDownloadError(f"Worksheet '{title}' is empty; refusing to export")


def build_merged_excel(
    artifacts: list[AutomationArtifactModel],
) -> bytes:
    """Merge report Excel files into one workbook (one sheet per report, catalog order)."""
    ordered = _ordered_artifact_paths(artifacts, file_type="excel")
    if not ordered:
        raise MergedDownloadError("No Excel artifacts available for merge")

    dest = Workbook()
    default = dest.active
    dest.remove(default)

    used_titles: set[str] = set()
    expected_titles: list[str] = []
    for entry, path in ordered:
        source_wb = load_workbook(path, data_only=False)
        source_ws = source_wb.active
        title = _safe_sheet_title(entry)
        base = title
        suffix = 1
        while title in used_titles:
            trimmed = base[: max(1, 28 - len(str(suffix)))]
            title = f"{trimmed}_{suffix}"
            suffix += 1
        used_titles.add(title)
        expected_titles.append(title)
        dest_ws = dest.create_sheet(title=title)
        _copy_worksheet(source_ws, dest_ws)
        source_wb.close()

    _validate_merged_workbook(dest, expected_titles)

    output = BytesIO()
    dest.save(output)
    dest.close()
    return output.getvalue()
