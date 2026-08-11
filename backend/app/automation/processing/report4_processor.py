"""Report 4 post-ingestion processor (Cause-wise Top 10 Trains - 7 Types)."""

from __future__ import annotations

import csv
import logging
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Flowable, KeepTogether, PageBreak, Paragraph, SimpleDocTemplate, Spacer

from app.automation.config import config
from app.automation.formatting.artifact_titles import build_artifact_main_title
from app.automation.formatting.excel_print import apply_uniform_center_alignment
from app.automation.formatting.scr import highlight_south_central_railway_rows, row_contains_scr
from app.automation.formatting.text_pipeline import (
    normalize_report_title,
    prepare_output_for_rendering,
    verify_text_rendering,
)
from app.automation.formatting.topn_pdf import (
    build_topn_fitted_table,
    build_topn_section_heading_block,
    build_topn_title_block,
    choose_topn_landscape_layout,
)
from app.automation.processing.base import ProcessingResult
from app.automation.processing.column_config import project_topn_for_output, resolve_projection_column_keys
from app.automation.processing.report3_processor import Report3Processor
from app.automation.processing.topn_output_columns import topn_default_ids, topn_labels
from app.automation.report4_filters import COMPLAINT_TYPES_ORDERED, get_type_configs, TypeConfig
from app.automation.date_range import date_range_for_processing
from app.automation.utils import (
    ensure_directory,
    log_automation_event,
    resolve_report_dir,
)

# Report 4 PDF (Report 5 Cause Wise): height-based section packing on A3 landscape.
_REPORT4_PDF_MARGIN_PT = 12.0
_REPORT4_TITLE_AFTER_PT = 5.0
_REPORT4_SECTION_GAP_PT = 8.0
_REPORT4_BEFORE_TABLE_PT = 3.0
_REPORT4_TABLE_CELL_PADDING_PT = 2.0


def _measure_flowable_height(flowable: Flowable, width: float) -> float:
    """Return wrapped height for a single flowable at the given frame width."""
    _unused_width, height = flowable.wrap(width, 1_000_000)
    return float(height)


def _measure_flowables_height(flowables: list[Flowable], width: float) -> float:
    return sum(_measure_flowable_height(flowable, width) for flowable in flowables)


def _report4_page_title_flowables(main_title: str, table_width: float) -> list[Flowable]:
    return [
        build_topn_title_block(main_title, table_width),
        Spacer(1, _REPORT4_TITLE_AFTER_PT),
    ]


def _build_report4_section_table(
    section: TypeDataset,
    *,
    splittable: bool,
) -> tuple[Flowable, tuple[float, float], float, list[float]]:
    """Build the data table (or empty placeholder) for one grievance section."""
    styles = getSampleStyleSheet()
    if not section.rows:
        placeholder = Paragraph("No data available for this type.", styles["Normal"])
        return placeholder, landscape(A3), _REPORT4_PDF_MARGIN_PT, []

    table_data: list[list[object]] = [list(section.headers)]
    scr_row_indices: set[int] = set()
    for row_values in section.rows:
        table_data.append(list(row_values))
        if row_contains_scr(row_values):
            scr_row_indices.add(len(table_data) - 1)

    style_commands: list[tuple] = [
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
    ]
    for row_idx in scr_row_indices:
        style_commands.append(("BACKGROUND", (0, row_idx), (-1, row_idx), colors.yellow))
        style_commands.append(("TEXTCOLOR", (0, row_idx), (-1, row_idx), colors.black))

    table, pagesize, margin, col_widths = build_topn_fitted_table(
        table_data,
        style_commands,
        margin=_REPORT4_PDF_MARGIN_PT,
        splittable=splittable,
        cell_padding=_REPORT4_TABLE_CELL_PADDING_PT,
    )
    return table, pagesize, margin, col_widths


def _build_report4_section_pack(
    section: TypeDataset,
    *,
    table_width: float,
    usable_width: float,
    usable_height: float,
    title_height: float,
) -> tuple[list[Flowable], float, tuple[float, float], float, list[float]]:
    """
    Build flowables for one complete section and its measured pack height.

    Uses KeepTogether(heading + table) when the section fits on a page with the
    report title; otherwise keeps the heading with the table header and allows
    the table body to split across pages.
    """
    heading = build_topn_section_heading_block(
        section.type_config.section_title,
        table_width,
    )
    before_table = Spacer(1, _REPORT4_BEFORE_TABLE_PT)
    table, section_pagesize, section_margin, section_col_widths = _build_report4_section_table(
        section,
        splittable=False,
    )

    inner: list[Flowable] = [heading, before_table, table]
    inner_height = _measure_flowables_height(inner, usable_width)
    max_with_title = usable_height - title_height

    if inner_height <= max_with_title:
        block = KeepTogether(inner)
        return (
            [block],
            inner_height,
            section_pagesize,
            section_margin,
            section_col_widths,
        )

    table, section_pagesize, section_margin, section_col_widths = _build_report4_section_table(
        section,
        splittable=True,
    )
    heading_block = KeepTogether([heading, before_table])
    heading_height = _measure_flowables_height([heading, before_table], usable_width)
    table_height = _measure_flowable_height(table, usable_width)
    return (
        [heading_block, table],
        heading_height + table_height,
        section_pagesize,
        section_margin,
        section_col_widths,
    )


def _pack_report5_pdf_story(
    sections: list[TypeDataset],
    *,
    main_title: str,
    pagesize: tuple[float, float],
    margin: float,
    table_width: float,
) -> tuple[list[Flowable], tuple[float, float], float, float]:
    """Height-based pagination: up to three sections per page when they fit."""
    usable_width = pagesize[0] - (2 * margin)
    usable_height = pagesize[1] - (2 * margin)

    title_flowables = _report4_page_title_flowables(main_title, table_width)
    title_height = _measure_flowables_height(title_flowables, usable_width)

    story: list[Flowable] = []
    remaining = 0.0
    sections_on_page = 0

    for section in sections:
        pack_flowables, pack_height, new_pagesize, new_margin, new_col_widths = (
            _build_report4_section_pack(
                section,
                table_width=table_width,
                usable_width=usable_width,
                usable_height=usable_height,
                title_height=title_height,
            )
        )
        if new_col_widths and new_pagesize[0] > pagesize[0]:
            pagesize = new_pagesize
            margin = new_margin
            table_width = float(sum(new_col_widths))
            usable_width = pagesize[0] - (2 * margin)
            usable_height = pagesize[1] - (2 * margin)
            title_flowables = _report4_page_title_flowables(main_title, table_width)
            title_height = _measure_flowables_height(title_flowables, usable_width)
            pack_flowables, pack_height, _, _, _ = _build_report4_section_pack(
                section,
                table_width=table_width,
                usable_width=usable_width,
                usable_height=usable_height,
                title_height=title_height,
            )

        gap_height = _REPORT4_SECTION_GAP_PT if sections_on_page > 0 else 0.0
        needed = gap_height + pack_height

        def _start_new_page() -> None:
            nonlocal remaining, sections_on_page
            if story:
                story.append(PageBreak())
            story.extend(title_flowables)
            remaining = usable_height - title_height
            sections_on_page = 0

        if sections_on_page == 0:
            if title_height + pack_height > usable_height and story:
                _start_new_page()
            elif not story:
                story.extend(title_flowables)
                remaining = usable_height - title_height
        elif needed > remaining:
            _start_new_page()

        if sections_on_page > 0:
            story.append(Spacer(1, _REPORT4_SECTION_GAP_PT))
            remaining -= gap_height

        story.extend(pack_flowables)
        remaining -= pack_height
        sections_on_page += 1

    return story, pagesize, margin, table_width


logger = logging.getLogger(__name__)

PROCESSOR_NAME = "report4_causewise_top10_processor"

TOP_N = 10

DEFAULT_OUTPUT_HEADERS = topn_labels(topn_default_ids("types"), "types")

OUTPUT_HEADERS = DEFAULT_OUTPUT_HEADERS

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


@dataclass
class TypeDataset:
    """Data for a single complaint type section."""

    type_config: TypeConfig
    headers: list[str]
    rows: list[list[str]]


class Report4Processor:
    """Process Cause-wise Top 10 Trains datasets and emit formatted Excel/PDF."""

    processor_name = PROCESSOR_NAME

    def process(
        self,
        *,
        source_a_path: Path,
        report_slug: str,
        source_b_path: Path | None = None,
        type_datasets: dict[str, list[dict[str, str]]] | None = None,
        column_selection: dict[str, Any] | None = None,
    ) -> ProcessingResult:
        if source_a_path.suffix.lower() == ".pdf":
            return ProcessingResult(success=False, error="PDF cannot be used as processing input")

        if column_selection:
            log_automation_event(
                logger,
                "report4_filter_payload_received",
                selected_column_ids=column_selection.get("selected_column_ids"),
                column_order=column_selection.get("column_order"),
            )

        raw_sections, total_input_rows = self._load_raw_sections(
            source_a_path,
            type_datasets=type_datasets,
        )

        try:
            sections, output_headers, keys, config_source = self._project_sections(
                raw_sections,
                report_slug=report_slug,
                column_selection=column_selection,
            )
        except ValueError as exc:
            return ProcessingResult(
                input_row_count=total_input_rows,
                success=False,
                error=str(exc),
                source_a_path=str(source_a_path),
                source_a_rows=total_input_rows,
            )

        log_automation_event(
            logger,
            "report4_all_sections_validated",
            section_count=len(sections),
            selected_headers=output_headers,
            selected_column_count=len(keys),
        )

        normalized_sections: list[TypeDataset] = []
        for section in sections:
            normalized_title = normalize_report_title(
                section.type_config.section_title,
                report_slug=report_slug,
            )
            prepared_headers, prepared_rows = prepare_output_for_rendering(
                report_slug,
                section.headers,
                section.rows,
            )
            normalized_sections.append(
                TypeDataset(
                    type_config=replace(section.type_config, section_title=normalized_title),
                    headers=prepared_headers,
                    rows=prepared_rows,
                )
            )
        sections = normalized_sections

        date_range = date_range_for_processing(column_selection)


        main_title = build_artifact_main_title("types", date_range)

        filename_suffix = date_range.filename_suffix()
        excel_dir = ensure_directory(resolve_report_dir(config.output_excel_dir, report_slug))
        pdf_dir = ensure_directory(resolve_report_dir(config.output_pdf_dir, report_slug))
        base_name = f"Rail_Madad_Report_4_Cause_Wise_Top_10_Trains_{filename_suffix}"
        excel_path = excel_dir / f"{base_name}.xlsx"
        pdf_path = pdf_dir / f"{base_name}.pdf"

        try:
            self._write_excel(excel_path, sections, main_title=main_title)
            log_automation_event(logger, "report4_excel_generated", excel_path=str(excel_path))
            self._write_pdf(pdf_path, sections, main_title=main_title)
            log_automation_event(logger, "report4_pdf_generated", pdf_path=str(pdf_path))
            for section in sections:
                verify_text_rendering(
                    report_slug=report_slug,
                    headers=section.headers,
                    rows=section.rows,
                    pdf_path=pdf_path,
                )
        except Exception as exc:
            return ProcessingResult(
                input_row_count=total_input_rows,
                success=False,
                error=str(exc),
                source_a_path=str(source_a_path),
                source_a_rows=total_input_rows,
            )

        total_output_rows = sum(len(s.rows) for s in sections)

        log_automation_event(
            logger,
            "phase8_dataset_loaded",
            source_a=str(source_a_path),
            input_row_count=total_input_rows,
            section_count=len(sections),
            total_output_rows=total_output_rows,
        )

        return ProcessingResult(
            success=True,
            input_row_count=total_input_rows,
            processed_row_count=total_output_rows,
            excel_path=str(excel_path),
            pdf_path=str(pdf_path),
            source_a_path=str(source_a_path),
            source_a_rows=total_input_rows,
            output_columns=output_headers,
            visible_columns=output_headers,
            selected_column_ids=keys,
            column_order=list(keys),
            configuration_source=config_source,
        )

    def _load_raw_sections(
        self,
        source_a_path: Path,
        *,
        type_datasets: dict[str, list[dict[str, str]]] | None = None,
    ) -> tuple[list[tuple[TypeConfig, list[dict[str, str]]]], int]:
        type_configs = get_type_configs()
        type_by_name = {cfg.name: cfg for cfg in type_configs}
        raw_sections: list[tuple[TypeConfig, list[dict[str, str]]]] = []
        total_input_rows = 0

        if type_datasets:
            for type_config in type_configs:
                raw_rows = type_datasets.get(type_config.name, [])
                data_rows = self._exclude_total_rows(raw_rows)
                top_rows = self._top_n_by_received(data_rows)
                total_input_rows += len(data_rows)
                raw_sections.append((type_config, top_rows))
            return raw_sections, total_input_rows

        index_entries = self._read_combined_index(source_a_path)
        if index_entries:
            for type_name in COMPLAINT_TYPES_ORDERED:
                type_config = type_by_name.get(type_name)
                if type_config is None:
                    continue
                entry = index_entries.get(type_name)
                top_rows: list[dict[str, str]] = []
                if entry is not None and str(entry.get("status", "")).lower() == "success":
                    csv_path = Path(str(entry.get("csv_path") or ""))
                    if csv_path.is_file():
                        raw_rows, _ = self._read_csv(csv_path)
                        data_rows = self._exclude_total_rows(raw_rows)
                        top_rows = self._top_n_by_received(data_rows)
                        total_input_rows += len(data_rows)
                    else:
                        log_automation_event(
                            logger,
                            "type_csv_not_found",
                            type_name=type_name,
                            expected_path=str(csv_path),
                        )
                raw_sections.append((type_config, top_rows))
            return raw_sections, total_input_rows

        base_dir = source_a_path.parent
        for type_config in type_configs:
            type_slug = type_config.name.lower().replace(" ", "_").replace("&", "and")
            csv_path = base_dir / f"report4_{type_slug}_raw.csv"
            top_rows = []
            if csv_path.exists():
                raw_rows, _ = self._read_csv(csv_path)
                data_rows = self._exclude_total_rows(raw_rows)
                top_rows = self._top_n_by_received(data_rows)
                total_input_rows += len(data_rows)
            else:
                log_automation_event(
                    logger,
                    "type_csv_not_found",
                    type_name=type_config.name,
                    expected_path=str(csv_path),
                )
            raw_sections.append((type_config, top_rows))
        return raw_sections, total_input_rows

    def _project_sections(
        self,
        raw_sections: list[tuple[TypeConfig, list[dict[str, str]]]],
        *,
        report_slug: str,
        column_selection: dict[str, Any] | None = None,
    ) -> tuple[list[TypeDataset], list[str], list[str], str]:
        keys, config_source = resolve_projection_column_keys(
            report_slug,
            column_selection=column_selection,
        )
        expected_headers: list[str] | None = None
        projected_sections: list[TypeDataset] = []

        for type_config, top_rows in raw_sections:
            log_automation_event(
                logger,
                "report4_section_projection_started",
                section_name=type_config.name,
                source_row_count=len(top_rows),
            )
            canonical_rows = Report3Processor.build_canonical_rows(top_rows)
            section_headers, section_rows, _visible, _keys, _source = project_topn_for_output(
                report_slug,
                canonical_rows,
                selected_keys=keys,
                config_source=config_source,
            )
            if expected_headers is None:
                expected_headers = list(section_headers)
            elif section_headers != expected_headers:
                raise ValueError(
                    "REPORT4_SECTION_COLUMN_MISMATCH: "
                    f"{type_config.name} expected={expected_headers} actual={section_headers}"
                )
            log_automation_event(
                logger,
                "report4_section_projection_completed",
                section_name=type_config.name,
                top10_row_count=len(section_rows),
                selected_headers=section_headers,
                projection_success=True,
            )
            projected_sections.append(
                TypeDataset(
                    type_config=type_config,
                    headers=section_headers,
                    rows=section_rows,
                )
            )

        if expected_headers is None:
            expected_headers = []

        return projected_sections, expected_headers, keys, config_source

    @classmethod
    def build_projected_sections(
        cls,
        source_a_path: Path,
        *,
        report_slug: str = "types",
        column_selection: dict[str, Any] | None = None,
        selected_keys: list[str] | None = None,
    ) -> tuple[list[TypeDataset], list[str], list[str], str]:
        processor = cls()
        raw_sections, _ = processor._load_raw_sections(source_a_path)
        if selected_keys is not None:
            selection = {
                "selected_column_ids": selected_keys,
                "column_order": selected_keys,
                "configuration_source": "manual_snapshot",
            }
            return processor._project_sections(
                raw_sections,
                report_slug=report_slug,
                column_selection=selection,
            )
        return processor._project_sections(
            raw_sections,
            report_slug=report_slug,
            column_selection=column_selection,
        )

    @staticmethod
    def _read_csv(path: Path) -> tuple[list[dict[str, str]], list[str]]:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            headers = list(reader.fieldnames or [])
            rows = [{header: row.get(header, "") for header in headers} for row in reader]
        return rows, headers

    @staticmethod
    def _read_combined_index(source_a_path: Path) -> dict[str, dict[str, str]]:
        if source_a_path.name != "types_combined_index.csv":
            return {}
        if not source_a_path.is_file():
            return {}
        entries: dict[str, dict[str, str]] = {}
        with source_a_path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                type_name = (row.get("type_name") or "").strip()
                if not type_name:
                    continue
                entries[type_name] = {
                    "type_name": type_name,
                    "csv_path": (row.get("csv_path") or "").strip(),
                    "row_count": (row.get("row_count") or "0").strip(),
                    "status": (row.get("status") or "").strip(),
                    "error": (row.get("error") or "").strip(),
                }
        return entries

    @staticmethod
    def _is_total_row(row: dict[str, str]) -> bool:
        for key in ("Train No.", "Train Name", "Owning Zone"):
            val = row.get(key, "")
            if "total" in val.strip().lower():
                return True
        return False

    def _exclude_total_rows(
        self,
        rows: list[dict[str, str]],
    ) -> list[dict[str, str]]:
        return [r for r in rows if not self._is_total_row(r)]

    @staticmethod
    def _received_sort_key(row: dict[str, str]) -> tuple[int, str]:
        raw = str(row.get("Received") or row.get("received") or "0").replace(",", "").strip()
        try:
            value = int(float(raw)) if raw else 0
        except ValueError:
            value = 0
        train_no = str(row.get("Train No.") or row.get("Train No") or "")
        return (-value, train_no)

    def _top_n_by_received(self, rows: list[dict[str, str]]) -> list[dict[str, str]]:
        """Take highest-Received trains after numeric sort (do not trust CSV order)."""
        ordered = sorted(rows, key=self._received_sort_key)
        return ordered[:TOP_N]

    @staticmethod
    def _train_no_column_index(headers: list[str]) -> int | None:
        for idx, header in enumerate(headers, start=1):
            if header.strip() in {"Train No.", "Train No"}:
                return idx
        return None

    @staticmethod
    def _section_width(sections: list[TypeDataset]) -> int:
        if not sections:
            return 1
        return max(len(section.headers) for section in sections)

    def _write_excel(
        self,
        target_path: Path,
        sections: list[TypeDataset],
        *,
        main_title: str,
    ) -> None:
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Report 4"

        col_count = self._section_width(sections)
        main_title = normalize_report_title(main_title, report_slug="types")
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        title_cell = worksheet.cell(row=1, column=1, value=main_title)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        current_row = 2

        for section_idx, section in enumerate(sections):
            section_cols = max(len(section.headers), 1)
            worksheet.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=section_cols,
            )
            section_title_cell = worksheet.cell(
                row=current_row,
                column=1,
                value=section.type_config.section_title,
            )
            section_title_cell.font = Font(bold=True, size=11)
            section_title_cell.alignment = Alignment(horizontal="left")
            current_row += 1

            for col_idx, header in enumerate(section.headers, start=1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.border = THIN_BORDER
            current_row += 1

            train_no_col = self._train_no_column_index(section.headers)
            data_start = current_row
            for row_values in section.rows:
                for col_idx, value in enumerate(row_values, start=1):
                    cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                    cell.border = THIN_BORDER
                    if train_no_col is not None and col_idx == train_no_col:
                        cell.number_format = "@"
                current_row += 1

            if section.rows:
                highlight_south_central_railway_rows(
                    worksheet,
                    start_row=data_start,
                    end_row=current_row - 1,
                    start_col=1,
                    end_col=len(section.headers),
                )

            if section_idx < len(sections) - 1:
                current_row += 1

        apply_uniform_center_alignment(worksheet)

        workbook.save(temp_path)
        temp_path.replace(target_path)

    def _write_pdf(
        self,
        target_path: Path,
        sections: list[TypeDataset],
        *,
        main_title: str,
    ) -> None:
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")

        sample_headers = next(
            (list(section.headers) for section in sections if section.headers),
            ["Train Name"],
        )
        pagesize, col_widths, margin = choose_topn_landscape_layout(
            sample_headers,
            margin=_REPORT4_PDF_MARGIN_PT,
        )
        table_width = float(sum(col_widths))
        main_title = normalize_report_title(main_title, report_slug="types")

        story, pagesize, margin, _table_width = _pack_report5_pdf_story(
            sections,
            main_title=main_title,
            pagesize=pagesize,
            margin=margin,
            table_width=table_width,
        )

        doc = SimpleDocTemplate(
            str(temp_path),
            pagesize=pagesize,
            leftMargin=margin,
            rightMargin=margin,
            topMargin=margin,
            bottomMargin=margin,
        )
        doc.build(story)
        temp_path.replace(target_path)
