"""Bottom Performed Trains Report — PDF and Excel generation from structured result.json."""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, Spacer, TableStyle

from app.automation.bottom_report_filters import BOTTOM_REPORT_FILE_STEM
from app.automation.config import config
from app.automation.date_range import ReportDateRange, date_range_for_processing
from app.automation.formatting.pdf_fonts import ensure_pdf_unicode_fonts, pdf_font_bold, pdf_font_regular
from app.automation.formatting.pdf_table import SAFE_MARGIN_PT
from app.automation.formatting.scr import SCR_FILL, SCR_FONT
from app.automation.processing.base import ProcessingResult
from app.automation.processing.bottom_report_models import (
    BOTTOM_REPORT_SLUG,
    NUM_OUTPUT_COLUMNS,
    OUTPUT_COLUMNS,
    RESULT_JSON_FILENAME,
    SECTION_DISPLAY_TITLES,
    SECTION_RENDER_ORDER,
    BottomReportResult,
    DivisionResult,
    SectionResult,
    footer_message,
    section_subheading,
)
from app.automation.utils import ensure_directory, log_automation_event, resolve_run_scoped_dir

logger = logging.getLogger(__name__)

PROCESSOR_NAME = "bottom_report_processor"
SHEET_TITLE = "Bottom Performed Trains"
PDF_MARGIN_PT = SAFE_MARGIN_PT
OWNING_COL_INDEX = 4

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
HEADER_FONT = Font(bold=True, color="000000", size=8)
TITLE_FONT = Font(bold=True, size=12)
SUBTITLE_FONT = Font(bold=True, size=8)
SECTION_FONT = Font(bold=True, size=10)
FOOTER_FONT = Font(bold=True, size=8)
CELL_FONT = Font(size=8)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

EXCEL_COL_WIDTHS = (8, 12, 48, 18, 14)

# Spec §17 proportions (must sum to 1.0).
PDF_COL_WIDTH_RATIOS = (0.08, 0.12, 0.45, 0.20, 0.15)

# Compact reference spacing (points).
PDF_GAP_TITLE_TO_FIRST_SECTION = 10
PDF_GAP_SECTION_TO_TABLE = 4
PDF_GAP_BETWEEN_DIVISION_TABLES = 8
PDF_GAP_BETWEEN_SECTIONS = 12
PDF_CELL_PAD = 3


@dataclass(frozen=True)
class BottomTableBlock:
    subtitle: str
    trains: tuple[Any, ...]
    message_row: str | None
    footer: str | None


def _escape_xml(text: str) -> str:
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _format_title_lines(date_range: ReportDateRange) -> tuple[str, str]:
    line1 = "Bottom performed trains based on comprehensive drop down"
    if date_range.date_from and date_range.date_to and date_range.date_from != date_range.date_to:
        date_label = (
            f"{date_range.date_from.strftime('%d.%m.%Y')} to "
            f"{date_range.date_to.strftime('%d.%m.%Y')}"
        )
    else:
        end = date_range.date_to or date_range.date_from
        date_label = end.strftime("%d.%m.%Y") if end else ""
    line2 = f"(territorial basis) as on {date_label}"
    return line1, line2


def _division_table_block(section_id: str, division: DivisionResult) -> BottomTableBlock:
    water_style = section_id == "water_availability"
    subtitle = section_subheading(
        section_id,
        None if water_style else division.division_code,
    )
    message_row = None
    if division.no_train_message and not division.qualifying_trains:
        message_row = division.no_train_message
    return BottomTableBlock(
        subtitle=subtitle,
        trains=tuple(division.qualifying_trains),
        message_row=message_row,
        footer=footer_message(division.division_received, division.division_code),
    )


def _no_division_table_block(section_id: str, message: str) -> BottomTableBlock:
    return BottomTableBlock(
        subtitle=section_subheading(section_id, None),
        trains=(),
        message_row=message,
        footer=None,
    )


def _section_table_blocks(section_id: str, section: SectionResult) -> list[BottomTableBlock]:
    """One bordered table per qualifying division (never merge divisions)."""
    if section.qualifying_divisions:
        return [
            _division_table_block(section_id, division)
            for division in section.qualifying_divisions
        ]
    message = section.no_division_message or "No Div. has figured with more than 20 complaints"
    return [_no_division_table_block(section_id, message)]


def _pdf_column_widths() -> list[float]:
    page_width, _ = landscape(A4)
    usable = page_width - 2 * PDF_MARGIN_PT
    return [usable * ratio for ratio in PDF_COL_WIDTH_RATIOS]


def _apply_excel_borders(ws, row: int, *, cols: int = NUM_OUTPUT_COLUMNS) -> None:
    for col in range(1, cols + 1):
        ws.cell(row=row, column=col).border = THIN_BORDER


def _merge_row(ws, row: int, value: str, *, font: Font, alignment: Alignment | None = None) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_OUTPUT_COLUMNS)
    cell = ws.cell(row=row, column=1, value=value)
    cell.font = font
    cell.alignment = alignment or CENTER
    _apply_excel_borders(ws, row)


def _highlight_sc_owning_cells(worksheet, *, start_row: int, end_row: int) -> None:
    for row_idx in range(start_row, end_row + 1):
        cell = worksheet.cell(row=row_idx, column=OWNING_COL_INDEX + 1)
        if str(cell.value or "").strip().upper() == "SC":
            cell.fill = SCR_FILL
            cell.font = SCR_FONT


class BottomReportProcessor:
    processor_name = PROCESSOR_NAME

    def process(
        self,
        *,
        source_a_path: Path,
        report_slug: str,
        source_b_path: Path | None = None,
        column_selection: dict[str, Any] | None = None,
    ) -> ProcessingResult:
        _ = source_b_path
        if source_a_path.suffix.lower() != ".json":
            return ProcessingResult(
                success=False,
                error="BOTTOM_REPORT_INVALID_SOURCE: expected result.json",
                source_a_path=str(source_a_path),
            )
        if not source_a_path.is_file():
            return ProcessingResult(
                success=False,
                error="BOTTOM_REPORT_MISSING: result.json not found",
                source_a_path=str(source_a_path),
            )

        try:
            result = BottomReportResult.load(source_a_path)
        except Exception as exc:
            return ProcessingResult(
                success=False,
                error=f"BOTTOM_REPORT_PARSE_FAILED: {exc}",
                source_a_path=str(source_a_path),
            )

        date_range = date_range_for_processing(column_selection)
        title_line1, title_line2 = _format_title_lines(date_range)
        run_id = (column_selection or {}).get("run_id") if column_selection else None
        slug = report_slug or BOTTOM_REPORT_SLUG

        if run_id:
            excel_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_excel_dir, slug, str(run_id))
            )
            pdf_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_pdf_dir, slug, str(run_id))
            )
        else:
            parent = source_a_path.parent
            excel_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_excel_dir, slug, parent.name)
            )
            pdf_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_pdf_dir, slug, parent.name)
            )

        suffix = date_range.filename_suffix().strip("_") or "report"
        excel_path = excel_dir / f"{BOTTOM_REPORT_FILE_STEM}_{suffix}.xlsx"
        pdf_path = pdf_dir / f"{BOTTOM_REPORT_FILE_STEM}_{suffix}.pdf"

        try:
            self._write_excel(excel_path, result, title_line1, title_line2)
            self._write_pdf(pdf_path, result, title_line1, title_line2)
        except Exception as exc:
            log_automation_event(
                logger,
                "bottom_report_processor_failed",
                error=str(exc),
            )
            return ProcessingResult(
                success=False,
                error=f"BOTTOM_REPORT_OUTPUT_FAILED: {exc}",
                source_a_path=str(source_a_path),
            )

        return ProcessingResult(
            success=True,
            excel_path=str(excel_path),
            pdf_path=str(pdf_path),
            processor_used=PROCESSOR_NAME,
            source_a_path=str(source_a_path),
            processed_row_count=self._count_output_rows(result),
        )

    @staticmethod
    def _count_output_rows(result: BottomReportResult) -> int:
        total = 0
        for section_id in SECTION_RENDER_ORDER:
            section = result.sections.get(section_id)
            if section is None:
                continue
            for division in section.qualifying_divisions:
                total += len(division.qualifying_trains)
        return total

    def _write_excel(
        self,
        path: Path,
        result: BottomReportResult,
        title_line1: str,
        title_line2: str,
    ) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_TITLE[:31]
        row = 1

        _merge_row(ws, row, title_line1, font=TITLE_FONT)
        row += 1
        _merge_row(ws, row, title_line2, font=TITLE_FONT)
        row += 2

        for col_idx, width in enumerate(EXCEL_COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for section_id in SECTION_RENDER_ORDER:
            section = result.sections.get(section_id)
            if section is None:
                continue

            _merge_row(ws, row, SECTION_DISPLAY_TITLES[section_id], font=SECTION_FONT)
            row += 1

            blocks = _section_table_blocks(section_id, section)
            for block_idx, block in enumerate(blocks):
                row = self._write_excel_table_block(ws, row, block)
                if block_idx < len(blocks) - 1:
                    row += 1

            row += 1

        wb.save(path)

    def _write_excel_table_block(self, ws, start_row: int, block: BottomTableBlock) -> int:
        row = start_row
        _merge_row(ws, row, block.subtitle, font=SUBTITLE_FONT)
        row += 1

        for col_idx, header in enumerate(OUTPUT_COLUMNS, start=1):
            display = header
            if header == "No of complaints":
                display = "No of\ncomplaints"
            elif header == "Owning Rly":
                display = "Owning\nRly"
            cell = ws.cell(row=row, column=col_idx, value=display)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = CENTER
        ws.row_dimensions[row].height = 28
        header_row = row
        row += 1
        data_start = row

        if block.message_row:
            _merge_row(ws, row, block.message_row, font=CELL_FONT)
            row += 1
        else:
            for sn, train in enumerate(block.trains, start=1):
                values = [
                    sn,
                    train.train_no,
                    train.from_to,
                    train.complaint_count,
                    train.owning_railway,
                ]
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.font = CELL_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = LEFT_WRAP if col_idx == 3 else CENTER
                ws.row_dimensions[row].height = 24 if len(str(train.from_to)) > 35 else 18
                row += 1

        if block.footer:
            _merge_row(ws, row, block.footer, font=FOOTER_FONT)
            row += 1

        if block.trains and not block.message_row:
            _highlight_sc_owning_cells(
                ws,
                start_row=data_start,
                end_row=row - (2 if block.footer else 1),
            )

        return row

    def _write_pdf(
        self,
        path: Path,
        result: BottomReportResult,
        title_line1: str,
        title_line2: str,
    ) -> None:
        ensure_pdf_unicode_fonts()
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "BottomTitle",
            parent=styles["Heading1"],
            fontName=pdf_font_bold(),
            fontSize=12,
            leading=14,
            alignment=TA_CENTER,
            spaceAfter=2,
        )
        section_style = ParagraphStyle(
            "BottomSection",
            parent=styles["Heading2"],
            fontName=pdf_font_bold(),
            fontSize=10,
            leading=12,
            alignment=TA_CENTER,
            spaceAfter=3,
        )
        subtitle_style = ParagraphStyle(
            "BottomSubtitle",
            parent=styles["Normal"],
            fontName=pdf_font_bold(),
            fontSize=8,
            leading=10,
            alignment=TA_CENTER,
        )
        cell_style = ParagraphStyle(
            "BottomCell",
            parent=styles["Normal"],
            fontName=pdf_font_regular(),
            fontSize=8,
            leading=10,
            alignment=TA_CENTER,
        )
        from_to_style = ParagraphStyle(
            "BottomFromTo",
            parent=cell_style,
            alignment=TA_LEFT,
        )
        footer_style = ParagraphStyle(
            "BottomFooter",
            parent=styles["Normal"],
            fontName=pdf_font_bold(),
            fontSize=8,
            leading=10,
            alignment=TA_CENTER,
        )

        story: list[Any] = []
        story.append(Paragraph(_escape_xml(title_line1), title_style))
        story.append(Paragraph(_escape_xml(title_line2), title_style))
        story.append(Spacer(1, PDF_GAP_TITLE_TO_FIRST_SECTION))

        col_widths = _pdf_column_widths()
        section_ids = [sid for sid in SECTION_RENDER_ORDER if sid in result.sections]

        for section_pos, section_id in enumerate(section_ids):
            section = result.sections[section_id]

            story.append(
                Paragraph(_escape_xml(SECTION_DISPLAY_TITLES[section_id]), section_style)
            )
            story.append(Spacer(1, PDF_GAP_SECTION_TO_TABLE))

            blocks = _section_table_blocks(section_id, section)
            for block_idx, block in enumerate(blocks):
                # repeatRows=2 keeps subtitle + column headers with continued pages.
                story.append(
                    self._build_pdf_table(
                        block,
                        col_widths,
                        subtitle_style,
                        cell_style,
                        from_to_style,
                        footer_style,
                    )
                )
                if block_idx < len(blocks) - 1:
                    story.append(Spacer(1, PDF_GAP_BETWEEN_DIVISION_TABLES))

            if section_pos < len(section_ids) - 1:
                story.append(Spacer(1, PDF_GAP_BETWEEN_SECTIONS))

        doc = SimpleDocTemplate(
            str(path),
            pagesize=landscape(A4),
            leftMargin=PDF_MARGIN_PT,
            rightMargin=PDF_MARGIN_PT,
            topMargin=PDF_MARGIN_PT,
            bottomMargin=PDF_MARGIN_PT,
        )
        doc.build(story)

    def _build_pdf_table(
        self,
        block: BottomTableBlock,
        col_widths: list[float],
        subtitle_style: ParagraphStyle,
        cell_style: ParagraphStyle,
        from_to_style: ParagraphStyle,
        footer_style: ParagraphStyle,
    ) -> LongTable:
        def _p(text: str, style: ParagraphStyle) -> Paragraph:
            return Paragraph(_escape_xml(str(text)), style)

        # Plain header strings — avoid <br/> in PDF (XML escape would show it literally).
        header_labels = list(OUTPUT_COLUMNS)

        table_data: list[list[Any]] = [
            [_p(block.subtitle, subtitle_style), "", "", "", ""],
            header_labels,
        ]

        data_row_start = 2
        if block.message_row:
            table_data.append([_p(block.message_row, cell_style), "", "", "", ""])
        else:
            for sn, train in enumerate(block.trains, start=1):
                table_data.append([
                    str(sn),
                    str(train.train_no),
                    _p(train.from_to, from_to_style),
                    str(train.complaint_count),
                    str(train.owning_railway),
                ])

        footer_row_idx: int | None = None
        if block.footer:
            footer_row_idx = len(table_data)
            table_data.append([_p(block.footer, footer_style), "", "", "", ""])

        table = LongTable(table_data, colWidths=col_widths, repeatRows=2)
        style_cmds: list[tuple] = [
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("SPAN", (0, 0), (-1, 0)),
            ("BACKGROUND", (0, 1), (-1, 1), colors.Color(0.91, 0.91, 0.91)),
            ("FONTNAME", (0, 1), (-1, 1), pdf_font_bold()),
            ("FONTSIZE", (0, 1), (-1, 1), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ALIGN", (2, data_row_start), (2, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), PDF_CELL_PAD),
            ("RIGHTPADDING", (0, 0), (-1, -1), PDF_CELL_PAD),
            ("TOPPADDING", (0, 0), (-1, -1), PDF_CELL_PAD),
            ("BOTTOMPADDING", (0, 0), (-1, -1), PDF_CELL_PAD),
        ]

        if block.message_row:
            style_cmds.append(("SPAN", (0, data_row_start), (-1, data_row_start)))

        if footer_row_idx is not None:
            style_cmds.append(("SPAN", (0, footer_row_idx), (-1, footer_row_idx)))

        if block.trains and not block.message_row:
            last_data_row = data_row_start + len(block.trains) - 1
            style_cmds.append(("FONTSIZE", (0, data_row_start), (-1, last_data_row), 8))
            style_cmds.append(("FONTNAME", (0, data_row_start), (-1, last_data_row), pdf_font_regular()))
            for r_idx, train in enumerate(block.trains, start=data_row_start):
                if str(train.owning_railway or "").strip().upper() == "SC":
                    style_cmds.append(
                        ("BACKGROUND", (OWNING_COL_INDEX, r_idx), (OWNING_COL_INDEX, r_idx), colors.yellow)
                    )

        table.setStyle(TableStyle(style_cmds))
        return table
