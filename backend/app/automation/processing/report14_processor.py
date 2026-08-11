"""Report 14 post-ingestion processor — Watering Complaints Previous + Upcoming.

Loads previous_watering.csv and upcoming_watering.csv via report14_combined_index.csv
and emits one combined XLSX + one PDF with side-by-side metric columns.
"""

from __future__ import annotations

import csv
import logging
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, Spacer, TableStyle

from app.automation.config import config
from app.automation.date_range import date_range_for_processing
from app.automation.formatting.excel_print import apply_uniform_center_alignment
from app.automation.formatting.pdf_fonts import ensure_pdf_unicode_fonts, pdf_font_bold, pdf_font_regular
from app.automation.formatting.pdf_table import SAFE_MARGIN_PT, fit_column_widths
from app.automation.formatting.text_pipeline import normalize_report_title
from app.automation.processing.base import ProcessingResult
from app.automation.processing.report14_output_columns import (
    REPORT14_LABEL_BY_ID,
    REPORT14_PREV_GROUP_TITLE,
    REPORT14_SUB_HEADERS,
    REPORT14_UP_GROUP_TITLE,
    report14_default_ids,
    report14_labels,
    validate_selected_report14_fields,
)
from app.automation.report14_filters import (
    OUTPUT_HEADERS,
    SOURCE_PREVIOUS,
    SOURCE_UPCOMING,
)
from app.automation.utils import (
    ensure_directory,
    log_automation_event,
    resolve_run_scoped_dir,
)

logger = logging.getLogger(__name__)

PROCESSOR_NAME = "report14_processor"
REPORT14_SHEET_TITLE = "Watering Complaints"
REPORT14_FILE_STEM = "Rail_Madad_Report_14_Watering_Complaints"

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
TOTAL_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
HEADER_FONT = Font(bold=True, color="000000", size=9)

_PDF_MARGIN_PT = min(SAFE_MARGIN_PT, 12.0)

# Official merged-table metrics (horizontal Division join).
_MERGE_METRICS = ("Received", "% Share", "Average Rating")
_MERGE_KEY = "Division"


def _escape_paragraph_xml(text: str) -> str:
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _parse_num(value: str) -> float:
    cleaned = re.sub(r"[^\d.\-]", "", str(value or "").replace(",", ""))
    if not cleaned or cleaned in {".", "-", "-."}:
        return 0.0
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _format_num(value: float) -> str:
    if value == int(value):
        return str(int(value))
    return f"{value:.2f}"


class Report14Processor:
    """Merge Previous + Upcoming watering extracts into combined Excel/PDF."""

    processor_name = PROCESSOR_NAME

    def process(
        self,
        *,
        source_a_path: Path,
        report_slug: str,
        source_b_path: Path | None = None,
        column_selection: dict[str, Any] | None = None,
    ) -> ProcessingResult:
        _ = source_b_path
        if source_a_path.suffix.lower() == ".pdf":
            return ProcessingResult(success=False, error="PDF cannot be used as processing input")

        prev_rows, prev_headers, up_rows, up_headers, total_input, prev_total, up_total = (
            self._load_sources(source_a_path)
        )
        if prev_rows is None or up_rows is None:
            return ProcessingResult(
                success=False,
                error="REPORT14_TABLE_MISSING: need both previous and upcoming sources",
                source_a_path=str(source_a_path),
                input_row_count=total_input,
            )

        selected_ids = report14_default_ids()
        if column_selection:
            raw = (
                column_selection.get("selected_column_ids")
                or column_selection.get("column_order")
                or []
            )
            if raw:
                selected_ids = validate_selected_report14_fields(raw)

        try:
            merged_headers, merged_rows = self._merge_by_division(
                prev_rows,
                prev_headers or [],
                up_rows,
                up_headers or [],
                prev_total=prev_total,
                up_total=up_total,
            )
        except ValueError as exc:
            return ProcessingResult(
                success=False,
                error=f"REPORT14_MERGE_FAILED: {exc}",
                source_a_path=str(source_a_path),
                input_row_count=total_input,
            )

        visible_headers = report14_labels(selected_ids)
        id_to_label = REPORT14_LABEL_BY_ID
        col_indexes = [
            merged_headers.index(id_to_label[cid])
            for cid in selected_ids
            if cid in id_to_label and id_to_label[cid] in merged_headers
        ]
        if not col_indexes:
            col_indexes = list(range(len(merged_headers)))
            visible_headers = list(merged_headers)
            selected_ids = report14_default_ids()

        projected_rows = [
            [row[i] if i < len(row) else "" for i in col_indexes] for row in merged_rows
        ]

        date_range = date_range_for_processing(column_selection)
        main_title = "WATERING COMPLAINTS"
        title_suffix = date_range.title_suffix()
        filename_suffix = date_range.filename_suffix()
        subtitle = f"South Central Railway — Division Wise {title_suffix}"
        report_date = date_range.display_to() or date_range.display_from() or title_suffix.strip("()")

        run_id = (column_selection or {}).get("run_id") if column_selection else None
        if run_id:
            excel_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_excel_dir, report_slug, str(run_id))
            )
            pdf_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_pdf_dir, report_slug, str(run_id))
            )
        else:
            parent = source_a_path.parent
            excel_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_excel_dir, report_slug, parent.name)
            )
            pdf_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_pdf_dir, report_slug, parent.name)
            )

        base_name = f"{REPORT14_FILE_STEM}_{filename_suffix}"
        excel_path = excel_dir / f"{base_name}.xlsx"
        pdf_path = pdf_dir / f"{base_name}.pdf"

        try:
            self._write_excel(
                excel_path,
                visible_headers,
                projected_rows,
                main_title=main_title,
                subtitle=subtitle,
                report_date=report_date,
            )
            log_automation_event(logger, "report14_excel_generated", excel_path=str(excel_path))
        except Exception as exc:
            return ProcessingResult(
                success=False,
                error=f"REPORT14_XLSX_FAILED: {exc}",
                source_a_path=str(source_a_path),
                input_row_count=total_input,
            )

        try:
            self._write_pdf(
                pdf_path,
                visible_headers,
                projected_rows,
                main_title=main_title,
                subtitle=subtitle,
                report_date=report_date,
            )
            log_automation_event(logger, "report14_pdf_generated", pdf_path=str(pdf_path))
        except Exception as exc:
            return ProcessingResult(
                success=False,
                error=f"REPORT14_PDF_FAILED: {exc}",
                source_a_path=str(source_a_path),
                input_row_count=total_input,
                excel_path=str(excel_path),
            )

        data_rows = [r for r in projected_rows if not self._is_total_row(r)]
        log_automation_event(
            logger,
            "report14_processing_completed",
            source_a=str(source_a_path),
            input_row_count=total_input,
            total_output_rows=len(data_rows),
        )

        return ProcessingResult(
            success=True,
            attempted=True,
            input_row_count=total_input,
            processed_row_count=len(data_rows),
            excel_path=str(excel_path),
            pdf_path=str(pdf_path),
            source_a_path=str(source_a_path),
            source_a_rows=total_input,
            output_columns=visible_headers,
            visible_columns=visible_headers,
            selected_column_ids=list(selected_ids),
            column_order=list(selected_ids),
            configuration_source="manual_snapshot"
            if column_selection and column_selection.get("selected_column_ids")
            else "default",
        )

    def _load_sources(
        self, source_a_path: Path
    ) -> tuple[
        list[dict[str, str]] | None,
        list[str] | None,
        list[dict[str, str]] | None,
        list[str] | None,
        int,
        dict[str, str] | None,
        dict[str, str] | None,
    ]:
        index_entries = self._read_combined_index(source_a_path)
        base_dir = source_a_path.parent

        def resolve_csv(source_id: str, filename: str) -> Path | None:
            entry = index_entries.get(source_id)
            if entry and str(entry.get("status", "")).lower() == "success":
                candidate = Path(str(entry.get("csv_path") or ""))
                if candidate.is_file():
                    return candidate
            fallback = base_dir / filename
            return fallback if fallback.is_file() else None

        prev_path = resolve_csv(SOURCE_PREVIOUS.source_id, SOURCE_PREVIOUS.filename)
        up_path = resolve_csv(SOURCE_UPCOMING.source_id, SOURCE_UPCOMING.filename)
        if prev_path is None or up_path is None:
            return None, None, None, None, 0, None, None

        prev_rows, prev_headers = self._read_csv(prev_path)
        up_rows, up_headers = self._read_csv(up_path)
        prev_data, prev_total = self._split_total_row(prev_rows)
        up_data, up_total = self._split_total_row(up_rows)
        total_input = len(prev_data) + len(up_data)
        return prev_data, prev_headers, up_data, up_headers, total_input, prev_total, up_total

    def _read_combined_index(self, source_a_path: Path) -> dict[str, dict[str, str]]:
        if source_a_path.name != "report14_combined_index.csv":
            return {}
        if not source_a_path.is_file():
            return {}
        entries: dict[str, dict[str, str]] = {}
        with source_a_path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                source_id = (row.get("source_id") or "").strip()
                if not source_id:
                    continue
                entries[source_id] = {
                    "source_id": source_id,
                    "csv_path": (row.get("csv_path") or "").strip(),
                    "status": (row.get("status") or "").strip(),
                }
        return entries

    @staticmethod
    def _read_csv(path: Path) -> tuple[list[dict[str, str]], list[str]]:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            headers = list(reader.fieldnames or [])
            rows = [{header: row.get(header, "") for header in headers} for row in reader]
        return rows, headers

    @staticmethod
    def _split_total_row(
        rows: list[dict[str, str]],
    ) -> tuple[list[dict[str, str]], dict[str, str] | None]:
        if not rows:
            return [], None
        last = rows[-1]
        if any(str(v).strip().lower() == "total" for v in last.values()):
            return rows[:-1], last
        return rows, None

    @staticmethod
    def _division_key(name: str) -> str:
        compact = re.sub(r"\s+", " ", (name or "").strip()).upper()
        compact = re.sub(r"\s+DIVISION$", "", compact)
        return compact

    @staticmethod
    def _pick_display_division(*names: str) -> str:
        cleaned = [re.sub(r"\s+", " ", n.strip()) for n in names if n and n.strip()]
        if not cleaned:
            return ""
        return max(cleaned, key=len)

    def _header_lookup(self, headers: list[str]) -> dict[str, str]:
        lookup: dict[str, str] = {}
        for h in headers:
            hl = h.strip().lower()
            lookup[hl] = h
            if hl in {"avg. rating", "avg rating", "average rating"}:
                lookup["average rating"] = h
            if hl in {"% share", "share %", "share"}:
                lookup["% share"] = h
        return lookup

    def _division_for_row(self, row: dict[str, str], headers: list[str]) -> tuple[str, str]:
        header_map = self._header_lookup(headers)
        div_col = header_map.get("division")
        if not div_col:
            raise ValueError("Division column missing from extracted CSV")
        display = str(row.get(div_col, "")).strip()
        if not display or display.lower() == "total":
            return "", ""
        return self._division_key(display), display

    def _metric_triplet(
        self, row: dict[str, str], headers: list[str]
    ) -> dict[str, str]:
        header_map = self._header_lookup(headers)
        out: dict[str, str] = {}
        aliases = {
            "Received": ("received",),
            "% Share": ("% share", "share %", "share"),
            "Average Rating": ("average rating", "avg. rating", "avg rating"),
        }
        for metric, keys in aliases.items():
            col = None
            for key in keys:
                if key in header_map:
                    col = header_map[key]
                    break
            out[metric] = str(row.get(col, "")).strip() if col else ""
        return out

    def _metric_from_total(
        self, total_row: dict[str, str] | None, headers: list[str], metric: str
    ) -> str:
        if not total_row:
            return ""
        return self._metric_triplet(total_row, headers).get(metric, "")

    def _merge_by_division(
        self,
        prev_rows: list[dict[str, str]],
        prev_headers: list[str],
        up_rows: list[dict[str, str]],
        up_headers: list[str],
        *,
        prev_total: dict[str, str] | None,
        up_total: dict[str, str] | None,
    ) -> tuple[list[str], list[list[str]]]:
        prev_by_key: dict[str, dict[str, str]] = {}
        prev_display: dict[str, str] = {}
        for row in prev_rows:
            key, display = self._division_for_row(row, prev_headers)
            if not key:
                continue
            if key in prev_by_key:
                raise ValueError(f"Duplicate Division in Previous extract: {display}")
            prev_by_key[key] = self._metric_triplet(row, prev_headers)
            prev_display[key] = display

        up_by_key: dict[str, dict[str, str]] = {}
        up_display: dict[str, str] = {}
        for row in up_rows:
            key, display = self._division_for_row(row, up_headers)
            if not key:
                continue
            if key in up_by_key:
                raise ValueError(f"Duplicate Division in Upcoming extract: {display}")
            up_by_key[key] = self._metric_triplet(row, up_headers)
            up_display[key] = display

        all_keys = list(dict.fromkeys([*prev_by_key.keys(), *up_by_key.keys()]))
        if not all_keys:
            raise ValueError("No Division rows found in Previous or Upcoming extracts")

        all_keys.sort(
            key=lambda k: (
                -max(
                    _parse_num(prev_by_key.get(k, {}).get("Received", "0")),
                    _parse_num(up_by_key.get(k, {}).get("Received", "0")),
                ),
                k,
            )
        )

        blank = {m: "" for m in _MERGE_METRICS}
        data_rows: list[list[str]] = []
        seen_divisions: set[str] = set()
        for idx, key in enumerate(all_keys, start=1):
            if key in seen_divisions:
                raise ValueError(f"Duplicate Division after merge: {key}")
            seen_divisions.add(key)
            division = self._pick_display_division(
                prev_display.get(key, ""), up_display.get(key, "")
            )
            prev = prev_by_key.get(key, blank)
            up = up_by_key.get(key, blank)
            data_rows.append(
                [
                    str(idx),
                    division,
                    prev.get("Received", ""),
                    prev.get("% Share", ""),
                    prev.get("Average Rating", ""),
                    up.get("Received", ""),
                    up.get("% Share", ""),
                    up.get("Average Rating", ""),
                ]
            )

        self._validate_merged_rows(data_rows)

        total = [""] * len(OUTPUT_HEADERS)
        total[1] = "Total"
        total[2] = _format_num(sum(_parse_num(r[2]) for r in data_rows))
        total[3] = "100" if _parse_num(total[2]) > 0 else ""
        total[4] = self._metric_from_total(prev_total, prev_headers, "Average Rating")
        total[5] = _format_num(sum(_parse_num(r[5]) for r in data_rows))
        total[6] = "100" if _parse_num(total[5]) > 0 else ""
        total[7] = self._metric_from_total(up_total, up_headers, "Average Rating")
        data_rows.append(total)
        return list(OUTPUT_HEADERS), data_rows

    @staticmethod
    def _validate_merged_rows(rows: list[list[str]]) -> None:
        divisions: set[str] = set()
        for row in rows:
            division = str(row[1] if len(row) > 1 else "").strip()
            key = Report14Processor._division_key(division)
            if not key:
                continue
            if key in divisions:
                raise ValueError(f"Merged output has duplicate Division: {division}")
            divisions.add(key)

    @staticmethod
    def _is_total_row(row: list[str]) -> bool:
        return any(str(c).strip().lower() == "total" for c in row)

    def _write_excel(
        self,
        target_path: Path,
        headers: list[str],
        rows: list[list[str]],
        *,
        main_title: str,
        subtitle: str,
        report_date: str = "",
    ) -> None:
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = REPORT14_SHEET_TITLE

        col_count = max(len(headers), len(OUTPUT_HEADERS))
        main_title = normalize_report_title(main_title, report_slug="report14")

        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count - 1)
        title_cell = worksheet.cell(row=1, column=1, value=main_title)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        if report_date:
            date_cell = worksheet.cell(row=1, column=col_count, value=report_date)
            date_cell.font = Font(bold=True, size=10)
            date_cell.alignment = Alignment(horizontal="right", vertical="center")

        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
        sub_cell = worksheet.cell(
            row=2,
            column=1,
            value=normalize_report_title(subtitle, report_slug="report14"),
        )
        sub_cell.font = Font(bold=True, size=10)
        sub_cell.alignment = Alignment(horizontal="center")

        header_row_group = 3
        header_row_cols = 4

        worksheet.merge_cells(
            start_row=header_row_group,
            start_column=1,
            end_row=header_row_cols,
            end_column=1,
        )
        worksheet.merge_cells(
            start_row=header_row_group,
            start_column=2,
            end_row=header_row_cols,
            end_column=2,
        )
        sno_cell = worksheet.cell(row=header_row_group, column=1, value="S.No.")
        div_cell = worksheet.cell(row=header_row_group, column=2, value="Division")
        for cell in (sno_cell, div_cell):
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        worksheet.merge_cells(
            start_row=header_row_group,
            start_column=3,
            end_row=header_row_group,
            end_column=5,
        )
        worksheet.merge_cells(
            start_row=header_row_group,
            start_column=6,
            end_row=header_row_group,
            end_column=8,
        )
        prev_group = worksheet.cell(row=header_row_group, column=3, value=REPORT14_PREV_GROUP_TITLE)
        up_group = worksheet.cell(row=header_row_group, column=6, value=REPORT14_UP_GROUP_TITLE)
        for cell in (prev_group, up_group):
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        sub_headers = ["S.No.", "Division", *REPORT14_SUB_HEADERS, *REPORT14_SUB_HEADERS]
        for col_idx, label in enumerate(sub_headers[:col_count], start=1):
            if col_idx <= 2:
                continue
            cell = worksheet.cell(row=header_row_cols, column=col_idx, value=label)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        data_start = header_row_cols + 1
        for row_idx, row_values in enumerate(rows, start=data_start):
            is_total = self._is_total_row(row_values)
            for col_idx, value in enumerate(row_values[:col_count], start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(
                    horizontal="left" if col_idx == 2 else "center",
                    wrap_text=True,
                )
                if is_total:
                    cell.font = Font(bold=True)
                    cell.fill = TOTAL_FILL

        worksheet.column_dimensions["A"].width = 6
        worksheet.column_dimensions["B"].width = 28
        for col_idx in range(3, col_count + 1):
            letter = worksheet.cell(row=header_row_cols, column=col_idx).column_letter
            worksheet.column_dimensions[letter].width = 14

        apply_uniform_center_alignment(worksheet)

        workbook.save(temp_path)
        temp_path.replace(target_path)

    def _write_pdf(
        self,
        target_path: Path,
        headers: list[str],
        rows: list[list[str]],
        *,
        main_title: str,
        subtitle: str,
        report_date: str = "",
    ) -> None:
        ensure_pdf_unicode_fonts()
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        margin = _PDF_MARGIN_PT
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Report14Title",
            parent=styles["Heading1"],
            fontSize=12,
            leading=14,
            alignment=TA_CENTER,
            spaceAfter=3,
            fontName=pdf_font_bold(),
        )
        subtitle_style = ParagraphStyle(
            "Report14Subtitle",
            parent=styles["Normal"],
            fontSize=9,
            leading=11,
            alignment=TA_CENTER,
            spaceAfter=6,
            fontName=pdf_font_regular(),
        )
        cell_style = ParagraphStyle(
            "Report14Cell",
            parent=styles["Normal"],
            fontSize=7,
            leading=8,
            fontName=pdf_font_regular(),
            alignment=TA_CENTER,
        )
        header_style = ParagraphStyle(
            "Report14Header",
            parent=styles["Normal"],
            fontSize=7,
            leading=8,
            fontName=pdf_font_bold(),
            alignment=TA_CENTER,
        )

        page_width, _ = landscape(A4)
        usable = page_width - 2 * margin
        n_cols = max(len(headers), len(OUTPUT_HEADERS))
        raw_widths = [24.0, 72.0] + [40.0] * max(n_cols - 2, 0)
        col_widths = fit_column_widths(raw_widths[:n_cols], usable)

        def _p(text: str, style: ParagraphStyle) -> Paragraph:
            return Paragraph(_escape_paragraph_xml(text), style)

        group_row: list[Any] = [
            _p("S.No.", header_style),
            _p("Division", header_style),
            _p(REPORT14_PREV_GROUP_TITLE, header_style),
            "",
            "",
            _p(REPORT14_UP_GROUP_TITLE, header_style),
            "",
            "",
        ]
        sub_row: list[Any] = [
            _p("S.No.", header_style),
            _p("Division", header_style),
            *[_p(h, header_style) for h in REPORT14_SUB_HEADERS],
            *[_p(h, header_style) for h in REPORT14_SUB_HEADERS],
        ]
        table_data: list[list[Any]] = [group_row[:n_cols], sub_row[:n_cols]]
        for row in rows:
            cells = [_p(str(val), cell_style) for val in row[:n_cols]]
            while len(cells) < n_cols:
                cells.append(_p("", cell_style))
            table_data.append(cells)

        table = LongTable(table_data, colWidths=col_widths, repeatRows=2)
        style_cmds: list[tuple] = [
            ("BACKGROUND", (0, 0), (-1, 1), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 1), pdf_font_bold()),
            ("SPAN", (0, 0), (0, 1)),
            ("SPAN", (1, 0), (1, 1)),
            ("SPAN", (2, 0), (4, 0)),
            ("SPAN", (5, 0), (7, 0)),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#666666")),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]
        for r_idx, row in enumerate(rows, start=2):
            if self._is_total_row(row):
                style_cmds.append(
                    ("BACKGROUND", (0, r_idx), (-1, r_idx), colors.HexColor("#D9D9D9"))
                )
                style_cmds.append(("FONTNAME", (0, r_idx), (-1, r_idx), pdf_font_bold()))
        table.setStyle(TableStyle(style_cmds))

        doc = SimpleDocTemplate(
            str(temp_path),
            pagesize=landscape(A4),
            leftMargin=margin,
            rightMargin=margin,
            topMargin=margin,
            bottomMargin=margin,
        )
        title_line = main_title
        if report_date:
            title_line = f"{main_title} — {report_date}"
        story = [
            Paragraph(_escape_paragraph_xml(title_line), title_style),
            Paragraph(_escape_paragraph_xml(subtitle), subtitle_style),
            Spacer(1, 4),
            table,
        ]
        doc.build(story)
        temp_path.replace(target_path)
