"""Report 10-13 post-ingestion processor (Comprehensive Reports).

Processes four sections extracted from the Comprehensive (with drill down) page:
- Report 10: C&W complaints division wise
- Report 11: Security complaints
- Report 12: Punctuality complaints
- Report 13: Electrical Equipment complaints division wise

Each section produces a stacked table with:
- Section heading
- Date range
- Selected columns
- Data rows sorted by Received descending
- Total row

Output: One combined XLSX and one combined PDF.
"""

from __future__ import annotations

import csv
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.automation.config import config
from app.automation.comprehensive1013_filters import (
    COMPREHENSIVE_1013_SECTION_IDS,
    SectionConfig,
    get_section_config_by_id,
)
from app.automation.date_range import date_range_for_processing
from app.automation.formatting.excel_print import apply_uniform_center_alignment
from app.automation.formatting.pdf_fonts import pdf_font_bold, pdf_font_regular
from app.automation.formatting.pdf_table import (
    MAX_FONT_SIZE,
    SAFE_MARGIN_PT,
    fit_column_widths,
    preferred_column_widths,
)
from app.automation.formatting.text_pipeline import normalize_report_title
from app.automation.processing.base import ProcessingResult
from app.automation.processing.comprehensive_output_columns import (
    ADDITIVE_COLUMNS,
    COMPREHENSIVE_COLUMN_IDS,
    NON_ADDITIVE_COLUMNS,
    column_labels,
    default_column_ids,
    normalize_header_to_column_id,
    sanitize_comprehensive_section_columns,
    sanitize_comprehensive_sections,
)
from app.automation.utils import (
    ensure_directory,
    log_automation_event,
    resolve_run_scoped_dir,
)

_SECTION_VALIDATION_NAMES: dict[str, str] = {
    "report10_cw": "Report 10 — C&W",
    "report11_security": "Report 11 — Security",
    "report12_punctuality": "Report 12 — Punctuality",
    "report13_electrical": "Report 13 — Electrical Equipment",
}

# Report 10-13 PDF only: compact margins/spacing so four sections fit one page.
_PDF_MARGIN_PT = min(SAFE_MARGIN_PT, 12.0)
_PDF_SECTION_GAP_PT = 8.0
_PDF_AFTER_HEADING_PT = 3.0
_PDF_TITLE_AFTER_PT = 5.0
_PDF_HEIGHT_BUFFER_PT = 4.0
_PDF_MAIN_TITLE = "Comprehensive Reports"
_DIVISION_RAW_HEADER_PRIORITY = ("Division", "Organisation")


def _headers_for_column_ids(raw_headers: list[str]) -> dict[str, list[str]]:
    """Map canonical column IDs to all raw CSV headers that project to them."""
    mapping: dict[str, list[str]] = {}
    for header in raw_headers:
        col_id = normalize_header_to_column_id(header)
        if col_id:
            mapping.setdefault(col_id, []).append(header)
    return mapping


def _division_header_sort_key(header: str) -> tuple[int, str]:
    if header in _DIVISION_RAW_HEADER_PRIORITY:
        return (_DIVISION_RAW_HEADER_PRIORITY.index(header), header)
    return (len(_DIVISION_RAW_HEADER_PRIORITY), header)


def _raw_cell_value(row: dict[str, str], raw_header: str) -> str:
    return str(row.get(raw_header, "") or "").strip()


def _value_for_column_id(
    row: dict[str, str],
    col_id: str,
    header_map: dict[str, list[str]],
) -> str:
    """Resolve a projected cell value, preferring non-empty matches."""
    raw_headers = header_map.get(col_id, [])
    if col_id == "division":
        ordered = sorted(raw_headers, key=_division_header_sort_key)
        for header in ordered:
            value = _raw_cell_value(row, header)
            if value:
                return value
        if ordered:
            return _raw_cell_value(row, ordered[0])
        return ""
    for header in raw_headers:
        value = _raw_cell_value(row, header)
        if value:
            return value
    return ""


def _col_widths_for_section_headers(
    section_headers: list[str],
    shared_headers: list[str],
    shared_widths: list[float],
) -> list[float]:
    """Map shared width vector onto a section's header order."""
    by_header = dict(zip(shared_headers, shared_widths))
    fallback = (sum(shared_widths) / len(shared_widths)) if shared_widths else 40.0
    return [by_header.get(header, fallback) for header in section_headers]


def _escape_paragraph_xml(text: str) -> str:
    """Escape XML special characters for reportlab Paragraph markup."""
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _section_date_text(date_range: Any) -> str:
    """Display date for PDF section heading rows (does not change date resolution)."""
    if date_range.date_from == date_range.date_to:
        return date_range.display_from()
    return f"{date_range.display_from()} to {date_range.display_to()}"


def _render_section_heading(
    title: str,
    date_text: str,
    table_width: float,
    style: ParagraphStyle,
) -> Table:
    """Centered section title with right-aligned date on the same row over table_width."""
    centered_style = ParagraphStyle(
        name=f"{style.name}Centered",
        parent=style,
        alignment=TA_CENTER,
        fontName=pdf_font_bold(),
    )
    date_style = ParagraphStyle(
        name=f"{style.name}Date",
        parent=style,
        alignment=TA_RIGHT,
        fontName=pdf_font_bold(),
    )
    side_width = max(
        stringWidth(date_text, pdf_font_bold(), style.fontSize) + 8.0,
        56.0,
    )
    side_width = min(side_width, table_width * 0.28)
    middle_width = max(table_width - (2.0 * side_width), 1.0)

    heading_table = Table(
        [
            [
                "",
                Paragraph(_escape_paragraph_xml(title), centered_style),
                Paragraph(_escape_paragraph_xml(date_text), date_style),
            ]
        ],
        colWidths=[side_width, middle_width, side_width],
        hAlign="CENTER",
    )
    heading_table.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (0, 0), "LEFT"),
                ("ALIGN", (1, 0), (1, 0), "CENTER"),
                ("ALIGN", (2, 0), (2, 0), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    return heading_table


def _column_align_commands(headers: list[str]) -> list[tuple]:
    """ALIGN commands: every column (including Division/Zone/Train names) centered."""
    return [("ALIGN", (col_idx, 0), (col_idx, -1), "CENTER") for col_idx in range(len(headers))]


def _build_section_table(
    headers: list[str],
    rows: list[list[str]],
    total_row: list[str],
    col_widths: list[float],
    font_size: float,
) -> Table:
    """Build one section data table with shared widths and Excel-like styling."""
    table_data: list[list[object]] = [list(headers)]
    for row_values in rows:
        table_data.append(list(row_values))
    table_data.append(list(total_row))

    style_commands: list[tuple] = [
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("BACKGROUND", (0, -1), (-1, -1), colors.Color(0.85, 0.85, 0.85)),
        ("FONTNAME", (0, 0), (-1, 0), pdf_font_bold()),
        ("FONTNAME", (0, 1), (-1, -2), pdf_font_regular()),
        ("FONTNAME", (0, -1), (-1, -1), pdf_font_bold()),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        *_column_align_commands(headers),
    ]
    table = Table(table_data, colWidths=list(col_widths), repeatRows=1)
    table.setStyle(TableStyle(style_commands))
    table.hAlign = "CENTER"
    return table


def _render_pdf_section(
    section: SectionDataset,
    *,
    section_date: str,
    col_widths: list[float],
    font_size: float,
    heading_style: ParagraphStyle,
) -> list:
    """Render heading + date, spacer, and table for one comprehensive section."""
    table = _build_section_table(
        section.headers,
        section.rows,
        section.total_row,
        col_widths,
        font_size,
    )
    table_width = float(sum(col_widths))
    heading = _render_section_heading(
        section.section_config.section_title,
        section_date,
        table_width,
        heading_style,
    )
    return [heading, Spacer(1, _PDF_AFTER_HEADING_PT), table]


def _shared_pdf_column_layout(
    sections: list[SectionDataset],
    *,
    margin: float,
) -> tuple[tuple[float, float], list[float], float]:
    """
    Compute one pagesize + identical col_widths + font_size for all sections.

    Uses the max preferred width per column across sections that share the same
    header set (typical case). Tables always span the same usable page width.
    """
    base_pagesize = landscape(A3)
    page_width = base_pagesize[0]
    usable = page_width - (2 * margin)
    font_size = MAX_FONT_SIZE

    # Prefer the widest common header layout; fall back to first section headers.
    header_key_counts: dict[tuple[str, ...], int] = {}
    for section in sections:
        key = tuple(section.headers)
        header_key_counts[key] = header_key_counts.get(key, 0) + 1
    shared_headers = list(
        max(header_key_counts.keys(), key=lambda key: (header_key_counts[key], len(key)))
    )

    combined_rows: list[list[object]] = [list(shared_headers)]
    for section in sections:
        if list(section.headers) != shared_headers:
            continue
        combined_rows.extend(list(row) for row in section.rows)
        combined_rows.append(list(section.total_row))

    preferred = preferred_column_widths(
        combined_rows,
        font_size=font_size,
        headers=shared_headers,
    )
    col_widths = fit_column_widths(preferred, usable)
    return base_pagesize, col_widths, font_size

logger = logging.getLogger(__name__)

PROCESSOR_NAME = "comprehensive1013_processor"

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

TOTAL_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


@dataclass
class SectionDataset:
    """Data for a single comprehensive report section."""

    section_config: SectionConfig
    headers: list[str]
    rows: list[list[str]]
    total_row: list[str]
    raw_headers: list[str]
    column_ids: list[str]


class Comprehensive1013Processor:
    """Process Comprehensive Reports 10-13 and emit combined Excel/PDF."""

    processor_name = PROCESSOR_NAME

    def process(
        self,
        *,
        source_a_path: Path,
        report_slug: str,
        source_b_path: Path | None = None,
        column_selection: dict[str, Any] | None = None,
    ) -> ProcessingResult:
        if source_a_path.suffix.lower() == ".pdf":
            return ProcessingResult(success=False, error="PDF cannot be used as processing input")

        column_selection = self._resolve_effective_column_selection(column_selection)
        log_automation_event(
            logger,
            "comprehensive1013_column_selection_received",
            run_id=column_selection.get("run_id"),
            configuration_source=column_selection.get("configuration_source"),
            snapshot_hash=column_selection.get("snapshot_hash"),
            sections={
                section_id: list(
                    (column_selection.get("sections") or {})
                    .get(section_id, {})
                    .get("selected_column_ids")
                    or []
                )
                for section_id in COMPREHENSIVE_1013_SECTION_IDS
            },
        )

        try:
            sections, total_input_rows = self._load_sections(source_a_path, column_selection)
        except ValueError as exc:
            return ProcessingResult(
                success=False,
                error=str(exc),
                source_a_path=str(source_a_path),
            )

        if not sections:
            return ProcessingResult(
                success=False,
                error="No section data found in combined index",
                source_a_path=str(source_a_path),
            )

        date_range = date_range_for_processing(column_selection)
        report_date = date_range.title_suffix()
        section_date = _section_date_text(date_range)
        filename_suffix = date_range.filename_suffix()

        run_id = column_selection.get("run_id")
        if run_id:
            excel_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_excel_dir, report_slug, str(run_id))
            )
            pdf_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_pdf_dir, report_slug, str(run_id))
            )
        else:
            parent = source_a_path.parent
            scope = parent.name if parent.name else "shared"
            excel_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_excel_dir, report_slug, scope)
            )
            pdf_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_pdf_dir, report_slug, scope)
            )
        base_name = f"Rail_Madad_Report_10_13_Comprehensive_{filename_suffix}"
        excel_path = excel_dir / f"{base_name}.xlsx"
        pdf_path = pdf_dir / f"{base_name}.pdf"

        try:
            self._write_excel(excel_path, sections, report_date=report_date)
            log_automation_event(logger, "comprehensive1013_excel_generated", excel_path=str(excel_path))
            self._write_pdf(pdf_path, sections, section_date=section_date)
            log_automation_event(logger, "comprehensive1013_pdf_generated", pdf_path=str(pdf_path))
        except Exception as exc:
            return ProcessingResult(
                input_row_count=total_input_rows,
                success=False,
                error=str(exc),
                source_a_path=str(source_a_path),
                source_a_rows=total_input_rows,
            )

        total_output_rows = sum(len(s.rows) for s in sections)

        section_row_counts = {s.section_config.section_id: len(s.rows) for s in sections}

        from app.automation.processing.comprehensive_output_columns import (
            build_comprehensive_artifact_metadata,
            build_comprehensive_column_snapshot,
        )

        # Snapshot must come from the resolved filter config (all four sections),
        # not only from sections that extracted successfully. Otherwise a single
        # failed portal section aborts PDF/Excel even when other sections are ready.
        selection_sections = (column_selection or {}).get("sections")
        if isinstance(selection_sections, dict) and len(selection_sections) == len(
            COMPREHENSIVE_1013_SECTION_IDS
        ):
            sections_for_snapshot = selection_sections
        else:
            sections_for_snapshot = {
                section_id: {
                    "selected_column_ids": list(
                        next(
                            (
                                s.column_ids
                                for s in sections
                                if s.section_config.section_id == section_id
                            ),
                            default_column_ids(),
                        )
                    )
                }
                for section_id in COMPREHENSIVE_1013_SECTION_IDS
            }

        date_from = (column_selection or {}).get("date_from")
        date_to = (column_selection or {}).get("date_to")
        try:
            column_snapshot = build_comprehensive_column_snapshot(
                sections_for_snapshot,
                date_from=str(date_from) if date_from else None,
                date_to=str(date_to) if date_to else None,
                configuration_source=str(
                    (column_selection or {}).get("configuration_source") or "manual_snapshot"
                ),
            )
        except ValueError as exc:
            return ProcessingResult(
                success=False,
                error=str(exc),
                input_row_count=total_input_rows,
                processed_row_count=total_output_rows,
                excel_path=str(excel_path),
                pdf_path=str(pdf_path),
                source_a_path=str(source_a_path),
                source_a_rows=total_input_rows,
            )
        union_column_ids = column_snapshot["selected_column_ids"]
        artifact_metadata = build_comprehensive_artifact_metadata(column_snapshot)
        # Record which sections were actually rendered (extraction may be partial).
        artifact_metadata["rendered_section_ids"] = [
            s.section_config.section_id for s in sections
        ]

        log_automation_event(
            logger,
            "comprehensive1013_processing_completed",
            source_a=str(source_a_path),
            input_row_count=total_input_rows,
            section_count=len(sections),
            total_output_rows=total_output_rows,
            section_row_counts=section_row_counts,
            snapshot_hash=column_snapshot.get("snapshot_hash"),
            configuration_source=column_selection.get("configuration_source"),
            rendered_section_ids=artifact_metadata["rendered_section_ids"],
        )

        return ProcessingResult(
            success=True,
            input_row_count=total_input_rows,
            processed_row_count=total_output_rows,
            excel_path=str(excel_path),
            pdf_path=str(pdf_path),
            source_a_path=str(source_a_path),
            source_a_rows=total_input_rows,
            output_columns=column_labels(union_column_ids),
            visible_columns=column_labels(union_column_ids),
            selected_column_ids=union_column_ids,
            column_order=list(union_column_ids),
            configuration_source=str(
                (column_selection or {}).get("configuration_source") or "manual_snapshot"
            ),
            artifact_metadata=artifact_metadata,
        )

    def _load_sections(
        self,
        source_a_path: Path,
        column_selection: dict[str, Any] | None,
    ) -> tuple[list[SectionDataset], int]:
        """Load all sections from the combined index and their CSVs."""
        sections: list[SectionDataset] = []
        total_input_rows = 0

        index_entries = self._read_combined_index(source_a_path)
        if not index_entries:
            base_dir = source_a_path.parent
            for section_id in COMPREHENSIVE_1013_SECTION_IDS:
                section_config = get_section_config_by_id(section_id)
                if section_config is None:
                    continue
                csv_path = base_dir / f"{section_id}.csv"
                if csv_path.is_file():
                    index_entries[section_id] = {
                        "section_id": section_id,
                        "csv_path": str(csv_path),
                        "status": "success",
                    }

        for section_id in COMPREHENSIVE_1013_SECTION_IDS:
            section_config = get_section_config_by_id(section_id)
            if section_config is None:
                continue

            entry = index_entries.get(section_id)
            if entry is None or str(entry.get("status", "")).lower() != "success":
                log_automation_event(
                    logger,
                    "comprehensive1013_section_skipped",
                    section_id=section_id,
                    reason="not in index or failed",
                )
                continue

            csv_path = Path(str(entry.get("csv_path") or ""))
            if not csv_path.is_file():
                log_automation_event(
                    logger,
                    "comprehensive1013_section_csv_not_found",
                    section_id=section_id,
                    expected_path=str(csv_path),
                )
                continue

            raw_rows, raw_headers = self._read_csv(csv_path)
            if not raw_rows:
                continue

            data_rows, portal_total_row = self._split_total_row(raw_rows)
            total_input_rows += len(data_rows)

            selected_ids = self._resolve_column_ids(section_id, column_selection)
            projected_headers, projected_rows = self._project_columns(
                raw_headers, data_rows, selected_ids
            )
            total_row = self._compute_total_row(
                projected_headers, projected_rows, portal_total_row, raw_headers, selected_ids
            )

            default_ids = default_column_ids()
            log_automation_event(
                logger,
                "comprehensive1013_section_column_filter",
                run_id=(column_selection or {}).get("run_id"),
                section_id=section_id,
                selected_column_ids=selected_ids,
                renderer_column_keys=selected_ids,
                missing_vs_default=[col for col in default_ids if col not in selected_ids],
                unexpected_columns=[
                    col for col in selected_ids if col not in COMPREHENSIVE_COLUMN_IDS
                ],
            )

            sections.append(
                SectionDataset(
                    section_config=section_config,
                    headers=projected_headers,
                    rows=projected_rows,
                    total_row=total_row,
                    raw_headers=raw_headers,
                    column_ids=selected_ids,
                )
            )
            log_automation_event(
                logger,
                "comprehensive1013_section_loaded",
                section_id=section_id,
                row_count=len(projected_rows),
                selected_columns=selected_ids,
            )

        return sections, total_input_rows

    def _read_combined_index(self, source_a_path: Path) -> dict[str, dict[str, str]]:
        """Read the combined index CSV if it exists."""
        if source_a_path.name != "comprehensive_combined_index.csv":
            return {}
        if not source_a_path.is_file():
            return {}
        entries: dict[str, dict[str, str]] = {}
        with source_a_path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                section_id = (row.get("section_id") or "").strip()
                if not section_id:
                    continue
                entries[section_id] = {
                    "section_id": section_id,
                    "section_name": (row.get("section_name") or "").strip(),
                    "csv_path": (row.get("csv_path") or "").strip(),
                    "row_count": (row.get("row_count") or "0").strip(),
                    "status": (row.get("status") or "").strip(),
                    "error": (row.get("error") or "").strip(),
                }
        return entries

    @staticmethod
    def _read_csv(path: Path) -> tuple[list[dict[str, str]], list[str]]:
        """Read a CSV file and return rows as dicts plus header list."""
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            headers = list(reader.fieldnames or [])
            rows = [{header: row.get(header, "") for header in headers} for row in reader]
        return rows, headers

    @staticmethod
    def _split_total_row(
        rows: list[dict[str, str]],
    ) -> tuple[list[dict[str, str]], dict[str, str] | None]:
        """Split data rows from total row (last row with 'Total' in Division column)."""
        if not rows:
            return [], None
        last_row = rows[-1]
        division_val = str(last_row.get("Division") or last_row.get("Organisation") or "").strip().lower()
        if division_val == "total":
            return rows[:-1], last_row
        return rows, None

    def _resolve_effective_column_selection(
        self,
        column_selection: dict[str, Any] | None,
    ) -> dict[str, Any]:
        """Normalize selection: run snapshot → saved config → legacy union → defaults."""
        selection = dict(column_selection or {})

        try:
            from app.automation.run_context import get_run_context

            ctx = get_run_context()
            if ctx is not None:
                if not selection.get("run_id") and ctx.run_id:
                    selection["run_id"] = ctx.run_id
                user_id = ctx.user_id
            else:
                user_id = None
        except Exception:
            user_id = None

        sections_raw = selection.get("sections")
        if isinstance(sections_raw, dict) and sections_raw:
            sanitized = sanitize_comprehensive_sections(sections_raw)
            if len(sanitized) == len(COMPREHENSIVE_1013_SECTION_IDS):
                selection["sections"] = sanitized
                selection.setdefault("configuration_source", "manual_snapshot")
                self._log_column_selection_diagnostics(selection, source="run_snapshot")
                return selection

        from app.features.reports.config_store import load_report_config

        saved = load_report_config("comprehensive-10-13", user_id=user_id) or {}
        saved_sections = saved.get("sections")
        if isinstance(saved_sections, dict) and saved_sections:
            sanitized = sanitize_comprehensive_sections(saved_sections)
            if len(sanitized) == len(COMPREHENSIVE_1013_SECTION_IDS):
                selection["sections"] = sanitized
                selection["configuration_source"] = "saved_user_config"
                if not selection.get("selected_column_ids") and saved.get("selected_column_ids"):
                    selection["selected_column_ids"] = list(saved["selected_column_ids"])
                if not selection.get("snapshot_hash") and saved.get("snapshot_hash"):
                    selection["snapshot_hash"] = saved["snapshot_hash"]
                self._log_column_selection_diagnostics(selection, source="saved_user_config")
                return selection

        union_ids = selection.get("selected_column_ids") or saved.get("selected_column_ids")
        sanitized_union = sanitize_comprehensive_section_columns(union_ids or [])
        if sanitized_union:
            selection["sections"] = {
                section_id: {"selected_column_ids": list(sanitized_union)}
                for section_id in COMPREHENSIVE_1013_SECTION_IDS
            }
            selection["configuration_source"] = (
                selection.get("configuration_source") or "legacy_union"
            )
            self._log_column_selection_diagnostics(selection, source="legacy_union")
            return selection

        defaults = default_column_ids()
        selection["sections"] = {
            section_id: {"selected_column_ids": list(defaults)}
            for section_id in COMPREHENSIVE_1013_SECTION_IDS
        }
        selection["configuration_source"] = "report_default"
        self._log_column_selection_diagnostics(selection, source="report_default")
        return selection

    @staticmethod
    def _log_column_selection_diagnostics(
        selection: dict[str, Any],
        *,
        source: str,
    ) -> None:
        sections = selection.get("sections") or {}
        for section_id in COMPREHENSIVE_1013_SECTION_IDS:
            selected = list((sections.get(section_id) or {}).get("selected_column_ids") or [])
            log_automation_event(
                logger,
                "comprehensive1013_column_selection_resolved",
                run_id=selection.get("run_id"),
                section_id=section_id,
                configuration_source=source,
                selected_column_ids=selected,
                selected_column_count=len(selected),
                missing_vs_default=[col for col in default_column_ids() if col not in selected],
            )

    def _resolve_column_ids(
        self,
        section_id: str,
        column_selection: dict[str, Any] | None,
    ) -> list[str]:
        """Resolve which columns to include for a section."""
        section_name = _SECTION_VALIDATION_NAMES.get(section_id, section_id)
        if column_selection and isinstance(column_selection.get("sections"), dict):
            sections = column_selection["sections"]
            section_columns = sections.get(section_id)
            if not isinstance(section_columns, dict):
                raise ValueError(f"Select at least one column for {section_name}.")
            selected = sanitize_comprehensive_section_columns(
                section_columns.get("selected_column_ids") or []
            )
            if not selected:
                raise ValueError(f"Select at least one column for {section_name}.")
            return selected

        if column_selection:
            all_selected = sanitize_comprehensive_section_columns(
                column_selection.get("selected_column_ids") or []
            )
            if all_selected:
                return all_selected
        return default_column_ids()

    def _project_columns(
        self,
        raw_headers: list[str],
        data_rows: list[dict[str, str]],
        selected_ids: list[str],
    ) -> tuple[list[str], list[list[str]]]:
        """Project rows to only selected columns, returning headers and row lists."""
        header_map = _headers_for_column_ids(raw_headers)
        output_headers = column_labels(selected_ids)
        output_rows: list[list[str]] = []

        for row_idx, row in enumerate(data_rows):
            output_row: list[str] = []
            for col_id in selected_ids:
                if col_id == "sno":
                    output_row.append(str(row_idx + 1))
                else:
                    output_row.append(_value_for_column_id(row, col_id, header_map))
            output_rows.append(output_row)

        return output_headers, output_rows

    def _compute_total_row(
        self,
        projected_headers: list[str],
        projected_rows: list[list[str]],
        portal_total_row: dict[str, str] | None,
        raw_headers: list[str],
        selected_ids: list[str],
    ) -> list[str]:
        """Compute total row for a section."""
        header_map = _headers_for_column_ids(raw_headers)

        total_row: list[str] = []
        for col_id in selected_ids:
            if col_id == "sno":
                total_row.append("")
            elif col_id == "division":
                total_row.append("Total")
            elif col_id == "share_percent":
                total_row.append("100.00")
            elif col_id in NON_ADDITIVE_COLUMNS:
                if portal_total_row:
                    total_row.append(_value_for_column_id(portal_total_row, col_id, header_map))
                else:
                    total_row.append("")
            elif col_id in ADDITIVE_COLUMNS:
                col_idx = selected_ids.index(col_id)
                col_sum = 0
                for row in projected_rows:
                    try:
                        val = row[col_idx].replace(",", "").strip()
                        if val:
                            col_sum += int(float(val))
                    except (ValueError, IndexError):
                        pass
                total_row.append(str(col_sum))
            else:
                total_row.append("")

        return total_row

    @staticmethod
    def _section_width(sections: list[SectionDataset]) -> int:
        """Return maximum column count across sections."""
        if not sections:
            return 1
        return max(len(section.headers) for section in sections)

    def _write_excel(
        self,
        target_path: Path,
        sections: list[SectionDataset],
        *,
        report_date: str,
    ) -> None:
        """Write combined Excel with all sections stacked vertically."""
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Comprehensive Reports 10-13"

        col_count = self._section_width(sections)
        main_title = normalize_report_title(
            f"Report 10-13 (Comprehensive Reports) {report_date}",
            report_slug="comprehensive-10-13",
        )
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(col_count, 1))
        title_cell = worksheet.cell(row=1, column=1, value=main_title)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        current_row = 2

        for section_idx, section in enumerate(sections):
            section_cols = max(len(section.headers), 1)
            worksheet.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=section_cols,
            )
            section_title_cell = worksheet.cell(
                row=current_row,
                column=1,
                value=section.section_config.section_title,
            )
            section_title_cell.font = Font(bold=True, size=11)
            section_title_cell.alignment = Alignment(horizontal="left")
            current_row += 1

            for col_idx, header in enumerate(section.headers, start=1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.border = THIN_BORDER
            current_row += 1

            for row_values in section.rows:
                for col_idx, value in enumerate(row_values, start=1):
                    cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                    cell.border = THIN_BORDER
                current_row += 1

            for col_idx, value in enumerate(section.total_row, start=1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                cell.font = Font(bold=True)
                cell.border = THIN_BORDER
                cell.fill = TOTAL_FILL
            current_row += 1

            if section_idx < len(sections) - 1:
                current_row += 1

        apply_uniform_center_alignment(worksheet)

        workbook.save(temp_path)
        temp_path.replace(target_path)

    def _write_pdf(
        self,
        target_path: Path,
        sections: list[SectionDataset],
        *,
        section_date: str,
    ) -> None:
        """Write all four sections stacked continuously on a single PDF page."""
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        margin = _PDF_MARGIN_PT

        styles = getSampleStyleSheet()
        section_style = ParagraphStyle(
            "ComprehensiveSection",
            parent=styles["Heading2"],
            fontSize=10,
            leading=12,
            spaceBefore=0,
            spaceAfter=0,
            textColor=colors.black,
            fontName=pdf_font_bold(),
        )
        empty_style = ParagraphStyle(
            "ComprehensiveEmpty",
            parent=styles["Normal"],
            fontSize=8,
            leading=10,
            alignment=TA_CENTER,
            spaceBefore=0,
            spaceAfter=0,
        )
        title_style = ParagraphStyle(
            "ComprehensiveTitle",
            parent=styles["Heading1"],
            fontSize=12,
            leading=14,
            alignment=TA_CENTER,
            spaceBefore=0,
            spaceAfter=0,
            fontName=pdf_font_bold(),
        )

        main_title = normalize_report_title(
            _PDF_MAIN_TITLE,
            report_slug="comprehensive-10-13",
        )

        data_sections = [section for section in sections if section.rows]
        base_pagesize, shared_col_widths, font_size = _shared_pdf_column_layout(
            data_sections if data_sections else sections,
            margin=margin,
        )
        page_width = base_pagesize[0]
        usable_width = page_width - (2 * margin)

        header_counts: dict[tuple[str, ...], int] = {}
        for section in data_sections:
            key = tuple(section.headers)
            header_counts[key] = header_counts.get(key, 0) + 1
        shared_headers = (
            list(max(header_counts.keys(), key=lambda key: (header_counts[key], len(key))))
            if header_counts
            else []
        )

        # One title only; no PageBreak / repeated titles between sections.
        story: list = [
            Paragraph(_escape_paragraph_xml(main_title), title_style),
            Spacer(1, _PDF_TITLE_AFTER_PT),
        ]

        for section_idx, section in enumerate(sections):
            if section_idx > 0:
                story.append(Spacer(1, _PDF_SECTION_GAP_PT))

            if not section.rows:
                story.append(Paragraph("No data available for this section.", empty_style))
                continue

            if list(section.headers) == shared_headers:
                col_widths = list(shared_col_widths)
            else:
                col_widths = _col_widths_for_section_headers(
                    list(section.headers),
                    shared_headers,
                    shared_col_widths,
                )

            story.extend(
                _render_pdf_section(
                    section,
                    section_date=section_date,
                    col_widths=col_widths,
                    font_size=font_size,
                    heading_style=section_style,
                )
            )

        content_height = 0.0
        for flowable in story:
            _width, height = flowable.wrap(usable_width, 100000)
            content_height += float(height)

        # Fit page to content (no forced A3 minimum); slight headroom avoids wrap under-estimate splits.
        page_height = (content_height * 1.04) + (2 * margin) + _PDF_HEIGHT_BUFFER_PT
        pagesize = (page_width, page_height)

        log_automation_event(
            logger,
            "comprehensive1013_pdf_single_page_layout",
            page_width=page_width,
            page_height=page_height,
            content_height=content_height,
            margin=margin,
            section_count=len(sections),
        )

        doc = SimpleDocTemplate(
            str(temp_path),
            pagesize=pagesize,
            leftMargin=margin,
            rightMargin=margin,
            topMargin=margin,
            bottomMargin=margin,
        )
        doc.build(story)
        temp_path.replace(target_path)
