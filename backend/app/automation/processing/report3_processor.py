"""Report 3 post-ingestion processor (Top 20 Trains)."""

from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.automation.config import config
from app.automation.formatting.excel_print import apply_uniform_center_alignment
from app.automation.formatting.pdf_fonts import ensure_pdf_unicode_fonts, pdf_font_bold, pdf_font_regular, pdf_title_style
from app.automation.formatting.pdf_table import (
    SAFE_MARGIN_PT,
    _cell_text,
    _escape_paragraph_text,
    _rescale_if_overflow,
    _table_style_commands,
)
from app.automation.formatting.scr import highlight_south_central_railway_rows, row_contains_scr
from app.automation.formatting.text_pipeline import (
    normalize_report_title,
    prepare_output_for_rendering,
    verify_text_rendering,
)
from app.automation.processing.base import ProcessingResult
from app.automation.processing.column_config import project_topn_for_output
from app.automation.processing.topn_output_columns import (
    build_canonical_topn_row,
    topn_default_ids,
    topn_labels,
)
from app.automation.date_range import date_range_for_processing
from app.automation.utils import (
    ensure_directory,
    log_automation_event,
    resolve_report_dir,
)

logger = logging.getLogger(__name__)

PROCESSOR_NAME = "report3_top20_trains_processor"

TOP_N = 20

DEFAULT_OUTPUT_HEADERS = topn_labels(topn_default_ids("train-no"), "train-no")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

OUTPUT_HEADERS = DEFAULT_OUTPUT_HEADERS

_PDF_MARGIN_PT = min(SAFE_MARGIN_PT, 12.0)
_PDF_FONT_SIZE = 8.0
_REPORT3_EXCEL_WIDTHS: dict[str, float] = {
    "S.No.": 8,
    "Train No.": 12,
    "Train Name": 36,
    "Owning Zone": 28,
    "Owning Division": 28,
    "Received": 10,
}
_REPORT3_COL_WEIGHTS: dict[str, float] = {
    "S.No.": 0.06,
    "Train No.": 0.10,
    "Train Name": 0.28,
    "Owning Zone": 0.22,
    "Owning Division": 0.22,
    "Received": 0.08,
}


def _allocate_report3_col_widths(headers: list[str], usable_width: float) -> list[float]:
    weights = [_REPORT3_COL_WEIGHTS.get(header, 0.15) for header in headers]
    total = sum(weights) or 1.0
    return [usable_width * weight / total for weight in weights]


def _prepare_report3_pdf_table_data(
    table_data: list[list[object]],
    headers: list[str],
    font_size: float,
) -> list[list[object]]:
    ensure_pdf_unicode_fonts()
    leading = font_size + 2.0
    cell_style = ParagraphStyle(
        "Report3PdfCell",
        fontName=pdf_font_regular(),
        fontSize=font_size,
        leading=leading,
        wordWrap="CJK",
        alignment=TA_CENTER,
    )
    header_style = ParagraphStyle(
        "Report3PdfHeader",
        fontName=pdf_font_bold(),
        fontSize=font_size,
        leading=leading,
        wordWrap="CJK",
        alignment=TA_CENTER,
    )
    wrapped: list[list[object]] = []
    for row_idx, row in enumerate(table_data):
        style = header_style if row_idx == 0 else cell_style
        wrapped.append(
            [
                Paragraph(_escape_paragraph_text(_cell_text(cell), field_kind="text"), style)
                for cell in row
            ]
        )
    return wrapped


class Report3Processor:
    """Process Top 20 Trains dataset and emit formatted Excel/PDF."""

    processor_name = PROCESSOR_NAME

    def process(
        self,
        *,
        source_a_path: Path,
        report_slug: str,
        source_b_path: Path | None = None,
        column_selection: dict[str, Any] | None = None,
    ) -> ProcessingResult:
        if source_a_path.suffix.lower() == ".pdf":
            return ProcessingResult(success=False, error="PDF cannot be used as processing input")

        if column_selection:
            log_automation_event(
                logger,
                "report3_filter_payload_received",
                selected_column_ids=column_selection.get("selected_column_ids"),
                column_order=column_selection.get("column_order"),
            )

        source_rows, _source_headers = self._read_csv(source_a_path)
        data_rows, _ = self._split_total_row(source_rows)
        top_n_rows = data_rows[:TOP_N]
        canonical_rows = self.build_canonical_rows(top_n_rows)

        try:
            output_headers, output_rows, visible_columns, keys, config_source = (
                project_topn_for_output(
                    report_slug,
                    canonical_rows,
                    column_selection=column_selection,
                )
            )
        except ValueError as exc:
            return ProcessingResult(
                input_row_count=len(data_rows),
                success=False,
                error=str(exc),
                source_a_path=str(source_a_path),
                source_a_rows=len(data_rows),
            )

        log_automation_event(
            logger,
            "report3_projection_completed",
            output_headers=output_headers,
            output_row_count=len(output_rows),
            selected_column_count=len(keys),
            configuration_source=config_source,
        )

        output_headers, output_rows = prepare_output_for_rendering(
            report_slug,
            output_headers,
            output_rows,
        )

        date_range = date_range_for_processing(column_selection)


        report_date = date_range.title_suffix()


        filename_suffix = date_range.filename_suffix()
        excel_dir = ensure_directory(resolve_report_dir(config.output_excel_dir, report_slug))
        pdf_dir = ensure_directory(resolve_report_dir(config.output_pdf_dir, report_slug))
        base_name = f"Rail_Madad_Report_3_Top_20_Trains_{filename_suffix}"
        excel_path = excel_dir / f"{base_name}.xlsx"
        pdf_path = pdf_dir / f"{base_name}.pdf"

        try:
            self._write_excel(excel_path, output_headers, output_rows, report_date=report_date)
            log_automation_event(logger, "report3_excel_generated", excel_path=str(excel_path))
            self._write_pdf(pdf_path, output_headers, output_rows, report_date=report_date)
            log_automation_event(logger, "report3_pdf_generated", pdf_path=str(pdf_path))
            verify_text_rendering(
                report_slug=report_slug,
                headers=output_headers,
                rows=output_rows,
                pdf_path=pdf_path,
            )
        except Exception as exc:
            return ProcessingResult(
                input_row_count=len(data_rows),
                success=False,
                error=str(exc),
                source_a_path=str(source_a_path),
                source_a_rows=len(data_rows),
            )

        log_automation_event(
            logger,
            "phase8_dataset_loaded",
            source_a=str(source_a_path),
            input_row_count=len(data_rows),
            top_n_selected=len(top_n_rows),
        )

        return ProcessingResult(
            success=True,
            input_row_count=len(data_rows),
            processed_row_count=len(output_rows),
            excel_path=str(excel_path),
            pdf_path=str(pdf_path),
            source_a_path=str(source_a_path),
            source_a_rows=len(data_rows),
            output_columns=output_headers,
            visible_columns=visible_columns,
            selected_column_ids=keys,
            column_order=list(keys),
            configuration_source=config_source,
        )

    @classmethod
    def build_canonical_rows(cls, rows: list[dict[str, str]]) -> list[dict[str, str]]:
        return [build_canonical_topn_row(row, serial=idx) for idx, row in enumerate(rows, start=1)]

    @classmethod
    def build_projected_table(
        cls,
        source_a_path: Path,
        *,
        report_slug: str = "train-no",
        column_selection: dict[str, Any] | None = None,
        selected_keys: list[str] | None = None,
    ) -> tuple[list[str], list[list[str]], list[str], list[str], str]:
        processor = cls()
        source_rows, _ = processor._read_csv(source_a_path)
        data_rows, _ = processor._split_total_row(source_rows)
        top_n_rows = data_rows[:TOP_N]
        canonical_rows = cls.build_canonical_rows(top_n_rows)
        return project_topn_for_output(
            report_slug,
            canonical_rows,
            selected_keys=selected_keys,
            column_selection=column_selection,
        )

    @staticmethod
    def _read_csv(path: Path) -> tuple[list[dict[str, str]], list[str]]:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            headers = list(reader.fieldnames or [])
            rows = [{header: row.get(header, "") for header in headers} for row in reader]
        return rows, headers

    @staticmethod
    def _is_total_row(row: dict[str, str]) -> bool:
        for key in ("Train No.", "Train Name", "Owning Zone"):
            val = row.get(key, "")
            if "total" in val.strip().lower():
                return True
        return False

    def _split_total_row(
        self,
        rows: list[dict[str, str]],
    ) -> tuple[list[dict[str, str]], dict[str, str] | None]:
        if not rows:
            return [], None
        if self._is_total_row(rows[-1]):
            return rows[:-1], rows[-1]
        return rows, None

    @staticmethod
    def _train_no_column_index(headers: list[str]) -> int | None:
        for idx, header in enumerate(headers, start=1):
            if header.strip() in {"Train No.", "Train No"}:
                return idx
        return None

    def _write_excel(
        self,
        target_path: Path,
        headers: list[str],
        rows: list[list[str]],
        *,
        report_date: str,
    ) -> None:
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Report 3"

        title = normalize_report_title(
            f"Rail Madad Report No 3 - 20 Trains having maximum grievances "
            f"{report_date}",
            report_slug="train-no",
        )
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = worksheet.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="center")

        header_row = 2
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            letter = get_column_letter(col_idx)
            worksheet.column_dimensions[letter].width = _REPORT3_EXCEL_WIDTHS.get(header, 14)

        train_no_col = self._train_no_column_index(headers)
        data_start = 3
        for row_offset, row_values in enumerate(rows):
            row_idx = data_start + row_offset
            for col_idx, value in enumerate(row_values, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                header = headers[col_idx - 1] if col_idx - 1 < len(headers) else ""
                wrap = header in {"Train Name", "Owning Zone", "Owning Division"}
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="top" if wrap else "center",
                    wrap_text=True,
                )
                if train_no_col is not None and col_idx == train_no_col:
                    cell.number_format = "@"

        highlight_south_central_railway_rows(
            worksheet,
            start_row=data_start,
            end_row=worksheet.max_row,
            start_col=1,
            end_col=len(headers),
        )

        apply_uniform_center_alignment(worksheet)

        workbook.save(temp_path)
        temp_path.replace(target_path)

    def _write_pdf(
        self,
        target_path: Path,
        headers: list[str],
        rows: list[list[str]],
        *,
        report_date: str,
    ) -> None:
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")

        table_data: list[list[object]] = [list(headers)]
        scr_row_indices: set[int] = set()
        for row_values in rows:
            table_data.append(list(row_values))
            if row_contains_scr(row_values):
                scr_row_indices.add(len(table_data) - 1)

        style_commands: list[tuple] = [
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 0), pdf_font_bold()),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
            ("VALIGN", (0, 1), (-1, -1), "TOP"),
        ]
        for row_idx in scr_row_indices:
            style_commands.append(("BACKGROUND", (0, row_idx), (-1, row_idx), colors.yellow))
            style_commands.append(("TEXTCOLOR", (0, row_idx), (-1, row_idx), colors.black))

        margin = _PDF_MARGIN_PT
        page_width, _ = landscape(A4)
        usable_width = page_width - (2 * margin)
        col_widths = _allocate_report3_col_widths(headers, usable_width)
        wrapped_data = _prepare_report3_pdf_table_data(table_data, headers, _PDF_FONT_SIZE)
        table = Table(wrapped_data, colWidths=col_widths, repeatRows=1)
        commands = _table_style_commands(
            style_commands,
            _PDF_FONT_SIZE,
            valign="TOP",
        )
        table.setStyle(TableStyle(commands))
        table = _rescale_if_overflow(
            wrapped_data,
            table,
            col_widths,
            commands,
            landscape(A4),
            margin,
            splittable=False,
        )

        title = normalize_report_title(
            f"Rail Madad Report No 3 - 20 Trains having maximum grievances "
            f"{report_date}",
            report_slug="train-no",
        )
        story = [
            Paragraph(title, pdf_title_style("Report3Title")),
            Spacer(1, 12),
            table,
        ]
        pagesize = landscape(A4)
        doc = SimpleDocTemplate(
            str(temp_path),
            pagesize=pagesize,
            leftMargin=margin,
            rightMargin=margin,
            topMargin=margin,
            bottomMargin=margin,
        )
        doc.build(story)
        temp_path.replace(target_path)
