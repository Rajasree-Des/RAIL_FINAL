"""Report 9 post-ingestion processor — Train/Station Cause Wise Grievances.

Loads four current-run CSVs via report9_combined_index.csv and emits one XLSX + one PDF
with sections stacked vertically:

1. All Zones — Train
2. All Zones — Station
3. SCR — Train
4. SCR — Station
"""

from __future__ import annotations

import csv
import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import KeepTogether, LongTable, Paragraph, SimpleDocTemplate, Spacer, TableStyle

from app.automation.config import config
from app.automation.date_range import date_range_for_processing
from app.automation.formatting.artifact_titles import build_artifact_main_title
from app.automation.formatting.excel_print import apply_uniform_center_alignment
from app.automation.formatting.pdf_fonts import ensure_pdf_unicode_fonts, pdf_font_bold, pdf_font_regular
from app.automation.formatting.pdf_table import SAFE_MARGIN_PT
from app.automation.formatting.text_pipeline import normalize_report_title
from app.automation.processing.base import ProcessingResult
from app.automation.processing.report9_output_columns import (
    project_report9_row,
    report9_default_ids,
    report9_labels,
    validate_selected_report9_fields,
)
from app.automation.report9_filters import (
    OUTPUT_HEADERS,
    SECTION_ORDER,
    Report9SourceConfig,
)
from app.automation.utils import (
    ensure_directory,
    log_automation_event,
    resolve_run_scoped_dir,
)

logger = logging.getLogger(__name__)

PROCESSOR_NAME = "report9_processor"
REPORT9_SHEET_TITLE = "Cause Wise on Date"  # Excel sheet name limit is 31 chars
REPORT9_FILE_STEM = "Rail_Madad_All_Zones_Train_Station_Cause_Wise_on_Date"

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

TOTAL_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

_PDF_MARGIN_PT = min(SAFE_MARGIN_PT, 14.0)
_PDF_SECTION_GAP_PT = 8.0
_REPORT9_PDF_COL_WIDTHS = [8, 22, 12, 12]


def _escape_paragraph_xml(text: str) -> str:
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _parse_received(value: str) -> float:
    cleaned = re.sub(r"[^\d.\-]", "", str(value or "").replace(",", ""))
    if not cleaned or cleaned in {".", "-", "-."}:
        return 0.0
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _format_share(value: str | float) -> str:
    try:
        num = float(str(value).replace("%", "").replace(",", "").strip())
        return f"{num:.2f}"
    except (TypeError, ValueError):
        return str(value or "")


@dataclass
class Report9Section:
    config: Report9SourceConfig
    headers: list[str]
    rows: list[list[str]]
    total_row: list[str]


class Report9Processor:
    """Process Report 9 cause-wise sections into combined Excel/PDF."""

    processor_name = PROCESSOR_NAME

    def process(
        self,
        *,
        source_a_path: Path,
        report_slug: str,
        source_b_path: Path | None = None,
        column_selection: dict[str, Any] | None = None,
    ) -> ProcessingResult:
        if source_a_path.suffix.lower() == ".pdf":
            return ProcessingResult(success=False, error="PDF cannot be used as processing input")

        sections, total_input_rows = self._load_sections(source_a_path)
        if len(sections) < len(SECTION_ORDER):
            missing = [
                cfg.source_id
                for cfg in SECTION_ORDER
                if not any(s.config.source_id == cfg.source_id for s in sections)
            ]
            return ProcessingResult(
                success=False,
                error=f"REPORT9_TABLE_MISSING: missing sections {missing}",
                source_a_path=str(source_a_path),
                input_row_count=total_input_rows,
            )

        selected_ids = report9_default_ids()
        config_source = "default"
        if column_selection:
            raw = (
                column_selection.get("selected_column_ids")
                or column_selection.get("column_order")
                or column_selection.get("default_selected_column_ids")
                or []
            )
            if raw:
                selected_ids = validate_selected_report9_fields(raw)
                config_source = str(
                    column_selection.get("configuration_source") or "manual_snapshot"
                )
        visible_headers = report9_labels(selected_ids)
        sections = [
            Report9Section(
                config=section.config,
                headers=list(visible_headers),
                rows=[project_report9_row(row, selected_ids) for row in section.rows],
                total_row=project_report9_row(section.total_row, selected_ids),
            )
            for section in sections
        ]

        date_range = date_range_for_processing(column_selection)
        main_title = build_artifact_main_title("report9", date_range)
        filename_suffix = date_range.filename_suffix()

        run_id = (column_selection or {}).get("run_id") if column_selection else None
        if run_id:
            excel_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_excel_dir, report_slug, str(run_id))
            )
            pdf_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_pdf_dir, report_slug, str(run_id))
            )
        else:
            # Fall back to parent of index path's run folder if present.
            parent = source_a_path.parent
            excel_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_excel_dir, report_slug, parent.name)
            )
            pdf_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_pdf_dir, report_slug, parent.name)
            )

        base_name = f"{REPORT9_FILE_STEM}_{filename_suffix}"
        excel_path = excel_dir / f"{base_name}.xlsx"
        pdf_path = pdf_dir / f"{base_name}.pdf"

        try:
            self._write_excel(excel_path, sections, main_title=main_title)
            log_automation_event(logger, "report9_excel_generated", excel_path=str(excel_path))
        except Exception as exc:
            return ProcessingResult(
                success=False,
                error=f"REPORT9_XLSX_FAILED: {exc}",
                source_a_path=str(source_a_path),
                input_row_count=total_input_rows,
            )

        try:
            self._write_pdf(pdf_path, sections, main_title=main_title)
            log_automation_event(logger, "report9_pdf_generated", pdf_path=str(pdf_path))
        except Exception as exc:
            return ProcessingResult(
                success=False,
                error=f"REPORT9_PDF_FAILED: {exc}",
                source_a_path=str(source_a_path),
                input_row_count=total_input_rows,
                excel_path=str(excel_path),
            )

        total_output_rows = sum(len(s.rows) for s in sections)
        log_automation_event(
            logger,
            "report9_processing_completed",
            source_a=str(source_a_path),
            input_row_count=total_input_rows,
            section_count=len(sections),
            total_output_rows=total_output_rows,
        )

        return ProcessingResult(
            success=True,
            attempted=True,
            input_row_count=total_input_rows,
            processed_row_count=total_output_rows,
            excel_path=str(excel_path),
            pdf_path=str(pdf_path),
            source_a_path=str(source_a_path),
            source_a_rows=total_input_rows,
            output_columns=list(visible_headers),
            visible_columns=list(visible_headers),
            selected_column_ids=list(selected_ids),
            column_order=list(selected_ids),
            configuration_source=config_source,
        )

    def _load_sections(self, source_a_path: Path) -> tuple[list[Report9Section], int]:
        sections: list[Report9Section] = []
        total_input_rows = 0
        index_entries = self._read_combined_index(source_a_path)

        if not index_entries:
            base_dir = source_a_path.parent
            for cfg in SECTION_ORDER:
                csv_path = base_dir / cfg.filename
                if csv_path.is_file():
                    index_entries[cfg.source_id] = {
                        "source_id": cfg.source_id,
                        "csv_path": str(csv_path),
                        "status": "success",
                    }

        for cfg in SECTION_ORDER:
            entry = index_entries.get(cfg.source_id)
            if entry is None or str(entry.get("status", "")).lower() != "success":
                continue
            csv_path = Path(str(entry.get("csv_path") or ""))
            if not csv_path.is_file():
                continue

            raw_rows, raw_headers = self._read_csv(csv_path)
            if not raw_rows and not raw_headers:
                continue

            data_rows, _portal_total = self._split_total_row(raw_rows)
            total_input_rows += len(data_rows)
            projected = self._project_and_sort(raw_headers, data_rows)
            total_row = self._compute_total_row(projected)

            sections.append(
                Report9Section(
                    config=cfg,
                    headers=list(OUTPUT_HEADERS),
                    rows=projected,
                    total_row=total_row,
                )
            )
            log_automation_event(
                logger,
                "report9_section_loaded",
                source_id=cfg.source_id,
                row_count=len(projected),
            )

        return sections, total_input_rows

    def _read_combined_index(self, source_a_path: Path) -> dict[str, dict[str, str]]:
        if source_a_path.name != "report9_combined_index.csv":
            return {}
        if not source_a_path.is_file():
            return {}
        entries: dict[str, dict[str, str]] = {}
        with source_a_path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                source_id = (row.get("source_id") or "").strip()
                if not source_id:
                    continue
                entries[source_id] = {
                    "source_id": source_id,
                    "section_title": (row.get("section_title") or "").strip(),
                    "zone": (row.get("zone") or "").strip(),
                    "csv_path": (row.get("csv_path") or "").strip(),
                    "row_count": (row.get("row_count") or "0").strip(),
                    "status": (row.get("status") or "").strip(),
                    "error": (row.get("error") or "").strip(),
                }
        return entries

    @staticmethod
    def _read_csv(path: Path) -> tuple[list[dict[str, str]], list[str]]:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            headers = list(reader.fieldnames or [])
            rows = [{header: row.get(header, "") for header in headers} for row in reader]
        return rows, headers

    @staticmethod
    def _split_total_row(
        rows: list[dict[str, str]],
    ) -> tuple[list[dict[str, str]], dict[str, str] | None]:
        if not rows:
            return [], None
        last = rows[-1]
        cause = str(last.get("Cause") or last.get("cause") or "").strip().lower()
        if cause == "total":
            return rows[:-1], last
        # Also detect Total in any value.
        if any(str(v).strip().lower() == "total" for v in last.values()):
            return rows[:-1], last
        return rows, None

    def _project_and_sort(
        self,
        raw_headers: list[str],
        data_rows: list[dict[str, str]],
    ) -> list[list[str]]:
        header_map = {h.strip().lower(): h for h in raw_headers}

        def col(*names: str) -> str | None:
            for name in names:
                key = name.lower()
                if key in header_map:
                    return header_map[key]
            for h in raw_headers:
                hl = h.strip().lower()
                for name in names:
                    if name.lower() in hl:
                        return h
            return None

        cause_key = col("Cause")
        received_key = col("Received")
        share_key = col("% Share", "%Share", "Share")

        projected: list[tuple[float, list[str]]] = []
        for row in data_rows:
            cause = str(row.get(cause_key or "", "") if cause_key else "").strip()
            if cause.lower() == "total":
                continue
            received_raw = str(row.get(received_key or "", "") if received_key else "").strip()
            share_raw = str(row.get(share_key or "", "") if share_key else "").strip()
            received_num = _parse_received(received_raw)
            projected.append(
                (
                    received_num,
                    [
                        "",  # S.No. filled after sort
                        cause,
                        str(int(received_num)) if received_num == int(received_num) else str(received_num),
                        share_raw,
                    ],
                )
            )

        # Received descending; stable ties keep relative order.
        projected.sort(key=lambda item: item[0], reverse=True)
        total_received = sum(item[0] for item in projected) or 0.0
        rows: list[list[str]] = []
        for idx, (recv, values) in enumerate(projected, start=1):
            values[0] = str(idx)
            if values[3].strip():
                values[3] = _format_share(values[3])
            elif total_received > 0:
                values[3] = f"{(recv / total_received) * 100.0:.2f}"
            else:
                values[3] = "0.00"
            rows.append(values)
        return rows

    @staticmethod
    def _compute_total_row(rows: list[list[str]]) -> list[str]:
        total_received = 0.0
        for row in rows:
            total_received += _parse_received(row[2] if len(row) > 2 else "0")
        received_str = (
            str(int(total_received))
            if total_received == int(total_received)
            else str(total_received)
        )
        return ["", "Total", received_str, "100.00"]

    def _write_excel(
        self,
        target_path: Path,
        sections: list[Report9Section],
        *,
        main_title: str,
    ) -> None:
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = REPORT9_SHEET_TITLE

        col_count = len(sections[0].headers) if sections else len(OUTPUT_HEADERS)
        main_title = normalize_report_title(main_title, report_slug="report9")
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        title_cell = worksheet.cell(row=1, column=1, value=main_title)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        current_row = 2
        for section_idx, section in enumerate(sections):
            worksheet.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=col_count,
            )
            heading_cell = worksheet.cell(
                row=current_row,
                column=1,
                value=section.config.section_title,
            )
            heading_cell.font = Font(bold=True, size=11)
            heading_cell.alignment = Alignment(horizontal="center")
            current_row += 1

            for col_idx, header in enumerate(section.headers, start=1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
            current_row += 1

            for row_values in section.rows:
                for col_idx, value in enumerate(row_values, start=1):
                    cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                    cell.border = THIN_BORDER
                    if col_idx in (1, 3, 4):
                        cell.alignment = Alignment(horizontal="center")
                current_row += 1

            for col_idx, value in enumerate(section.total_row, start=1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                cell.font = Font(bold=True)
                cell.border = THIN_BORDER
                cell.fill = TOTAL_FILL
                if col_idx in (1, 3, 4):
                    cell.alignment = Alignment(horizontal="center")
            current_row += 1
            if section_idx < len(sections) - 1:
                current_row += 1

        equal_col_width = 14
        for col_idx in range(1, col_count + 1):
            worksheet.column_dimensions[chr(64 + col_idx)].width = equal_col_width

        apply_uniform_center_alignment(worksheet)

        workbook.save(temp_path)
        temp_path.replace(target_path)

    def _write_pdf(
        self,
        target_path: Path,
        sections: list[Report9Section],
        *,
        main_title: str,
    ) -> None:
        ensure_pdf_unicode_fonts()
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        margin = _PDF_MARGIN_PT
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Report9Title",
            parent=styles["Heading1"],
            fontSize=14,
            leading=16,
            alignment=TA_CENTER,
            spaceAfter=4,
            fontName=pdf_font_bold(),
        )
        section_style = ParagraphStyle(
            "Report9Section",
            parent=styles["Heading2"],
            fontSize=10,
            leading=12,
            alignment=TA_CENTER,
            spaceBefore=2,
            spaceAfter=2,
            fontName=pdf_font_bold(),
        )

        normalized_main_title = normalize_report_title(main_title, report_slug="report9")

        story: list[Any] = [
            Paragraph(_escape_paragraph_xml(normalized_main_title), title_style),
        ]

        page_width, _page_height = landscape(A4)
        usable_width = page_width - (2 * margin)

        for section_idx, section in enumerate(sections):
            heading = Paragraph(
                _escape_paragraph_xml(section.config.section_title),
                section_style,
            )
            table_data: list[list[object]] = [list(section.headers)]
            table_data.extend([list(r) for r in section.rows])
            table_data.append(list(section.total_row))

            col_width_total = sum(_REPORT9_PDF_COL_WIDTHS)
            col_widths = [
                usable_width * width / col_width_total for width in _REPORT9_PDF_COL_WIDTHS
            ]
            table = LongTable(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8E8E8")),
                        ("FONTNAME", (0, 0), (-1, 0), pdf_font_bold()),
                        ("FONTNAME", (0, 1), (-1, -2), pdf_font_regular()),
                        ("FONTNAME", (0, -1), (-1, -1), pdf_font_bold()),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#D9D9D9")),
                        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 3),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                        ("TOPPADDING", (0, 0), (-1, -1), 2),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                    ]
                )
            )
            section_flow: list[Any] = [heading, table]
            if len(table_data) <= 18:
                story.append(KeepTogether(section_flow))
            else:
                story.extend(section_flow)
            if section_idx < len(sections) - 1:
                story.append(Spacer(1, _PDF_SECTION_GAP_PT))

        doc = SimpleDocTemplate(
            str(temp_path),
            pagesize=landscape(A4),
            leftMargin=margin,
            rightMargin=margin,
            topMargin=margin,
            bottomMargin=margin,
            title=normalized_main_title,
        )
        doc.build(story)
        temp_path.replace(target_path)
