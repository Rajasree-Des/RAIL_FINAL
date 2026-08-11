"""Report Vande Bharat post-ingestion processor.

Reads the detailed complaint CSV produced after TOTAL→Received drill-down and
emits the final VB RAILMADAD REPORT - SCR Excel + PDF.
"""

from __future__ import annotations

import csv
import json
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table as RLTable, TableStyle

from app.automation.config import config
from app.automation.date_range import ReportDateRange, date_range_for_processing
from app.automation.formatting.excel_print import apply_uniform_center_alignment
from app.automation.formatting.pdf_fonts import ensure_pdf_unicode_fonts, pdf_font_bold, pdf_font_regular
from app.automation.formatting.pdf_table import _cell_text, _escape_paragraph_text
from app.automation.formatting.text_safe import field_kind_for_header
from app.automation.processing.base import ProcessingResult
from app.automation.processing.report18_output_columns import (
    report18_default_ids,
    resolve_report18_header_indexes,
    validate_selected_report18_fields,
)
from app.automation.report18_detail_extract import (
    REPORT18_FINAL_HEADERS,
    REPORT18_LEGACY_HEADER_ALIASES,
    REPORT18_SUMMARY_META_FILENAME,
)
from app.automation.report18_filters import REPORT18_FILE_STEM, REPORT18_LOG_PREFIX
from app.automation.utils import (
    ensure_directory,
    log_automation_event,
    resolve_run_scoped_dir,
)

logger = logging.getLogger(__name__)

PROCESSOR_NAME = "report18_processor"
REPORT18_SHEET_TITLE = "Vande Bharat"
REPORT18_MAIN_TITLE = "VB RAILMADAD REPORT - SCR"

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_PDF_MARGIN_PT = 10.0
_PDF_TITLE_GAP_PT = 6.0
_PDF_HEIGHT_BUFFER_PT = 14.0
_PDF_FONT_SIZE = 6.8
_PDF_CELL_PADDING_PT = 3.0
_PDF_REMARKS_MAX_LINES = 5

# Minimum column widths (pt) so headers and typical values do not overlap.
_REPORT18_MIN_COL_PT: dict[str, float] = {
    "Sl No": 24,
    "complaintRefNo": 66,
    "createdOn": 56,
    "modifiedOn": 56,
    "trainStation": 40,
    "channelType": 32,
    "compTypeName": 60,
    "ownZoneCode": 30,
    "deptCode": 30,
    "sla": 24,
    "rating": 40,
    "status": 34,
    "feedbackRemarks": 44,
    "restStation": 32,
    "contactNo": 50,
    "physicalCoachNo": 46,
    "trainNameForReport": 54,
    "complaintDesc": 76,
    "remarks": 76,
    "userid": 46,
    "feedbackRemark": 44,
    "nextStation": 32,
    "contactId": 50,
    "userId": 46,
}

# Extra width distribution after minimums are satisfied.
REPORT18_COL_WEIGHTS: dict[str, float] = {
    "Sl No": 0.020,
    "complaintRefNo": 0.070,
    "createdOn": 0.058,
    "modifiedOn": 0.058,
    "trainStation": 0.034,
    "channelType": 0.024,
    "compTypeName": 0.068,
    "ownZoneCode": 0.024,
    "deptCode": 0.024,
    "sla": 0.020,
    "rating": 0.040,
    "status": 0.028,
    "feedbackRemarks": 0.040,
    "restStation": 0.028,
    "contactNo": 0.048,
    "physicalCoachNo": 0.042,
    "trainNameForReport": 0.060,
    "complaintDesc": 0.118,
    "remarks": 0.118,
    "userid": 0.052,
    "feedbackRemark": 0.040,
    "nextStation": 0.028,
    "contactId": 0.048,
    "userId": 0.052,
}
_WRAP_HEADERS = frozenset({
    "complaintDesc",
    "remarks",
    "feedbackRemarks",
    "trainNameForReport",
    "userid",
})


def _escape_paragraph_xml(text: str) -> str:
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _format_report18_date(date_range: ReportDateRange) -> str:
    """Display date for VB report header (DD-MMM-YYYY)."""
    if date_range.date_from == date_range.date_to:
        return date_range.date_from.strftime("%d-%b-%Y")
    start = date_range.date_from.strftime("%d-%b-%Y")
    end = date_range.date_to.strftime("%d-%b-%Y")
    return f"{start} to {end}"


def _migrate_report18_header(header: str) -> str:
    key = str(header or "").strip()
    return REPORT18_LEGACY_HEADER_ALIASES.get(key, key)


def _report18_header_min_width(header: str, font_size: float) -> float:
    ensure_pdf_unicode_fonts()
    measured = stringWidth(str(header or ""), pdf_font_bold(), font_size) + 10
    floor = _REPORT18_MIN_COL_PT.get(header, 30)
    return max(measured, floor)


def _allocate_report18_col_widths(
    headers: list[str],
    usable_width: float,
    *,
    font_size: float,
) -> list[float]:
    mins = [_report18_header_min_width(header, font_size) for header in headers]
    min_total = sum(mins)
    if min_total >= usable_width:
        scale = usable_width / min_total
        return [width * scale for width in mins]

    extra = usable_width - min_total
    weights = [REPORT18_COL_WEIGHTS.get(header, 0.04) for header in headers]
    weight_total = sum(weights) or 1.0
    return [
        mins[idx] + (extra * weights[idx] / weight_total)
        for idx in range(len(headers))
    ]


def _wrap_text_to_width_lines(
    text: str,
    *,
    avail_width: float,
    font_name: str,
    font_size: float,
) -> list[str]:
    words = str(text or "").split()
    if not words:
        return []
    lines: list[str] = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip() if current else word
        if stringWidth(candidate, font_name, font_size) <= avail_width:
            current = candidate
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines


def _truncate_remarks_for_pdf(
    text: str,
    *,
    col_width_pt: float,
    font_size: float,
    max_lines: int = _PDF_REMARKS_MAX_LINES,
) -> str:
    """Limit remarks to max_lines in the PDF (full text remains in Excel/CSV)."""
    normalized = str(text or "").strip()
    if not normalized:
        return ""

    ensure_pdf_unicode_fonts()
    font_name = pdf_font_regular()
    avail_width = max(col_width_pt - (2 * _PDF_CELL_PADDING_PT), 24.0)
    lines = _wrap_text_to_width_lines(
        normalized,
        avail_width=avail_width,
        font_name=font_name,
        font_size=font_size,
    )
    if len(lines) <= max_lines:
        return normalized

    kept = lines[:max_lines]
    suffix = "..."
    last = kept[-1]
    while last and stringWidth(last + suffix, font_name, font_size) > avail_width:
        if " " in last:
            last = last.rsplit(" ", 1)[0]
        elif last:
            last = last[:-1]
        else:
            break
    kept[-1] = (last + suffix) if last else suffix
    return "\n".join(kept)


def _prepare_report18_pdf_table_data(
    table_data: list[list[str]],
    headers: list[str],
    font_size: float,
    col_widths: list[float],
) -> list[list[object]]:
    """Wrap every cell as a Paragraph so content respects column boundaries."""
    ensure_pdf_unicode_fonts()
    leading = font_size + 2.5
    header_style = ParagraphStyle(
        "Report18PdfHeader",
        fontName=pdf_font_bold(),
        fontSize=font_size,
        leading=leading,
        wordWrap="CJK",
        alignment=TA_CENTER,
    )
    center_style = ParagraphStyle(
        "Report18PdfCenter",
        fontName=pdf_font_regular(),
        fontSize=font_size,
        leading=leading,
        wordWrap="CJK",
        alignment=TA_CENTER,
    )
    left_style = ParagraphStyle(
        "Report18PdfLeft",
        fontName=pdf_font_regular(),
        fontSize=font_size,
        leading=leading,
        wordWrap="CJK",
        alignment=TA_LEFT,
    )
    header_kinds = [field_kind_for_header(header) for header in headers]
    wrapped: list[list[object]] = []
    for row_idx, row in enumerate(table_data):
        new_row: list[object] = []
        for col_idx, cell in enumerate(row):
            header = headers[col_idx] if col_idx < len(headers) else ""
            kind = header_kinds[col_idx] if col_idx < len(header_kinds) else "text"
            text = _cell_text(cell)
            if row_idx > 0 and header == "remarks":
                col_width = col_widths[col_idx] if col_idx < len(col_widths) else 76.0
                text = _truncate_remarks_for_pdf(
                    text,
                    col_width_pt=col_width,
                    font_size=font_size,
                )
            if row_idx == 0:
                style = header_style
            elif header in _WRAP_HEADERS:
                style = left_style
            else:
                style = center_style
            new_row.append(
                Paragraph(_escape_paragraph_text(text, field_kind=kind), style)
            )
        wrapped.append(new_row)
    return wrapped


def _report18_table_style_commands(headers: list[str], *, font_size: float) -> list[tuple]:
    pad = _PDF_CELL_PADDING_PT
    style_commands: list[tuple] = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8E8E8")),
        ("FONTNAME", (0, 0), (-1, 0), pdf_font_bold()),
        ("FONTNAME", (0, 1), (-1, -1), pdf_font_regular()),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
        ("VALIGN", (0, 1), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), pad),
        ("RIGHTPADDING", (0, 0), (-1, -1), pad),
        ("TOPPADDING", (0, 0), (-1, -1), pad),
        ("BOTTOMPADDING", (0, 0), (-1, -1), pad),
    ]
    for idx, header in enumerate(headers):
        if header in _WRAP_HEADERS:
            style_commands.append(("ALIGN", (idx, 1), (idx, -1), "LEFT"))
    return style_commands


class Report18Processor:
    """Process detailed Vande Bharat complaints into Excel/PDF artifacts."""

    processor_name = PROCESSOR_NAME

    def process(
        self,
        *,
        source_a_path: Path,
        report_slug: str,
        source_b_path: Path | None = None,
        column_selection: dict[str, Any] | None = None,
    ) -> ProcessingResult:
        _ = source_b_path
        if source_a_path.suffix.lower() == ".pdf":
            return ProcessingResult(success=False, error="PDF cannot be used as processing input")

        headers, rows = self._load_table(source_a_path)
        if not headers:
            return ProcessingResult(
                success=False,
                error="REPORT18_TABLE_MISSING: extracted CSV has no headers",
                source_a_path=str(source_a_path),
                input_row_count=0,
            )
        if not rows:
            return ProcessingResult(
                success=False,
                error="REPORT18_TABLE_MISSING: extracted CSV has no data rows",
                source_a_path=str(source_a_path),
                input_row_count=0,
            )

        headers = [_migrate_report18_header(h) for h in headers]

        selected_ids = report18_default_ids()
        config_source = "default"
        if column_selection:
            raw = (
                column_selection.get("selected_column_ids")
                or column_selection.get("column_order")
                or column_selection.get("default_selected_column_ids")
                or []
            )
            if raw:
                migrated = [
                    REPORT18_LEGACY_HEADER_ALIASES.get(str(item).strip(), str(item).strip())
                    for item in raw
                ]
                selected_ids = validate_selected_report18_fields(migrated)
                config_source = str(
                    column_selection.get("configuration_source") or "manual_snapshot"
                )

        # If source already uses final headers, project by selection; otherwise keep as-is.
        if set(h.strip() for h in headers) & set(REPORT18_FINAL_HEADERS):
            visible_headers, col_indexes = resolve_report18_header_indexes(headers, selected_ids)
            rows = [[row[i] if i < len(row) else "" for i in col_indexes] for row in rows]
            headers = list(visible_headers)
        else:
            logger.warning(
                "%s Source CSV headers are not the final VB schema; writing as provided. headers=%s",
                REPORT18_LOG_PREFIX,
                headers,
            )

        meta = self._load_summary_meta(source_a_path.parent)
        summary_total = meta.get("summary_total")
        detail_excel_rows = meta.get("detail_excel_rows")
        logger.info(
            "%s Summary total: %s | Detailed Excel records: %s | Final report records: %s",
            REPORT18_LOG_PREFIX,
            summary_total,
            detail_excel_rows if detail_excel_rows is not None else len(rows),
            len(rows),
        )
        if summary_total is not None and int(summary_total) != len(rows):
            logger.error(
                "%s Count mismatch — Summary total=%s Detailed/Final rows=%s",
                REPORT18_LOG_PREFIX,
                summary_total,
                len(rows),
            )
            return ProcessingResult(
                success=False,
                error=(
                    f"vande_bharat_detail_reconciliation_failed: "
                    f"aggregate={summary_total} details={len(rows)}"
                ),
                source_a_path=str(source_a_path),
                input_row_count=len(rows),
            )

        date_range = date_range_for_processing(column_selection)
        report_date = _format_report18_date(date_range)

        run_id = (column_selection or {}).get("run_id") if column_selection else None
        if run_id:
            excel_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_excel_dir, report_slug, str(run_id))
            )
            pdf_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_pdf_dir, report_slug, str(run_id))
            )
        else:
            parent = source_a_path.parent
            excel_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_excel_dir, report_slug, parent.name)
            )
            pdf_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_pdf_dir, report_slug, parent.name)
            )

        excel_path = excel_dir / f"{REPORT18_FILE_STEM}.xlsx"
        pdf_path = pdf_dir / f"{REPORT18_FILE_STEM}.pdf"

        try:
            self._write_excel(
                excel_path,
                headers,
                rows,
                report_date=report_date,
            )
            logger.info("%s Excel generated", REPORT18_LOG_PREFIX)
            log_automation_event(logger, "report18_excel_generated", excel_path=str(excel_path))
        except Exception as exc:
            return ProcessingResult(
                success=False,
                error=f"REPORT18_XLSX_FAILED: {exc}",
                source_a_path=str(source_a_path),
                input_row_count=len(rows),
            )

        try:
            self._write_pdf(
                pdf_path,
                headers,
                rows,
                report_date=report_date,
            )
            logger.info("%s PDF generated", REPORT18_LOG_PREFIX)
            log_automation_event(logger, "report18_pdf_generated", pdf_path=str(pdf_path))
        except Exception as exc:
            return ProcessingResult(
                success=False,
                error=f"REPORT18_PDF_FAILED: {exc}",
                source_a_path=str(source_a_path),
                input_row_count=len(rows),
                excel_path=str(excel_path),
            )

        log_automation_event(
            logger,
            "report18_processing_completed",
            source_a=str(source_a_path),
            input_row_count=len(rows),
            total_output_rows=len(rows),
            summary_total=summary_total,
        )

        return ProcessingResult(
            success=True,
            attempted=True,
            input_row_count=len(rows),
            processed_row_count=len(rows),
            excel_path=str(excel_path),
            pdf_path=str(pdf_path),
            source_a_path=str(source_a_path),
            source_a_rows=len(rows),
            output_columns=list(headers),
            visible_columns=list(headers),
            selected_column_ids=list(selected_ids),
            column_order=list(selected_ids),
            configuration_source=config_source,
        )

    @staticmethod
    def _load_summary_meta(directory: Path) -> dict[str, Any]:
        meta_path = directory / REPORT18_SUMMARY_META_FILENAME
        if not meta_path.is_file():
            return {}
        try:
            payload = json.loads(meta_path.read_text(encoding="utf-8"))
            return payload if isinstance(payload, dict) else {}
        except Exception:
            return {}

    def _load_table(self, source_a_path: Path) -> tuple[list[str], list[list[str]]]:
        if not source_a_path.is_file():
            return [], []
        with source_a_path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            rows = [list(row) for row in reader]
        if not rows:
            return [], []
        headers = [str(c).strip() for c in rows[0]]
        data = [
            [str(c).strip() for c in row]
            for row in rows[1:]
            if any(str(c).strip() for c in row)
        ]
        width = len(headers)
        normalized: list[list[str]] = []
        for row in data:
            if len(row) < width:
                row = row + [""] * (width - len(row))
            elif len(row) > width:
                row = row[:width]
            normalized.append(row)
        return headers, normalized

    def _write_excel(
        self,
        target_path: Path,
        headers: list[str],
        rows: list[list[str]],
        *,
        report_date: str,
    ) -> None:
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = REPORT18_SHEET_TITLE

        col_count = max(len(headers), 1)
        header_row = 2
        data_start_row = 3

        # Row 1: centered title; DATE label + value stacked at top-right.
        if col_count > 1:
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count - 1)
        title_cell = worksheet.cell(row=1, column=1, value=REPORT18_MAIN_TITLE)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        date_cell = worksheet.cell(row=1, column=col_count, value=f"DATE\n{report_date}")
        date_cell.font = Font(bold=True, size=11)
        date_cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        worksheet.row_dimensions[1].height = 34

        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx, row_values in enumerate(rows, start=data_start_row):
            for col_idx, value in enumerate(row_values, start=1):
                header = headers[col_idx - 1] if col_idx - 1 < len(headers) else ""
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                wrap = header in _WRAP_HEADERS
                cell.alignment = Alignment(
                    horizontal="left" if wrap else "center",
                    vertical="top",
                    wrap_text=True,
                )
            worksheet.row_dimensions[row_idx].height = 30

        widths = {
            "Sl No": 8,
            "complaintRefNo": 16,
            "createdOn": 14,
            "modifiedOn": 14,
            "trainStation": 12,
            "channelType": 12,
            "compTypeName": 22,
            "ownZoneCode": 12,
            "deptCode": 10,
            "sla": 8,
            "rating": 14,
            "status": 10,
            "feedbackRemarks": 18,
            "restStation": 12,
            "contactNo": 14,
            "physicalCoachNo": 14,
            "trainNameForReport": 26,
            "complaintDesc": 36,
            "remarks": 36,
            "userid": 18,
        }
        for col_idx, header in enumerate(headers, start=1):
            letter = get_column_letter(col_idx)
            worksheet.column_dimensions[letter].width = widths.get(header, 14)

        apply_uniform_center_alignment(worksheet)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        date_cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        for row_idx in range(header_row, header_row + 1):
            for col_idx, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_idx in range(data_start_row, data_start_row + len(rows)):
            for col_idx, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                wrap = header in _WRAP_HEADERS
                cell.alignment = Alignment(
                    horizontal="left" if wrap else "center",
                    vertical="top",
                    wrap_text=True,
                )

        worksheet.freeze_panes = worksheet.cell(row=data_start_row, column=1)

        workbook.save(temp_path)
        temp_path.replace(target_path)

    def _write_pdf(
        self,
        target_path: Path,
        headers: list[str],
        rows: list[list[str]],
        *,
        report_date: str,
    ) -> None:
        ensure_pdf_unicode_fonts()
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        margin = _PDF_MARGIN_PT
        page_width, _ = landscape(A4)
        usable_width = page_width - (2 * margin)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Report18Title",
            parent=styles["Heading1"],
            fontSize=13,
            leading=15,
            alignment=TA_CENTER,
            spaceAfter=0,
            fontName=pdf_font_bold(),
        )
        date_style = ParagraphStyle(
            "Report18Date",
            parent=styles["Normal"],
            fontSize=9,
            leading=12,
            alignment=TA_RIGHT,
            fontName=pdf_font_bold(),
        )

        table_data: list[list[str]] = [headers, *rows]
        col_widths = _allocate_report18_col_widths(
            headers,
            usable_width,
            font_size=_PDF_FONT_SIZE,
        )
        wrapped_data = _prepare_report18_pdf_table_data(
            table_data,
            headers,
            _PDF_FONT_SIZE,
            col_widths,
        )
        table = RLTable(wrapped_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle(_report18_table_style_commands(headers, font_size=_PDF_FONT_SIZE)))

        date_block = Paragraph(
            f"DATE<br/>{_escape_paragraph_xml(report_date)}",
            date_style,
        )
        header_table = RLTable(
            [
                [
                    Paragraph(_escape_paragraph_xml(REPORT18_MAIN_TITLE), title_style),
                    date_block,
                ]
            ],
            colWidths=[usable_width * 0.78, usable_width * 0.22],
            hAlign="LEFT",
        )
        header_table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (0, 0), "CENTER"),
                    ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )

        story: list[Any] = [
            header_table,
            Spacer(1, _PDF_TITLE_GAP_PT),
            table,
        ]

        content_height = 0.0
        for flowable in story:
            _width, height = flowable.wrap(usable_width, 1_000_000)
            content_height += float(height)

        page_height = (content_height * 1.02) + (2 * margin) + _PDF_HEIGHT_BUFFER_PT
        pagesize = (page_width, page_height)

        log_automation_event(
            logger,
            "report18_pdf_single_page_layout",
            page_width=page_width,
            page_height=page_height,
            content_height=content_height,
            row_count=len(rows),
            column_count=len(headers),
        )

        doc = SimpleDocTemplate(
            str(temp_path),
            pagesize=pagesize,
            leftMargin=margin,
            rightMargin=margin,
            topMargin=margin,
            bottomMargin=margin,
            title=REPORT18_MAIN_TITLE,
        )
        doc.build(story)
        temp_path.replace(target_path)
