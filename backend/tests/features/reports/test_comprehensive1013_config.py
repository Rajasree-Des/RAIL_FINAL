"""Tests for Report 10-13 comprehensive column configuration."""

from __future__ import annotations

import pytest

from app.automation.comprehensive1013_filters import COMPREHENSIVE_1013_SECTION_IDS
from app.automation.processing.column_config import (
    comprehensive_union_column_keys,
    sanitize_comprehensive_sections,
    validate_comprehensive_sections,
)
from app.automation.processing.comprehensive1013_processor import Comprehensive1013Processor
from app.features.reports.schemas import ManualGenerateRequest
from app.features.reports.service import build_config_snapshot


def _full_sections(**overrides: list[str]) -> dict[str, dict[str, list[str]]]:
    defaults = {
        section_id: {
            "selected_column_ids": [
                "sno",
                "division",
                "opening_balance",
                "received",
                "share_percent",
                "closed",
                "closing_balance",
                "disposal_percent",
                "avg_disposal_time",
                "avg_rating",
                "avg_pendency_time",
            ]
        }
        for section_id in COMPREHENSIVE_1013_SECTION_IDS
    }
    for section_id, selected in overrides.items():
        defaults[section_id] = {"selected_column_ids": selected}
    return defaults


def test_build_config_snapshot_writes_top_level_sections_and_dates():
    sections = _full_sections(
        report10_cw=["sno", "division", "received", "closed"],
        report11_security=["sno", "division", "received", "share_percent", "avg_rating"],
    )
    body = ManualGenerateRequest(
        date_from="2026-08-01",
        date_to="2026-08-02",
        sections=sections,
    )
    snapshot = build_config_snapshot(body, report_slug="comprehensive-10-13")

    assert snapshot["date_from"] == "2026-08-01"
    assert snapshot["date_to"] == "2026-08-02"
    assert snapshot["sections"]["report10_cw"]["selected_column_ids"] == [
        "sno",
        "division",
        "received",
        "closed",
    ]
    assert snapshot["sections"]["report11_security"]["selected_column_ids"] == [
        "sno",
        "division",
        "received",
        "share_percent",
        "avg_rating",
    ]
    assert "opening_balance" in comprehensive_union_column_keys(snapshot["sections"])


def test_validate_comprehensive_sections_requires_each_section():
    with pytest.raises(ValueError, match="Report 10 — C&W"):
        validate_comprehensive_sections(
            {
                "report10_cw": {"selected_column_ids": []},
                "report11_security": {"selected_column_ids": ["sno"]},
                "report12_punctuality": {"selected_column_ids": ["sno"]},
                "report13_electrical": {"selected_column_ids": ["sno"]},
            }
        )


def test_sanitize_comprehensive_sections_filters_invalid_ids():
    sections = _full_sections(
        report10_cw=["sno", "division", "invalid_column", "received"],
    )
    sanitized = sanitize_comprehensive_sections(sections)
    assert sanitized["report10_cw"]["selected_column_ids"] == [
        "sno",
        "division",
        "received",
    ]


def test_resolve_column_ids_uses_section_specific_selection():
    processor = Comprehensive1013Processor()
    column_selection = {
        "sections": _full_sections(
            report10_cw=["sno", "division", "received", "closed"],
            report11_security=["sno", "division", "received"],
        )
    }

    report10_cols = processor._resolve_column_ids("report10_cw", column_selection)
    report11_cols = processor._resolve_column_ids("report11_security", column_selection)

    assert report10_cols == ["sno", "division", "received", "closed"]
    assert report11_cols == ["sno", "division", "received"]
    assert "opening_balance" not in report10_cols
    assert "opening_balance" in processor._resolve_column_ids(
        "report12_punctuality", column_selection
    )


def test_resolve_column_ids_errors_when_sections_present_but_section_empty():
    processor = Comprehensive1013Processor()
    column_selection = {
        "sections": {
            "report10_cw": {"selected_column_ids": ["sno", "division"]},
            "report11_security": {"selected_column_ids": []},
            "report12_punctuality": {"selected_column_ids": ["division"]},
            "report13_electrical": {"selected_column_ids": ["division"]},
        }
    }

    with pytest.raises(ValueError, match="Report 11 — Security"):
        processor._resolve_column_ids("report11_security", column_selection)


def test_project_columns_omits_deselected_fields():
    processor = Comprehensive1013Processor()
    raw_headers = [
        "S.No.",
        "Division",
        "Opening Balance",
        "Received",
        "Closed",
    ]
    data_rows = [
        {
            "S.No.": "1",
            "Division": "Hyderabad",
            "Opening Balance": "10",
            "Received": "5",
            "Closed": "3",
        }
    ]
    headers, rows = processor._project_columns(
        raw_headers,
        data_rows,
        ["sno", "division", "received", "closed"],
    )

    assert headers == ["S.No.", "Division", "Received", "Closed"]
    assert rows == [["1", "Hyderabad", "5", "3"]]


def test_resolve_effective_selection_uses_saved_config_when_snapshot_missing(monkeypatch):
    ten_cols = [
        "sno",
        "division",
        "received",
        "share_percent",
        "closed",
        "closing_balance",
        "disposal_percent",
        "avg_disposal_time",
        "avg_rating",
        "avg_pendency_time",
    ]
    saved = {
        "sections": {
            section_id: {"selected_column_ids": list(ten_cols)}
            for section_id in COMPREHENSIVE_1013_SECTION_IDS
        }
    }

    monkeypatch.setattr(
        "app.features.reports.config_store.load_report_config",
        lambda *_args, **_kwargs: saved,
    )

    processor = Comprehensive1013Processor()
    resolved = processor._resolve_effective_column_selection(None)

    assert resolved["configuration_source"] == "saved_user_config"
    for section_id in COMPREHENSIVE_1013_SECTION_IDS:
        assert "opening_balance" not in resolved["sections"][section_id]["selected_column_ids"]
        assert len(resolved["sections"][section_id]["selected_column_ids"]) == 10


def test_process_omits_opening_balance_from_pdf_and_excel(tmp_path, monkeypatch):
    import csv

    from openpyxl import load_workbook
    from pypdf import PdfReader

    from app.automation.processing.comprehensive_output_columns import (
        COMPREHENSIVE_COLUMN_IDS,
        COMPREHENSIVE_COLUMN_LABELS,
    )

    monkeypatch.setattr(
        "app.automation.processing.comprehensive1013_processor.config.output_excel_dir",
        str(tmp_path / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.comprehensive1013_processor.config.output_pdf_dir",
        str(tmp_path / "pdf"),
    )

    ten_ids = [c for c in COMPREHENSIVE_COLUMN_IDS if c != "opening_balance"]
    sections = {
        sid: {"selected_column_ids": list(ten_ids)} for sid in COMPREHENSIVE_1013_SECTION_IDS
    }

    base = tmp_path / "src"
    base.mkdir()
    headers = list(COMPREHENSIVE_COLUMN_LABELS.values())
    for section_id in COMPREHENSIVE_1013_SECTION_IDS:
        path = base / f"{section_id}.csv"
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers)
            writer.writeheader()
            writer.writerow(
                {
                    "S.No.": "1",
                    "Division": "HYB",
                    "Opening Balance": "9",
                    "Received": "5",
                    "% Share": "100.00",
                    "Closed": "5",
                    "Closing Balance": "0",
                    "% Disposal": "100.00",
                    "Avg. Disposal Time": "1.0",
                    "Avg. Rating": "4.0",
                    "Avg. Pendency Time": "0.0",
                }
            )
            writer.writerow(
                {
                    "S.No.": "",
                    "Division": "Total",
                    "Opening Balance": "9",
                    "Received": "5",
                    "% Share": "100.00",
                    "Closed": "5",
                    "Closing Balance": "0",
                    "% Disposal": "100.00",
                    "Avg. Disposal Time": "1.0",
                    "Avg. Rating": "4.0",
                    "Avg. Pendency Time": "0.0",
                }
            )

    index = base / "comprehensive_combined_index.csv"
    with index.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["section_id", "section_name", "csv_path", "row_count", "status", "error"],
        )
        writer.writeheader()
        for section_id in COMPREHENSIVE_1013_SECTION_IDS:
            writer.writerow(
                {
                    "section_id": section_id,
                    "section_name": section_id,
                    "csv_path": str(base / f"{section_id}.csv"),
                    "row_count": "1",
                    "status": "success",
                    "error": "",
                }
            )

    result = Comprehensive1013Processor().process(
        source_a_path=index,
        report_slug="comprehensive-10-13",
        column_selection={
            "sections": sections,
            "date_from": "2026-08-02",
            "date_to": "2026-08-02",
            "run_id": "filter-test-run",
        },
    )
    assert result.success, result.error
    assert "opening_balance" not in (result.selected_column_ids or [])

    pdf_text = "\n".join(
        (page.extract_text() or "") for page in PdfReader(result.pdf_path).pages
    )
    assert "Opening Balance" not in pdf_text

    wb = load_workbook(result.excel_path)
    ws = wb.active
    header_row = [ws.cell(3, c).value for c in range(1, 12)]
    assert "Opening Balance" not in header_row
    assert header_row[:10] == [
        "S.No.",
        "Division",
        "Received",
        "% Share",
        "Closed",
        "Closing Balance",
        "% Disposal",
        "Avg. Disposal Time",
        "Avg. Rating",
        "Avg. Pendency Time",
    ]


def test_process_partial_extraction_uses_saved_filters(tmp_path, monkeypatch):
    """If Report 10 extraction fails, still emit 11–13 with configured columns."""
    import csv

    from openpyxl import load_workbook
    from pypdf import PdfReader

    from app.automation.processing.comprehensive_output_columns import (
        COMPREHENSIVE_COLUMN_IDS,
        COMPREHENSIVE_COLUMN_LABELS,
    )

    monkeypatch.setattr(
        "app.automation.processing.comprehensive1013_processor.config.output_excel_dir",
        str(tmp_path / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.comprehensive1013_processor.config.output_pdf_dir",
        str(tmp_path / "pdf"),
    )

    ten_ids = [c for c in COMPREHENSIVE_COLUMN_IDS if c != "opening_balance"]
    sections_payload = {
        sid: {"selected_column_ids": list(ten_ids)} for sid in COMPREHENSIVE_1013_SECTION_IDS
    }
    monkeypatch.setattr(
        "app.features.reports.config_store.load_report_config",
        lambda *_args, **_kwargs: {"sections": sections_payload},
    )

    base = tmp_path / "src"
    base.mkdir()
    headers = list(COMPREHENSIVE_COLUMN_LABELS.values())
    # Only create sections 11–13 (simulate Report 10 extraction failure).
    present = [
        "report11_security",
        "report12_punctuality",
        "report13_electrical",
    ]
    for section_id in present:
        path = base / f"{section_id}.csv"
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers)
            writer.writeheader()
            writer.writerow(
                {
                    "S.No.": "1",
                    "Division": "HYB",
                    "Opening Balance": "9",
                    "Received": "5",
                    "% Share": "100.00",
                    "Closed": "5",
                    "Closing Balance": "0",
                    "% Disposal": "100.00",
                    "Avg. Disposal Time": "1.0",
                    "Avg. Rating": "4.0",
                    "Avg. Pendency Time": "0.0",
                }
            )
            writer.writerow(
                {
                    "S.No.": "",
                    "Division": "Total",
                    "Opening Balance": "9",
                    "Received": "5",
                    "% Share": "100.00",
                    "Closed": "5",
                    "Closing Balance": "0",
                    "% Disposal": "100.00",
                    "Avg. Disposal Time": "1.0",
                    "Avg. Rating": "4.0",
                    "Avg. Pendency Time": "0.0",
                }
            )

    index = base / "comprehensive_combined_index.csv"
    with index.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["section_id", "section_name", "csv_path", "row_count", "status", "error"],
        )
        writer.writeheader()
        writer.writerow(
            {
                "section_id": "report10_cw",
                "section_name": "report10_cw",
                "csv_path": "",
                "row_count": "0",
                "status": "failed",
                "error": "submit timeout",
            }
        )
        for section_id in present:
            writer.writerow(
                {
                    "section_id": section_id,
                    "section_name": section_id,
                    "csv_path": str(base / f"{section_id}.csv"),
                    "row_count": "1",
                    "status": "success",
                    "error": "",
                }
            )

    result = Comprehensive1013Processor().process(
        source_a_path=index,
        report_slug="comprehensive-10-13",
        column_selection={
            "date_from": "2026-08-02",
            "date_to": "2026-08-02",
            "run_id": "partial-extraction-run",
        },
    )
    assert result.success, result.error
    assert "opening_balance" not in (result.selected_column_ids or [])
    assert result.artifact_metadata["rendered_section_ids"] == present

    pdf_text = "\n".join(
        (page.extract_text() or "") for page in PdfReader(result.pdf_path).pages
    )
    assert "Opening Balance" not in pdf_text
    assert "Security complaints" in pdf_text
    assert "C&W complaints" not in pdf_text

    wb = load_workbook(result.excel_path)
    ws = wb.active
    header_row = [ws.cell(3, c).value for c in range(1, 12)]
    assert "Opening Balance" not in header_row


def test_process_independent_avg_rating_per_section(tmp_path, monkeypatch):
    import csv

    from openpyxl import load_workbook

    from app.automation.processing.comprehensive_output_columns import (
        COMPREHENSIVE_COLUMN_IDS,
        COMPREHENSIVE_COLUMN_LABELS,
    )

    monkeypatch.setattr(
        "app.automation.processing.comprehensive1013_processor.config.output_excel_dir",
        str(tmp_path / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.comprehensive1013_processor.config.output_pdf_dir",
        str(tmp_path / "pdf"),
    )

    all_ids = list(COMPREHENSIVE_COLUMN_IDS)
    without_rating = [c for c in all_ids if c != "avg_rating"]
    sections = {
        sid: {"selected_column_ids": list(all_ids)} for sid in COMPREHENSIVE_1013_SECTION_IDS
    }
    sections["report11_security"] = {"selected_column_ids": list(without_rating)}

    base = tmp_path / "src"
    base.mkdir()
    headers = list(COMPREHENSIVE_COLUMN_LABELS.values())
    for section_id in COMPREHENSIVE_1013_SECTION_IDS:
        path = base / f"{section_id}.csv"
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers)
            writer.writeheader()
            writer.writerow(
                {
                    "S.No.": "1",
                    "Division": "HYB",
                    "Opening Balance": "0",
                    "Received": "5",
                    "% Share": "100.00",
                    "Closed": "5",
                    "Closing Balance": "0",
                    "% Disposal": "100.00",
                    "Avg. Disposal Time": "1.0",
                    "Avg. Rating": "4.0",
                    "Avg. Pendency Time": "0.0",
                }
            )
            writer.writerow(
                {
                    "S.No.": "",
                    "Division": "Total",
                    "Opening Balance": "0",
                    "Received": "5",
                    "% Share": "100.00",
                    "Closed": "5",
                    "Closing Balance": "0",
                    "% Disposal": "100.00",
                    "Avg. Disposal Time": "1.0",
                    "Avg. Rating": "4.0",
                    "Avg. Pendency Time": "0.0",
                }
            )

    index = base / "comprehensive_combined_index.csv"
    with index.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["section_id", "section_name", "csv_path", "row_count", "status", "error"],
        )
        writer.writeheader()
        for section_id in COMPREHENSIVE_1013_SECTION_IDS:
            writer.writerow(
                {
                    "section_id": section_id,
                    "section_name": section_id,
                    "csv_path": str(base / f"{section_id}.csv"),
                    "row_count": "1",
                    "status": "success",
                    "error": "",
                }
            )

    result = Comprehensive1013Processor().process(
        source_a_path=index,
        report_slug="comprehensive-10-13",
        column_selection={
            "sections": sections,
            "date_from": "2026-08-02",
            "date_to": "2026-08-02",
            "run_id": "independence-test-run",
        },
    )
    assert result.success, result.error

    wb = load_workbook(result.excel_path)
    ws = wb.active
    report10_headers = [ws.cell(3, c).value for c in range(1, 13)]
    assert "Avg. Rating" in report10_headers

    report11_header_row = None
    for row in range(1, 40):
        vals = [ws.cell(row, c).value for c in range(1, 13)]
        if vals[0] == "S.No." and row > 3:
            report11_header_row = vals
            break
    assert report11_header_row is not None
    assert "Avg. Rating" not in report11_header_row
