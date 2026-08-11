"""Unit tests for Report 9 (Train/Station Cause Wise Grievances)."""

from __future__ import annotations

import csv
from pathlib import Path

import pytest

from app.automation.handlers.registry import HANDLER_REGISTRY, get_handler
from app.automation.processing.registry import PROCESSORS
from app.automation.processing.report9_processor import Report9Processor
from app.automation.report_keys import CANONICAL_KEYS, is_supported_report_key
from app.automation.report9_filters import (
    SECTION_ORDER,
    SOURCE_A_CONFIGS,
    SOURCE_B_CONFIGS,
    ZONE_ALL,
    ZONE_SCR,
    filters_for_zone,
)
from app.automation.reports import (
    DEFAULT_CATALOG,
    REPORT_9,
    REPORT_10_13_COMPREHENSIVE,
    catalog,
)
from app.features.datasets.service import SUPPORTED_REPORT_IDS
from app.features.reports.slug_map import MANUAL_REPORT_SLUGS, PAGE_ID_TO_SLUG, resolve_manual_slug


class TestReport9Registration:
    def test_report9_in_canonical_keys(self):
        assert "report9" in CANONICAL_KEYS
        assert is_supported_report_key("report9")

    def test_report9_in_default_catalog_once(self):
        slugs = [r.slug for r in DEFAULT_CATALOG]
        assert slugs.count("report9") == 1

    def test_report9_above_comprehensive(self):
        slugs = [r.slug for r in DEFAULT_CATALOG]
        assert slugs.index("report9") < slugs.index("comprehensive-10-13")
        assert DEFAULT_CATALOG[-4] == REPORT_9
        assert DEFAULT_CATALOG[-3] == REPORT_10_13_COMPREHENSIVE

    def test_catalog_instance_order(self):
        slugs = [r.slug for r in catalog.reports]
        assert slugs == [
            "report1",
            "division",
            "train-no",
            "types",
            "scr-train",
            "scr-station",
            "report9",
            "comprehensive-10-13",
            "report14",
            "report18",
        ]

    def test_handler_registered(self):
        assert "report9" in HANDLER_REGISTRY
        handler = get_handler("report9")
        assert handler.__class__.__name__ == "Report9Handler"

    def test_processor_registered(self):
        assert "report9" in PROCESSORS
        assert isinstance(PROCESSORS["report9"], Report9Processor)

    def test_manual_slug_and_page_id(self):
        assert "report9" in MANUAL_REPORT_SLUGS
        assert PAGE_ID_TO_SLUG["report9"] == "report9"
        assert resolve_manual_slug("report9") == "report9"

    def test_supported_dataset_id(self):
        assert "report9" in SUPPORTED_REPORT_IDS

    def test_portal_definition(self):
        assert REPORT_9.page_path == "/mis_reports/report7"
        assert REPORT_9.url_fragment == "mis_reports/report7"
        assert REPORT_9.name == "All Zones Train/Station Cause Wise on Date"


class TestReport9Filters:
    def test_source_a_zone_all(self):
        assert all(cfg.zone == ZONE_ALL for cfg in SOURCE_A_CONFIGS)

    def test_source_b_zone_scr(self):
        assert all(cfg.zone == ZONE_SCR for cfg in SOURCE_B_CONFIGS)

    def test_filters_for_zone(self):
        all_filters = filters_for_zone(ZONE_ALL)
        scr_filters = filters_for_zone(ZONE_SCR)
        assert all_filters[0].value == ZONE_ALL
        assert scr_filters[0].value == ZONE_SCR
        assert all_filters[0].selector == "#complaintZoneInput"

    def test_section_order(self):
        ids = [cfg.source_id for cfg in SECTION_ORDER]
        assert ids == [
            "all_zone_train",
            "all_zone_station",
            "scr_train",
            "scr_station",
        ]


class TestReport9Processor:
    def _write_section_csv(self, path: Path, rows: list[tuple[str, str, str]]) -> None:
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(["S.No.", "Cause", "Received", "% Share"])
            for idx, (cause, received, share) in enumerate(rows, start=1):
                writer.writerow([str(idx), cause, received, share])
            total = sum(int(r[1]) for r in rows)
            writer.writerow(["", "Total", str(total), "100.00"])

    def _write_index(self, base: Path, sources: dict[str, Path]) -> Path:
        index_path = base / "report9_combined_index.csv"
        with index_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(
                [
                    "source_id",
                    "section_title",
                    "zone",
                    "csv_path",
                    "row_count",
                    "status",
                    "error",
                    "heading",
                ]
            )
            for cfg in SECTION_ORDER:
                path = sources[cfg.source_id]
                writer.writerow(
                    [
                        cfg.source_id,
                        cfg.section_title,
                        cfg.zone,
                        str(path),
                        2,
                        "success",
                        "",
                        cfg.section_title,
                    ]
                )
        return index_path

    def test_processor_sorts_and_stacks(self, tmp_path: Path):
        sources = {}
        samples = {
            "all_zone_train": [("B Cause", "10", "40.00"), ("A Cause", "15", "60.00")],
            "all_zone_station": [("Y", "3", "30.00"), ("X", "7", "70.00")],
            "scr_train": [("Low", "1", "25.00"), ("High", "3", "75.00")],
            "scr_station": [("S2", "2", "40.00"), ("S1", "3", "60.00")],
        }
        for source_id, rows in samples.items():
            path = tmp_path / f"{source_id}.csv"
            self._write_section_csv(path, rows)
            sources[source_id] = path

        index_path = self._write_index(tmp_path, sources)
        processor = Report9Processor()
        result = processor.process(
            source_a_path=index_path,
            report_slug="report9",
            column_selection={
                "run_id": "test-run",
                "date_from": "2026-07-28",
                "date_to": "2026-07-28",
            },
        )

        assert result.success is True
        assert result.excel_path and Path(result.excel_path).is_file()
        assert result.pdf_path and Path(result.pdf_path).is_file()
        assert "test-run" in result.excel_path
        assert "test-run" in result.pdf_path

        from openpyxl import load_workbook

        workbook = load_workbook(result.excel_path)
        ws = workbook.active
        main_title = ws.cell(row=1, column=1).value
        assert "Report 9: All Zones Train/Station Cause Wise on Date" in str(main_title)
        assert "28-07-2026" in str(main_title)
        assert "Rail Madad Train Cause Wise Grievances" in str(ws.cell(row=2, column=1).value)
        assert ws.cell(row=2, column=2).value is None

        # Verify sort + total via re-load of first section CSV processing path
        sections, _ = processor._load_sections(index_path)
        assert [s.config.source_id for s in sections] == [
            "all_zone_train",
            "all_zone_station",
            "scr_train",
            "scr_station",
        ]
        train = sections[0]
        assert train.headers == ["S.No.", "Cause", "Received", "% Share"]
        assert train.rows[0][1] == "A Cause"
        assert train.rows[0][2] == "15"
        assert train.rows[1][1] == "B Cause"
        assert train.total_row == ["", "Total", "25", "100.00"]

    def test_processor_computes_share_from_cause_received_only(self, tmp_path: Path):
        sources: dict[str, Path] = {}
        for cfg in SECTION_ORDER:
            p = tmp_path / cfg.filename
            with p.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(["Cause", "Received"])
                writer.writerow(["B", "25"])
                writer.writerow(["A", "75"])
            sources[cfg.source_id] = p
        index_path = self._write_index(tmp_path, sources)
        sections, _ = Report9Processor()._load_sections(index_path)
        assert sections[0].rows[0] == ["1", "A", "75", "75.00"]
        assert sections[0].rows[1] == ["2", "B", "25", "25.00"]
        assert sections[0].total_row == ["", "Total", "100", "100.00"]

    def test_missing_section_fails(self, tmp_path: Path):
        # Only one section present
        path = tmp_path / "all_zone_train.csv"
        self._write_section_csv(path, [("A", "5", "100.00")])
        index_path = tmp_path / "report9_combined_index.csv"
        with index_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(
                [
                    "source_id",
                    "section_title",
                    "zone",
                    "csv_path",
                    "row_count",
                    "status",
                    "error",
                    "heading",
                ]
            )
            writer.writerow(
                [
                    "all_zone_train",
                    "Train",
                    "ALL",
                    str(path),
                    1,
                    "success",
                    "",
                    "",
                ]
            )
            for cfg in SECTION_ORDER[1:]:
                writer.writerow(
                    [
                        cfg.source_id,
                        cfg.section_title,
                        cfg.zone,
                        "",
                        0,
                        "failed",
                        cfg.missing_error,
                        "",
                    ]
                )

        result = Report9Processor().process(
            source_a_path=index_path,
            report_slug="report9",
            column_selection={"run_id": "missing-run"},
        )
        assert result.success is False
        assert "REPORT9" in (result.error or "")


class TestReport9TableMatching:
    def test_match_ignores_non_cause_tables(self):
        from app.automation.handlers.report9_handler import Report9Handler
        from app.automation.report9_filters import SOURCE_A_TRAIN

        handler = Report9Handler()
        tables = [
            {
                "heading": "Some Pie Chart",
                "tableId": "",
                "headers": ["Label", "Value"],
                "rows": [["A", "1"], ["B", "2"]],
            },
            {
                "heading": "7.1) Train Complaints Cause Wise to Zone:ALL Division:ALL",
                "tableId": "tabled1",
                "headers": ["Cause", "Received"],
                "rows": [
                    ["Cause", "Received"],
                    ["A", "10"],
                ],
            },
        ]
        matched = handler._match_table(tables, SOURCE_A_TRAIN, zone=ZONE_ALL)
        assert matched is not None
        assert "Train" in matched["heading"] or matched.get("tableId") == "tabled1"
