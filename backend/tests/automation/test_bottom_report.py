"""Unit tests for Bottom Performed Trains Report."""

from __future__ import annotations

import json
from pathlib import Path

from app.automation.bottom_report_detail_extract import (
    parse_summary_table_from_csv,
    scan_division_summary_rows,
)
from app.automation.bottom_report_divisions import division_code_from_name
from app.automation.bottom_report_train_frequency import (
    aggregate_division_trains,
    aggregate_train_frequencies,
    extract_train_number,
)
from app.automation.processing.bottom_report_models import (
    DIVISION_RECEIVED_THRESHOLD,
    MSG_NO_QUALIFYING_DIVISION,
    MSG_NO_TRAIN_GE2,
    MSG_NO_VALID_TRAINS,
    TRAIN_INCLUSION_THRESHOLD,
    BottomReportResult,
    DivisionResult,
    QualifyingTrain,
    SectionResult,
    apply_train_inclusion_filter,
    footer_message,
    resolve_no_train_message,
    section_subheading,
)
from app.automation.processing.bottom_report_processor import BottomReportProcessor
from app.automation.report18_detail_extract import dedupe_portal_rows_by_ref
from openpyxl import load_workbook


class TestDivisionThreshold:
    def test_no_qualifying_division(self):
        headers = ["Division", "Received"]
        rows = [
            ["SECUNDERABAD DIVISION", "15"],
            ["HYDERABAD DIVISION", "3"],
            ["NANDED DIVISION", "2"],
        ]
        qualifying = parse_summary_table_from_csv(headers, rows)
        assert qualifying == []
        section = SectionResult(
            section_id="security",
            no_division_message=MSG_NO_QUALIFYING_DIVISION,
        )
        assert section.no_division_message == "No Div. has figured with more than 20 complaints"

    def test_exactly_20_does_not_qualify(self):
        """Received == 20 must NOT drill (strictly more than 20)."""
        headers = ["Division", "Received"]
        rows = [["SECUNDERABAD DIVISION", "20"]]
        qualifying = parse_summary_table_from_csv(headers, rows)
        assert qualifying == []

    def test_nineteen_does_not_qualify(self):
        headers = ["Division", "Received"]
        rows = [["SECUNDERABAD DIVISION", "19"]]
        assert parse_summary_table_from_csv(headers, rows) == []

    def test_twenty_one_qualifies(self):
        headers = ["Division", "Received"]
        rows = [["SECUNDERABAD DIVISION", "21"]]
        qualifying = parse_summary_table_from_csv(headers, rows)
        assert len(qualifying) == 1
        assert qualifying[0].received == 21

    def test_threshold_case_sc20_hyb21_ned22(self):
        """§24: SC=20 skip, HYB=21 and NED=22 process."""
        headers = ["Division", "Received"]
        rows = [
            ["SECUNDERABAD DIVISION", "20"],
            ["HYDERABAD DIVISION", "21"],
            ["NANDED DIVISION", "22"],
        ]
        qualifying = parse_summary_table_from_csv(headers, rows)
        assert [q.division_code for q in qualifying] == ["HYB", "NED"]
        assert [q.received for q in qualifying] == [21, 22]

    def test_multiple_qualifying_divisions(self):
        headers = ["Division", "Received"]
        rows = [
            ["SECUNDERABAD DIVISION", "27"],
            ["HYDERABAD DIVISION", "24"],
            ["NANDED DIVISION", "21"],
            ["GUNTAKAL DIVISION", "20"],  # exactly 20 must be excluded
        ]
        qualifying = parse_summary_table_from_csv(headers, rows)
        assert len(qualifying) == 3
        codes = {q.division_code for q in qualifying}
        assert codes == {"SC", "HYB", "NED"}
        # Must retain later divisions — not first/max only.
        assert [q.division_code for q in qualifying] == ["SC", "HYB", "NED"]


class TestTrainInclusionFilter:
    def _rows(self, train_counts: dict[str, int]) -> list[dict[str, str]]:
        rows = []
        for train_no, count in train_counts.items():
            for _ in range(count):
                rows.append(
                    {
                        "Ref. No.": f"REF-{train_no}-{len(rows)}",
                        "Train/Station": train_no,
                        "Mode": "Train",
                        "Train Name For Report": f"NAME-{train_no}",
                        "Owning Zone": "SC",
                    }
                )
        return rows

    def test_test1_excludes_single_complaint_trains(self):
        rows = self._rows({"12722": 5, "67759": 4, "17009": 2, "20629": 1})
        agg = aggregate_division_trains(rows, division_code="SC")
        filtered = apply_train_inclusion_filter(agg.trains)
        numbers = {t.train_no: t.complaint_count for t in filtered}
        assert numbers == {"12722": 5, "67759": 4, "17009": 2}
        assert "20629" not in numbers

    def test_test2_sc_division_only_keeps_ge2(self):
        rows = self._rows({"20819": 2, "11019": 1, "12252": 1})
        agg = aggregate_division_trains(rows, division_code="SC")
        filtered = apply_train_inclusion_filter(agg.trains)
        assert len(filtered) == 1
        assert filtered[0].train_no == "20819"
        assert filtered[0].complaint_count == 2
        assert resolve_no_train_message(agg) is None
        assert footer_message(20, "SC") == "Total 20 complaints figured in SC division."

    def test_test5_punctuality_excludes_count_one(self):
        rows = self._rows(
            {"67759": 5, "17009": 5, "67764": 4, "67763": 3, "12722": 1}
        )
        filtered = apply_train_inclusion_filter(aggregate_division_trains(rows).trains)
        numbers = [t.train_no for t in filtered]
        assert numbers == ["17009", "67759", "67764", "67763"]
        assert "12722" not in numbers

    def test_qualifying_division_all_trains_count_one(self):
        rows = self._rows({"12722": 1, "17009": 1, "67764": 1})
        agg = aggregate_division_trains(rows, division_code="SC")
        filtered = apply_train_inclusion_filter(agg.trains)
        assert filtered == []
        assert resolve_no_train_message(agg) == MSG_NO_TRAIN_GE2

    def test_no_valid_train_numbers(self):
        rows = [{"Ref. No.": "R1", "Train/Station": "STATION-ONLY", "Mode": "Station"}]
        agg = aggregate_division_trains(rows, division_code="SC")
        assert apply_train_inclusion_filter(agg.trains) == []
        assert resolve_no_train_message(agg) == MSG_NO_VALID_TRAINS

    def test_sort_order(self):
        rows = self._rows({"18503": 2, "12722": 7, "20629": 3})
        filtered = apply_train_inclusion_filter(aggregate_division_trains(rows).trains)
        assert [t.train_no for t in filtered] == ["12722", "20629", "18503"]

    def test_same_train_separate_divisions(self):
        sc_rows = self._rows({"12722": 4})
        hyb_rows = self._rows({"12722": 3})
        sc_filtered = apply_train_inclusion_filter(
            aggregate_division_trains(sc_rows, division_code="SC").trains
        )
        hyb_filtered = apply_train_inclusion_filter(
            aggregate_division_trains(hyb_rows, division_code="HYB").trains
        )
        assert sc_filtered[0].complaint_count == 4
        assert hyb_filtered[0].complaint_count == 3

    def test_leading_zero_train_preserved(self):
        assert extract_train_number("06509", mode="Train") == "06509"
        rows = self._rows({"06509": 4})
        filtered = apply_train_inclusion_filter(aggregate_train_frequencies(rows))
        assert filtered[0].train_no == "06509"

    def test_threshold_constant(self):
        assert TRAIN_INCLUSION_THRESHOLD == 2
        assert DIVISION_RECEIVED_THRESHOLD == 20
        from app.automation.processing.bottom_report_models import (
            division_meets_received_threshold,
        )

        assert not division_meets_received_threshold(19)
        assert not division_meets_received_threshold(20)
        assert division_meets_received_threshold(21)
        assert division_meets_received_threshold(25)
        assert division_meets_received_threshold(40)


class TestPaginationDedupe:
    def test_dedupe_27_entries(self):
        rows = []
        for i in range(1, 28):
            rows.append({"Ref. No.": f"RM{i:05d}", "Train/Station": "12722", "Mode": "Train"})
        rows.append({"Ref. No.": "RM00001", "Train/Station": "12722", "Mode": "Train"})
        unique = dedupe_portal_rows_by_ref(rows)
        assert len(unique) == 27


class TestDivisionCodeMapping:
    def test_codes(self):
        assert division_code_from_name("SECUNDERABAD DIVISION") == "SC"
        assert division_code_from_name("HYDERABAD DIVISION") == "HYB"
        assert division_code_from_name("NANDED DIVISION") == "NED"


class TestProcessor:
    def _build_result(self) -> BottomReportResult:
        return BottomReportResult(
            report_slug="bottom-report",
            date_from="2026-08-10",
            date_to="2026-08-10",
            sections={
                "water_availability": SectionResult(
                    section_id="water_availability",
                    no_division_message=MSG_NO_QUALIFYING_DIVISION,
                ),
                "electrical_equipment": SectionResult(
                    section_id="electrical_equipment",
                    no_division_message=MSG_NO_QUALIFYING_DIVISION,
                ),
                "security": SectionResult(
                    section_id="security",
                    qualifying_divisions=[
                        DivisionResult(
                            division_name="SECUNDERABAD DIVISION",
                            division_code="SC",
                            division_received=27,
                            qualifying_trains=[
                                QualifyingTrain(
                                    "20819",
                                    "PURI-OKHA WEEKLY EXP. [SUPERFAST]",
                                    2,
                                    "EO",
                                ),
                            ],
                        ),
                        DivisionResult(
                            division_name="HYDERABAD DIVISION",
                            division_code="HYB",
                            division_received=24,
                            qualifying_trains=[
                                QualifyingTrain("12722", "NZM-HYB EXP", 3, "SC"),
                            ],
                        ),
                    ],
                ),
                "punctuality": SectionResult(
                    section_id="punctuality",
                    qualifying_divisions=[
                        DivisionResult(
                            division_name="SECUNDERABAD DIVISION",
                            division_code="SC",
                            division_received=29,
                            qualifying_trains=[
                                QualifyingTrain("67759", "SC-CT MEMU [MEMU]", 5, "SC"),
                                QualifyingTrain("17009", "BIDR-HYB [MAIL EXPRESS]", 5, "SC"),
                                QualifyingTrain("67764", "KZJ-SC MEMU [MEMU]", 4, "SC"),
                                QualifyingTrain("67763", "SC-KZJ MEMU [MEMU]", 3, "SC"),
                            ],
                        ),
                    ],
                ),
            },
        )

    def test_pdf_excel_from_result_json(self, tmp_path: Path):
        result = self._build_result()
        json_path = tmp_path / "result.json"
        result.save(json_path)

        processor = BottomReportProcessor()
        out = processor.process(
            source_a_path=json_path,
            report_slug="bottom-report",
            column_selection={
                "date_from": "2026-08-10",
                "date_to": "2026-08-10",
                "run_id": "test-run",
            },
        )
        assert out.success, out.error
        assert out.excel_path and Path(out.excel_path).is_file()
        assert out.pdf_path and Path(out.pdf_path).is_file()

        wb = load_workbook(out.excel_path)
        ws = wb.active
        flat = [str(cell.value or "") for row in ws.iter_rows() for cell in row if cell.value]

        assert "Bottom performed trains based on comprehensive drop down" in flat
        assert "(territorial basis) as on 10.08.2026" in flat
        assert "Water availability" in flat
        assert "SC DIVISION" not in flat
        assert section_subheading("security", "SC") in flat
        assert section_subheading("security", "HYB") in flat
        assert MSG_NO_QUALIFYING_DIVISION in flat
        assert "20819" in flat
        assert "Total 27 complaints figured in SC division." in flat
        assert "Total 24 complaints figured in HYB division." in flat
        assert "67759" in flat
        assert "67763" in flat
        assert all(str(v) != "1" or "1)" not in str(v) for v in flat if v == "1")

    def test_test4_electrical_no_division_table_structure(self, tmp_path: Path):
        result = BottomReportResult(
            report_slug="bottom-report",
            date_from="2026-08-10",
            date_to="2026-08-10",
            sections={
                "water_availability": SectionResult(
                    section_id="water_availability",
                    no_division_message=MSG_NO_QUALIFYING_DIVISION,
                ),
                "electrical_equipment": SectionResult(
                    section_id="electrical_equipment",
                    no_division_message=MSG_NO_QUALIFYING_DIVISION,
                ),
                "security": SectionResult(
                    section_id="security",
                    no_division_message=MSG_NO_QUALIFYING_DIVISION,
                ),
                "punctuality": SectionResult(
                    section_id="punctuality",
                    no_division_message=MSG_NO_QUALIFYING_DIVISION,
                ),
            },
        )
        json_path = tmp_path / "result.json"
        result.save(json_path)
        out = BottomReportProcessor().process(
            source_a_path=json_path,
            report_slug="bottom-report",
            column_selection={"date_from": "2026-08-10", "date_to": "2026-08-10"},
        )
        assert out.success, out.error
        wb = load_workbook(out.excel_path)
        ws = wb.active
        values = [str(cell.value or "") for cell in ws["A"] if cell.value]
        assert section_subheading("electrical_equipment", None) in values
        assert section_subheading("security", None) in values
        assert section_subheading("punctuality", None) in values
        assert section_subheading("security", None) == "Bottom performed Trains w.r.to Security"
        assert section_subheading("punctuality", None) == "Bottom performed Trains w.r.to Punctuality"
        assert MSG_NO_QUALIFYING_DIVISION in values


class TestWaterPreviousOnly:
    def test_water_section_config(self):
        from app.automation.bottom_report_filters import (
            BOTTOM_COMPREHENSIVE_SECTIONS,
            BOTTOM_WATER_SECTION,
        )
        from app.automation.report14_filters import SOURCE_UPCOMING

        assert BOTTOM_WATER_SECTION.watering_point == "Previous Watering Point"
        assert BOTTOM_WATER_SECTION.watering_point != SOURCE_UPCOMING.watering_point
        assert len(BOTTOM_COMPREHENSIVE_SECTIONS) == 3


class TestScanDivisionRowsJS:
    def test_scan_payload_threshold(self):
        payload = {
            "found": True,
            "divisions": [
                {
                    "divisionName": "SECUNDERABAD DIVISION",
                    "received": 27,
                    "rowIndex": 0,
                    "receivedIdx": 3,
                    "tableIdx": 0,
                    "hasLink": True,
                },
                {
                    "divisionName": "HYDERABAD DIVISION",
                    "received": 20,
                    "rowIndex": 1,
                    "receivedIdx": 3,
                    "tableIdx": 0,
                    "hasLink": True,
                },
                {
                    "divisionName": "NANDED DIVISION",
                    "received": 12,
                    "rowIndex": 2,
                    "receivedIdx": 3,
                    "tableIdx": 0,
                    "hasLink": True,
                },
            ],
        }
        qualifying = scan_division_summary_rows(payload, threshold=DIVISION_RECEIVED_THRESHOLD)
        assert len(qualifying) == 1
        assert qualifying[0].received == 27

    def test_scan_keeps_all_qualifying_not_just_first(self):
        """§2 / §25: SC=27 and HYB=24 both qualify; NED=12 does not."""
        from app.automation.bottom_report_detail_extract import rematch_division_from_scan

        payload = {
            "found": True,
            "divisions": [
                {
                    "divisionName": "SECUNDERABAD DIVISION",
                    "received": 27,
                    "rowIndex": 0,
                    "receivedIdx": 3,
                    "tableIdx": 0,
                    "hasLink": True,
                },
                {
                    "divisionName": "HYDERABAD DIVISION",
                    "received": 24,
                    "rowIndex": 1,
                    "receivedIdx": 3,
                    "tableIdx": 0,
                    "hasLink": True,
                },
                {
                    "divisionName": "NANDED DIVISION",
                    "received": 12,
                    "rowIndex": 2,
                    "receivedIdx": 3,
                    "tableIdx": 0,
                    "hasLink": True,
                },
            ],
        }
        qualifying = scan_division_summary_rows(payload)
        assert [q.division_code for q in qualifying] == ["SC", "HYB"]

        # Rematch after a simulated row-index shift still finds HYB.
        shifted = {
            "found": True,
            "divisions": [
                {
                    "divisionName": "SECUNDERABAD DIVISION",
                    "received": 27,
                    "rowIndex": 5,
                    "receivedIdx": 3,
                    "tableIdx": 0,
                    "hasLink": True,
                },
                {
                    "divisionName": "HYDERABAD DIVISION",
                    "received": 24,
                    "rowIndex": 6,
                    "receivedIdx": 3,
                    "tableIdx": 0,
                    "hasLink": True,
                },
            ],
        }
        rematched = rematch_division_from_scan(shifted, qualifying[1])
        assert rematched is not None
        assert rematched.division_code == "HYB"
        assert rematched.row_index == 6


class TestMultiDivisionOutput:
    def _pair_divs(self, sc_received: int, hyb_received: int) -> list[DivisionResult]:
        return [
            DivisionResult(
                division_name="SECUNDERABAD DIVISION",
                division_code="SC",
                division_received=sc_received,
                qualifying_trains=[
                    QualifyingTrain("12722", "NZM-HYB EXP", 4, "SC"),
                ],
            ),
            DivisionResult(
                division_name="HYDERABAD DIVISION",
                division_code="HYB",
                division_received=hyb_received,
                qualifying_trains=[
                    QualifyingTrain("12722", "NZM-HYB EXP", 3, "SC"),
                ],
            ),
        ]

    def test_all_four_types_separate_division_tables(self, tmp_path: Path):
        """§3–§4 / §26: each type renders separate SC + HYB tables, not merged."""
        result = BottomReportResult(
            report_slug="bottom-report",
            date_from="2026-08-10",
            date_to="2026-08-10",
            sections={
                "water_availability": SectionResult(
                    section_id="water_availability",
                    qualifying_divisions=self._pair_divs(25, 22),
                ),
                "electrical_equipment": SectionResult(
                    section_id="electrical_equipment",
                    qualifying_divisions=self._pair_divs(30, 21),
                ),
                "security": SectionResult(
                    section_id="security",
                    qualifying_divisions=self._pair_divs(27, 24),
                ),
                "punctuality": SectionResult(
                    section_id="punctuality",
                    qualifying_divisions=self._pair_divs(35, 28),
                ),
            },
        )
        json_path = tmp_path / "result.json"
        result.save(json_path)
        out = BottomReportProcessor().process(
            source_a_path=json_path,
            report_slug="bottom-report",
            column_selection={"date_from": "2026-08-10", "date_to": "2026-08-10"},
        )
        assert out.success, out.error
        wb = load_workbook(out.excel_path)
        ws = wb.active
        col_a = [str(cell.value or "") for cell in ws["A"] if cell.value]

        # Exact section order.
        order_idx = [
            col_a.index("Water availability"),
            col_a.index("Electrical Equipment"),
            col_a.index("Security"),
            col_a.index("Punctuality"),
        ]
        assert order_idx == sorted(order_idx)

        # Security: separate SC and HYB subtitles + footers (not one merged table).
        assert section_subheading("security", "SC") in col_a
        assert section_subheading("security", "HYB") in col_a
        assert "Total 27 complaints figured in SC division." in col_a
        assert "Total 24 complaints figured in HYB division." in col_a
        assert col_a.count("Security") == 1

        # Punctuality / Electrical: division-specific subtitles.
        assert section_subheading("punctuality", "SC") in col_a
        assert section_subheading("punctuality", "HYB") in col_a
        assert section_subheading("electrical_equipment", "SC") in col_a
        assert section_subheading("electrical_equipment", "HYB") in col_a

        # Water keeps shared subtitle; footers distinguish divisions.
        water_subtitle = section_subheading("water_availability", None)
        assert col_a.count(water_subtitle) == 2
        assert "Total 25 complaints figured in SC division." in col_a
        assert "Total 22 complaints figured in HYB division." in col_a

        # Division-specific counting: same train number kept separate (4 vs 3).
        complaint_vals = [
            cell.value for row in ws.iter_rows() for cell in row if cell.value in (3, 4)
        ]
        assert 4 in complaint_vals and 3 in complaint_vals

        assert Path(out.pdf_path).is_file()

    def test_qualifying_division_no_train_ge2(self, tmp_path: Path):
        """§15: division qualifies but no train >=2 → message + Received footer."""
        result = BottomReportResult(
            report_slug="bottom-report",
            date_from="2026-08-10",
            date_to="2026-08-10",
            sections={
                "water_availability": SectionResult(
                    section_id="water_availability",
                    no_division_message=MSG_NO_QUALIFYING_DIVISION,
                ),
                "electrical_equipment": SectionResult(
                    section_id="electrical_equipment",
                    no_division_message=MSG_NO_QUALIFYING_DIVISION,
                ),
                "security": SectionResult(
                    section_id="security",
                    qualifying_divisions=[
                        DivisionResult(
                            division_name="SECUNDERABAD DIVISION",
                            division_code="SC",
                            division_received=25,
                            qualifying_trains=[],
                            no_train_message=MSG_NO_TRAIN_GE2,
                        ),
                    ],
                ),
                "punctuality": SectionResult(
                    section_id="punctuality",
                    no_division_message=MSG_NO_QUALIFYING_DIVISION,
                ),
            },
        )
        json_path = tmp_path / "result.json"
        result.save(json_path)
        out = BottomReportProcessor().process(
            source_a_path=json_path,
            report_slug="bottom-report",
            column_selection={"date_from": "2026-08-10", "date_to": "2026-08-10"},
        )
        assert out.success, out.error
        wb = load_workbook(out.excel_path)
        flat = [str(c.value or "") for row in wb.active.iter_rows() for c in row if c.value]
        assert section_subheading("security", "SC") in flat
        assert MSG_NO_TRAIN_GE2 in flat
        assert "Total 25 complaints figured in SC division." in flat
        # Must NOT replace with the no-division message for this section table.
        security_idx = flat.index("Security")
        next_section = flat.index("Punctuality")
        security_slice = flat[security_idx:next_section]
        assert MSG_NO_QUALIFYING_DIVISION not in security_slice


class TestResultJsonSerialization:
    def test_save_coerces_date_objects(self, tmp_path: Path):
        from datetime import date

        result = BottomReportResult(
            report_slug="bottom-report",
            date_from=date(2026, 8, 10),  # type: ignore[arg-type]
            date_to=date(2026, 8, 10),  # type: ignore[arg-type]
            sections={
                "security": SectionResult(
                    section_id="security",
                    no_division_message=MSG_NO_QUALIFYING_DIVISION,
                ),
            },
        )
        path = tmp_path / "result.json"
        result.save(path)
        payload = json.loads(path.read_text(encoding="utf-8"))
        assert payload["date_from"] == "2026-08-10"
        assert payload["date_to"] == "2026-08-10"
