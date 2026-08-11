"""Unit tests for merged PDF / Excel download builders."""

from __future__ import annotations

import pytest
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from pypdf import PdfReader
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from app.automation.config import config
from app.automation.merged_downloads import (
    MergedDownloadError,
    build_merged_excel,
    build_merged_pdf,
    cover_report_date_display,
    is_separator_title_page,
    merged_excel_filename,
)
from app.automation.merged_report_catalog import (
    MERGED_REPORT_CATALOG,
    consolidated_download_catalog,
    is_excluded_from_consolidated_download,
)
from app.infrastructure.database.models import AutomationArtifactModel


def _minimal_pdf(path: Path, label: str) -> None:
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    pdf.drawString(72, 800, label)
    pdf.showPage()
    pdf.save()
    path.write_bytes(buffer.getvalue())


def _artifact(slug: str, file_type: str, path: Path) -> AutomationArtifactModel:
    return AutomationArtifactModel(
        id=f"{slug}-{file_type}",
        run_id="run-test",
        report_slug=slug,
        report_name=slug,
        artifact_type=file_type,
        file_path=str(path),
        status="ready",
    )


def test_cover_date_display():
    when = datetime(2026, 8, 4, 6, 0, tzinfo=UTC)
    assert cover_report_date_display(generated_at=when) == "03-08-2026"


def test_separator_page_detection():
    divider = BytesIO()
    pdf = canvas.Canvas(divider, pagesize=A4)
    pdf.drawString(72, 700, "Report 6")
    pdf.drawString(72, 670, "SCR Train Report")
    pdf.drawString(72, 645, "-" * 50)
    pdf.showPage()
    pdf.save()
    reader = PdfReader(BytesIO(divider.getvalue()))
    assert is_separator_title_page(reader.pages[0]) is True

    content = BytesIO()
    pdf2 = canvas.Canvas(content, pagesize=A4)
    pdf2.drawString(72, 800, "Division")
    pdf2.drawString(72, 780, "Received")
    pdf2.drawString(72, 760, "10")
    pdf2.drawString(72, 740, "Total")
    pdf2.showPage()
    pdf2.save()
    reader2 = PdfReader(BytesIO(content.getvalue()))
    assert is_separator_title_page(reader2.pages[0]) is False


def test_merged_catalog_includes_report18_but_consolidated_excludes_it():
    all_slugs = [entry.slug for entry in MERGED_REPORT_CATALOG]
    assert "report18" in all_slugs
    assert all_slugs[-1] == "report18"

    consolidated_slugs = [entry.slug for entry in consolidated_download_catalog()]
    assert "report18" not in consolidated_slugs
    assert is_excluded_from_consolidated_download("report18") is True
    assert is_excluded_from_consolidated_download("report1") is False


def test_build_merged_pdf_cover_toc_and_no_dividers(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "output_pdf_dir", str(tmp_path / "pdf"))
    monkeypatch.setattr(config, "output_excel_dir", str(tmp_path / "excel"))

    artifacts: list[AutomationArtifactModel] = []
    for slug, label, num in (
        ("report1", "Zone Wise content", 1),
        ("division", "Division content", 2),
    ):
        pdf_dir = tmp_path / "pdf" / slug
        pdf_dir.mkdir(parents=True)
        pdf_path = pdf_dir / f"{slug}.pdf"
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=A4)
        pdf.drawString(72, 700, f"Report {num}")
        pdf.drawString(72, 670, label)
        pdf.drawString(72, 645, "-" * 50)
        pdf.showPage()
        pdf.drawString(72, 800, f"{label} table Received 99 Total")
        pdf.showPage()
        pdf.save()
        pdf_path.write_bytes(buffer.getvalue())
        artifacts.append(_artifact(slug, "pdf", pdf_path))

    payload = build_merged_pdf(
        artifacts,
        run_id="run-test",
        generated_at=datetime(2026, 8, 4, 6, 0, tzinfo=UTC),
    )
    reader = PdfReader(BytesIO(payload))
    text = "".join(page.extract_text() or "" for page in reader.pages)

    assert "CRB RM Reports as on 03-08-2026" in text
    assert "TABLE OF CONTENTS" in text
    assert "RailMadad Report Center" not in text
    assert "Run ID" not in text
    assert "Daily Report" not in text
    assert "Generated Time" not in text
    assert "Received 99" in text
    # Cover + TOC + 2 content pages (title pages stripped from each report PDF)
    assert len(reader.pages) == 4


def test_build_merged_pdf_excludes_report18(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "output_pdf_dir", str(tmp_path / "pdf"))

    pdf_dir = tmp_path / "pdf" / "report18"
    pdf_dir.mkdir(parents=True)
    pdf_path = pdf_dir / "report18.pdf"
    _minimal_pdf(pdf_path, "Vande Bharat Received Total")

    with pytest.raises(MergedDownloadError, match="No PDF artifacts"):
        build_merged_pdf(
            [_artifact("report18", "pdf", pdf_path)],
            run_id="run-test",
            generated_at=datetime(2026, 8, 4, 6, 0, tzinfo=UTC),
        )


def test_build_merged_pdf_ignores_report18_when_other_reports_present(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "output_pdf_dir", str(tmp_path / "pdf"))

    artifacts: list[AutomationArtifactModel] = []
    for slug, label in (("report1", "Zone Wise Received Total"), ("report18", "Vande Bharat Received Total")):
        pdf_dir = tmp_path / "pdf" / slug
        pdf_dir.mkdir(parents=True)
        pdf_path = pdf_dir / f"{slug}.pdf"
        _minimal_pdf(pdf_path, label)
        artifacts.append(_artifact(slug, "pdf", pdf_path))

    payload = build_merged_pdf(
        artifacts,
        run_id="run-test",
        generated_at=datetime(2026, 8, 4, 6, 0, tzinfo=UTC),
    )
    reader = PdfReader(BytesIO(payload))
    text = "".join(page.extract_text() or "" for page in reader.pages)
    assert "Zone Wise" in text
    assert "Vande Bharat" not in text
    assert "Report Vande Bharat" not in text


def test_build_merged_excel_sheet_order(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "output_pdf_dir", str(tmp_path / "pdf"))
    monkeypatch.setattr(config, "output_excel_dir", str(tmp_path / "excel"))

    artifacts: list[AutomationArtifactModel] = []
    for slug, title, value in (
        ("report1", "Report 1 - Zone Wise", "zone"),
        ("division", "Report 2 - Division", "division"),
    ):
        excel_dir = tmp_path / "excel" / slug
        excel_dir.mkdir(parents=True)
        excel_path = excel_dir / f"{slug}.xlsx"
        wb = Workbook()
        wb.active["A1"] = value
        wb.save(excel_path)
        artifacts.append(_artifact(slug, "excel", excel_path))

    payload = build_merged_excel(artifacts)
    merged = load_workbook(BytesIO(payload), data_only=True)
    assert merged.sheetnames == ["Report 1 - Zone Wise", "Report 2 - Division"]
    assert merged["Report 1 - Zone Wise"]["A1"].value == "zone"
    assert merged["Report 2 - Division"]["A1"].value == "division"
    merged.close()


def test_build_merged_excel_excludes_report18(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "output_pdf_dir", str(tmp_path / "pdf"))
    monkeypatch.setattr(config, "output_excel_dir", str(tmp_path / "excel"))

    excel_dir = tmp_path / "excel" / "report18"
    excel_dir.mkdir(parents=True)
    excel_path = excel_dir / "report18.xlsx"
    wb = Workbook()
    wb.active["A1"] = "vande"
    wb.save(excel_path)

    with pytest.raises(MergedDownloadError, match="No Excel artifacts"):
        build_merged_excel([_artifact("report18", "excel", excel_path)])


def test_build_merged_excel_ignores_report18_when_other_reports_present(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "output_pdf_dir", str(tmp_path / "pdf"))
    monkeypatch.setattr(config, "output_excel_dir", str(tmp_path / "excel"))

    artifacts: list[AutomationArtifactModel] = []
    for slug, value in (("report1", "zone"), ("report18", "vande")):
        excel_dir = tmp_path / "excel" / slug
        excel_dir.mkdir(parents=True)
        excel_path = excel_dir / f"{slug}.xlsx"
        wb = Workbook()
        wb.active["A1"] = value
        wb.save(excel_path)
        artifacts.append(_artifact(slug, "excel", excel_path))

    payload = build_merged_excel(artifacts)
    merged = load_workbook(BytesIO(payload), data_only=True)
    assert merged.sheetnames == ["Report 1 - Zone Wise"]
    assert "Vande Bharat" not in merged.sheetnames
    assert merged["Report 1 - Zone Wise"]["A1"].value == "zone"
    merged.close()


def test_merged_excel_filename_format():
    when = datetime(2026, 8, 4, 6, 0, tzinfo=UTC)
    assert merged_excel_filename(now=when) == "CRB_RM_Reports_03-08-2026.xlsx"


def test_build_merged_excel_with_merged_cells_and_custom_widths(tmp_path, monkeypatch):
    """Regression test: source sheets with a merged title row + explicit column
    widths (as every real report writer produces) must not crash the merge.

    Guards against two historical bugs:
      1. Setting ColumnDimension.customWidth (read-only in openpyxl >= 3.1).
      2. Autofit reading .column_letter off a MergedCell (raises AttributeError).
    """
    monkeypatch.setattr(config, "output_pdf_dir", str(tmp_path / "pdf"))
    monkeypatch.setattr(config, "output_excel_dir", str(tmp_path / "excel"))

    excel_dir = tmp_path / "excel" / "report1"
    excel_dir.mkdir(parents=True)
    excel_path = excel_dir / "report1.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("A1:C1")
    ws["A1"] = "Merged Title Row"
    ws["A2"] = "Header1"
    ws["B2"] = "Header2"
    ws["C2"] = "Header3"
    ws["A3"] = "Row1"
    ws["B3"] = 10
    ws["C3"] = 20
    ws.column_dimensions["A"].width = 18.0
    ws.column_dimensions["B"].width = 12.0
    ws.column_dimensions["C"].width = 12.0
    wb.save(excel_path)

    artifacts = [_artifact("report1", "excel", excel_path)]
    payload = build_merged_excel(artifacts)
    merged = load_workbook(BytesIO(payload))
    sheet = merged["Report 1 - Zone Wise"]
    assert sheet["A1"].value == "Merged Title Row"
    assert sheet["B3"].value == 10
    merged.close()


def test_build_merged_excel_rejects_empty_worksheet(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "output_pdf_dir", str(tmp_path / "pdf"))
    monkeypatch.setattr(config, "output_excel_dir", str(tmp_path / "excel"))

    excel_dir = tmp_path / "excel" / "report1"
    excel_dir.mkdir(parents=True)
    excel_path = excel_dir / "report1.xlsx"
    Workbook().save(excel_path)

    artifacts = [_artifact("report1", "excel", excel_path)]
    try:
        build_merged_excel(artifacts)
        raise AssertionError("Expected MergedDownloadError for empty worksheet")
    except MergedDownloadError:
        pass
