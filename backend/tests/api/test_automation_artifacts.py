"""Tests for CDP run artifact preview/download APIs."""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from uuid import uuid4

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from sqlalchemy.ext.asyncio import AsyncSession

from app.automation.config import config
from app.automation.dependencies import get_automation_service
from app.automation.run_registry import (
    ArtifactPathError,
    create_cdp_run,
    ensure_cdp_profile,
    register_artifact,
    validate_artifact_file,
)
from app.domain.entities.user import User, UserRole
from app.features.auth.dependencies import (
    require_admin,
    require_officer_or_admin,
    validate_csrf_token,
)
from app.infrastructure.database.session import get_db_session
from app.main import app
from unittest.mock import AsyncMock


@pytest.fixture
def admin_user() -> User:
    now = datetime.now(UTC)
    return User(
        id="test-admin",
        username="admin",
        email="admin@test.local",
        password_hash="hash",
        role=UserRole.ADMIN,
        is_active=True,
        created_at=now,
        updated_at=now,
    )


@pytest.fixture
async def api_client(admin_user: User, test_session: AsyncSession):
    async def override_admin() -> User:
        return admin_user

    def override_csrf() -> None:
        return None

    async def override_db():
        yield test_session

    app.dependency_overrides[get_automation_service] = lambda: AsyncMock()
    app.dependency_overrides[require_admin] = override_admin
    app.dependency_overrides[require_officer_or_admin] = override_admin
    app.dependency_overrides[validate_csrf_token] = override_csrf
    app.dependency_overrides[get_db_session] = override_db

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        yield client

    app.dependency_overrides.clear()


def test_validate_artifact_blocks_traversal(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "output_pdf_dir", str(tmp_path / "pdf"))
    monkeypatch.setattr(config, "output_excel_dir", str(tmp_path / "excel"))
    monkeypatch.setattr(config, "extracted_data_dir", str(tmp_path / "extracted"))
    monkeypatch.setattr(config, "pdf_archive_dir", str(tmp_path / "archive"))
    (tmp_path / "pdf").mkdir()
    evil = Path(tmp_path / "outside" / "secret.pdf")
    evil.parent.mkdir(parents=True)
    evil.write_bytes(b"%PDF-1.4x")
    with pytest.raises(ArtifactPathError):
        validate_artifact_file(evil)


def _write_minimal_pdf(path: Path, title: str) -> None:
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    pdf.setFont("Helvetica", 12)
    pdf.drawString(72, 800, title)
    pdf.showPage()
    pdf.save()
    path.write_bytes(buffer.getvalue())


@pytest.mark.asyncio
async def test_artifact_preview_download_and_merged_outputs(
    tmp_path, monkeypatch, api_client, test_session: AsyncSession
):
    pdf_dir = tmp_path / "pdf" / "division"
    excel_dir = tmp_path / "excel" / "division"
    pdf_dir.mkdir(parents=True)
    excel_dir.mkdir(parents=True)
    pdf_path = pdf_dir / "sample.pdf"
    excel_path = excel_dir / "sample.xlsx"
    _write_minimal_pdf(pdf_path, "Division Report")
    workbook = Workbook()
    workbook.active["A1"] = "sample"
    workbook.save(excel_path)

    monkeypatch.setattr(config, "output_pdf_dir", str(tmp_path / "pdf"))
    monkeypatch.setattr(config, "output_excel_dir", str(tmp_path / "excel"))
    monkeypatch.setattr(config, "extracted_data_dir", str(tmp_path / "extracted"))
    monkeypatch.setattr(config, "pdf_archive_dir", str(tmp_path / "archive"))
    (tmp_path / "extracted").mkdir(exist_ok=True)
    (tmp_path / "archive").mkdir(exist_ok=True)

    await ensure_cdp_profile(test_session)
    run = await create_cdp_run(test_session)
    pdf_art = await register_artifact(
        test_session,
        run_id=run.id,
        report_slug="division",
        report_name="division",
        file_type="pdf",
        file_path=pdf_path,
    )
    excel_art = await register_artifact(
        test_session,
        run_id=run.id,
        report_slug="division",
        report_name="division",
        file_type="excel",
        file_path=excel_path,
    )

    preview = await api_client.get(f"/api/v1/automation/artifacts/{pdf_art.id}/preview")
    assert preview.status_code == 200
    assert "inline" in preview.headers.get("content-disposition", "")
    assert preview.content.startswith(b"%PDF-")

    download = await api_client.get(f"/api/v1/automation/artifacts/{pdf_art.id}/download")
    assert download.status_code == 200
    assert "attachment" in download.headers.get("content-disposition", "")

    excel_dl = await api_client.get(
        f"/api/v1/automation/artifacts/{excel_art.id}/download"
    )
    assert excel_dl.status_code == 200

    run_resp = await api_client.get(f"/api/v1/automation/runs/{run.id}")
    assert run_resp.status_code == 200
    body = run_resp.json()
    assert body["run_id"] == run.id
    assert body["download_pdf_all_url"].endswith("/download/pdf/all")
    assert body["download_excel_all_url"].endswith("/download/excel/all")

    arts = await api_client.get(f"/api/v1/automation/runs/{run.id}/artifacts")
    assert arts.status_code == 200
    assert len(arts.json()) == 2
    assert all("file_path" not in a for a in arts.json())

    merged_pdf = await api_client.get(f"/api/v1/automation/runs/{run.id}/download/pdf/all")
    assert merged_pdf.status_code == 200
    assert merged_pdf.headers["content-type"].startswith("application/pdf")
    assert "RailMadad_Report_" in merged_pdf.headers.get("content-disposition", "")
    pdf_reader = PdfReader(BytesIO(merged_pdf.content))
    merged_text = "".join(page.extract_text() or "" for page in pdf_reader.pages)
    assert "CRB RM Reports as on" in merged_text
    assert "TABLE OF CONTENTS" in merged_text
    assert "Division Report" in merged_text
    # Cover + TOC + report content (no divider pages)
    assert len(pdf_reader.pages) == 3

    merged_excel = await api_client.get(f"/api/v1/automation/runs/{run.id}/download/excel/all")
    assert merged_excel.status_code == 200
    assert merged_excel.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "CRB_RM_Reports_" in merged_excel.headers.get("content-disposition", "")
    merged_wb = load_workbook(BytesIO(merged_excel.content), data_only=True)
    assert "Report 2 - Division" in merged_wb.sheetnames
    assert merged_wb["Report 2 - Division"]["A1"].value == "sample"
    merged_wb.close()


def test_validate_artifact_rejects_invalid_excel(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "output_pdf_dir", str(tmp_path / "pdf"))
    monkeypatch.setattr(config, "output_excel_dir", str(tmp_path / "excel"))
    monkeypatch.setattr(config, "extracted_data_dir", str(tmp_path / "extracted"))
    monkeypatch.setattr(config, "pdf_archive_dir", str(tmp_path / "archive"))
    excel_dir = tmp_path / "excel"
    excel_dir.mkdir()
    broken = excel_dir / "broken.xlsx"
    broken.write_bytes(b"PK\x03\x04not-a-real-xlsx")

    with pytest.raises(ArtifactPathError, match="Invalid Excel|Incomplete Excel|Corrupt Excel"):
        validate_artifact_file(broken, file_type="excel")


@pytest.mark.asyncio
async def test_missing_artifact_returns_404(api_client):
    resp = await api_client.get(f"/api/v1/automation/artifacts/{uuid4()}/download")
    assert resp.status_code == 404


@pytest.mark.asyncio
async def test_run_detail_scrubs_filesystem_paths(
    tmp_path, monkeypatch, api_client, test_session: AsyncSession
):
    from app.automation.schemas import MultiReportResult, ReportResult
    from app.automation.run_registry import finalize_cdp_run

    monkeypatch.setattr(config, "output_pdf_dir", str(tmp_path / "pdf"))
    monkeypatch.setattr(config, "output_excel_dir", str(tmp_path / "excel"))
    monkeypatch.setattr(config, "extracted_data_dir", str(tmp_path / "extracted"))
    monkeypatch.setattr(config, "pdf_archive_dir", str(tmp_path / "archive"))
    for name in ("pdf", "excel", "extracted", "archive"):
        (tmp_path / name).mkdir(exist_ok=True)

    run = await create_cdp_run(test_session)
    result = MultiReportResult(
        success=True,
        connected=True,
        tab_found=True,
        run_id=run.id,
        reports=[
            ReportResult(
                slug="division",
                dataset_key="division",
                status="success",
                excel_path=str(tmp_path / "excel" / "x.xlsx"),
                pdf_path=str(tmp_path / "pdf" / "x.pdf"),
                source_csv_path=str(tmp_path / "extracted" / "x.csv"),
                source_paths=[str(tmp_path / "extracted" / "x.csv")],
                archive_path=str(tmp_path / "archive" / "x.pdf"),
                pdf_download_url="/api/v1/automation/artifacts/a/download",
                pdf_preview_url="/api/v1/automation/artifacts/a/preview",
            )
        ],
    )
    await finalize_cdp_run(test_session, run.id, result)

    resp = await api_client.get(f"/api/v1/automation/runs/{run.id}")
    assert resp.status_code == 200
    body = resp.json()
    report = body["reports"][0]
    assert "excel_path" not in report
    assert "pdf_path" not in report
    assert "source_csv_path" not in report
    assert "source_paths" not in report
    assert "archive_path" not in report
    assert report["pdf_download_url"].endswith("/download")


@pytest.mark.asyncio
async def test_start_accepts_report_slugs_subset(api_client, admin_user):
    mock_service = AsyncMock()
    from app.automation.schemas import MultiReportResult

    mock_service.start = AsyncMock(
        return_value=MultiReportResult(
            success=True,
            connected=True,
            tab_found=True,
            run_id="run-subset",
            reports=[],
        )
    )
    app.dependency_overrides[get_automation_service] = lambda: mock_service

    resp = await api_client.post(
        "/api/v1/automation/start",
        json={"report_slugs": ["division"]},
    )
    assert resp.status_code == 200
    mock_service.start.assert_awaited()
    kwargs = mock_service.start.await_args.kwargs
    assert kwargs.get("report_slugs") == ["division"]
    assert kwargs.get("user_id") == admin_user.id
